# =================== ARCHIVO: colombia.py  ===================
"""
Procesamiento de Consolidado CAPEX para Colombia
Estructura específica del Reporte Pago Programado
"""

from utils import (APIHelper, ExcelProcessor, leer_excel_safe, 
                   validar_columnas_colombia, validar_monedas_colombia, 
                   validar_reporte_absoluto, analizar_estructura_archivo)
import pandas as pd
from pathlib import Path
import os
from typing import Optional, Dict, Any
import tempfile
from openpyxl import Workbook
from openpyxl.styles import PatternFill


try:
    import gspread
    from google.oauth2.service_account import Credentials
    from google.auth import default
    GOOGLE_SHEETS_AVAILABLE = True
except ImportError:
    GOOGLE_SHEETS_AVAILABLE = False

def configurar_google_sheets():
    """Configurar conexión a Google Sheets"""
    
    if not GOOGLE_SHEETS_AVAILABLE:
        print("   ⚠️ Librerías de Google Sheets no disponibles. Instala: pip install gspread google-auth")
        return None
    
    try:
        # Configurar scopes necesarios
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive.file'
        ]
        
        # Buscar archivo de credenciales en diferentes ubicaciones
        posibles_archivos_cred = [
            'credentials.json',
            'google_credentials.json', 
            'service_account.json',
            os.path.join('config', 'credentials.json'),
            os.path.join('keys', 'credentials.json')
        ]
        
        archivo_credenciales = None
        for archivo in posibles_archivos_cred:
            if os.path.exists(archivo):
                archivo_credenciales = archivo
                break
        
        if archivo_credenciales:
            # Usar archivo de credenciales si está disponible (desarrollo local)
            credentials = Credentials.from_service_account_file(archivo_credenciales, scopes=scopes)
            gc = gspread.authorize(credentials)
            print(f"   ✓ Conexión a Google Sheets configurada usando archivo: {archivo_credenciales}")
            return gc
        else:
            # Intentar usar Application Default Credentials (Cloud Run, GCE, etc.)
            try:
                credentials, project = default(scopes=scopes)
                # gspread puede tener problemas con credenciales de Compute Engine
                # Intentar autorizar con las credenciales de ADC
                gc = gspread.authorize(credentials)
                print(f"   ✓ Conexión a Google Sheets configurada usando Application Default Credentials")
                return gc
            except Exception as adc_error:
                print(f"   ⚠️ No se pudo configurar Google Sheets con ADC: {str(adc_error)}")
                print("   → Para usar Google Sheets en Cloud Run, necesitas un archivo de credenciales con clave privada")
                print("   → O configura la cuenta de servicio con permisos de Google Sheets API")
                return None
        
    except Exception as e:
        print(f"   ❌ Error al configurar Google Sheets: {str(e)}")
        return None

def leer_google_sheet_con_configuracion() -> Optional[pd.DataFrame]:
    """Leer Google Sheet usando configuración predefinida para proveedores"""
    
    # CONFIGURACIÓN - Modifica estos valores según tu Google Sheet
    GOOGLE_SHEET_CONFIG = {
        # Opción 1: Usar URL completa
        'sheet_url': 'https://docs.google.com/spreadsheets/d/1bzxDTIamkVrbRIRPZTq4Uoac_el5u4PWQn_2iAHNmxY/edit?gid=0#gid=0',
        
        # Opción 2: Usar solo el ID (alternativa)
        'sheet_id': '1bzxDTIamkVrbRIRPZTq4Uoac_el5u4PWQn_2iAHNmxY',
        
        # Nombre de la hoja (opcional, si no se especifica usa la primera)
        'sheet_name': 'Solicitantes',  # o 'Solicitantes', etc.
        
        # Método a usar ('url' o 'id')
        'method': 'url'  # Cambia a 'id' si prefieres usar solo el ID
    }
    
    print("   🔗 Leyendo Google Sheet con configuración predefinida...")
    
    # Verificar configuración
    if GOOGLE_SHEET_CONFIG['method'] == 'url':
        if 'TU_SHEET_ID' in GOOGLE_SHEET_CONFIG['sheet_url']:
            print("   ⚠️ CONFIGURACIÓN PENDIENTE:")
            print("   → Reemplaza 'TU_SHEET_ID' en sheet_url con el ID real de tu Google Sheet")
            print("   → El ID se encuentra en la URL: https://docs.google.com/spreadsheets/d/[ESTE_ES_EL_ID]/edit")
            return None
        return leer_google_sheet_proyectos_solicitantes(GOOGLE_SHEET_CONFIG['sheet_url'], GOOGLE_SHEET_CONFIG.get('sheet_name'))
    
    else:  # method == 'id'
        if GOOGLE_SHEET_CONFIG['sheet_id'] == 'TU_SHEET_ID_AQUI':
            print("   ⚠️ CONFIGURACIÓN PENDIENTE:")
            print("   → Reemplaza 'TU_SHEET_ID_AQUI' en sheet_id con el ID real de tu Google Sheet")
            return None
        return leer_google_sheet_proyectos_solicitantes(GOOGLE_SHEET_CONFIG['sheet_id'], GOOGLE_SHEET_CONFIG.get('sheet_name'))


def leer_google_sheet_proyectos_solicitantes(sheet_url_or_id: str, sheet_name: str = None) -> Optional[pd.DataFrame]:
    """Leer datos de proveedores desde Google Sheets"""
    
    print("   📊 Conectando a Google Sheets para obtener datos de solicitantes...")
    
    # Configurar conexión
    gc = configurar_google_sheets()
    if not gc:
        return None
    
    try:
        # Abrir spreadsheet
        if 'docs.google.com' in sheet_url_or_id:
            # Es una URL completa
            spreadsheet = gc.open_by_url(sheet_url_or_id)
            print(f"   ✓ Spreadsheet abierto por URL: {spreadsheet.title}")
        else:
            # Es un ID
            spreadsheet = gc.open_by_key(sheet_url_or_id)
            print(f"   ✓ Spreadsheet abierto por ID: {spreadsheet.title}")
        
        # Seleccionar hoja
        if sheet_name:
            try:
                worksheet = spreadsheet.worksheet(sheet_name)
                print(f"   ✓ Hoja seleccionada: {sheet_name}")
            except gspread.WorksheetNotFound:
                print(f"   ⚠️ Hoja '{sheet_name}' no encontrada. Usando primera hoja disponible.")
                worksheet = spreadsheet.sheet1
        else:
            worksheet = spreadsheet.sheet1
            print(f"   ✓ Usando primera hoja: {worksheet.title}")
        
        # Obtener todos los datos
        data = worksheet.get_all_records()
        
        if not data:
            print("   ⚠️ No se encontraron datos en la hoja")
            return None
        
        # Convertir a DataFrame
        df = pd.DataFrame(data)
        
        # Limpiar datos
        df = df.dropna(how='all')  # Eliminar filas completamente vacías
        
        # Limpiar espacios en blanco
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace('', pd.NA)
        
        print(f"   ✓ Datos obtenidos: {len(df)} filas, columnas: {list(df.columns)}")
        
        # Verificar que tenemos las columnas necesarias
        columnas_esperadas = ['SOLICITANTE', 'ÁREA']
        columnas_faltantes = []
        
        for col in columnas_esperadas:
            if col not in df.columns:
                # Buscar columnas similares (case insensitive)
                col_similar = None
                for df_col in df.columns:
                    if col.upper() in df_col.upper():
                        col_similar = df_col
                        break
                
                if col_similar:
                    df = df.rename(columns={col_similar: col})
                    print(f"   → Columna '{col_similar}' renombrada a '{col}'")
                else:
                    columnas_faltantes.append(col)
        
        if columnas_faltantes:
            print(f"   ⚠️ Columnas faltantes: {columnas_faltantes}")
            print(f"   → Columnas disponibles: {list(df.columns)}")
        
        return df
        
    except Exception as e:
        print(f"   ❌ Error al leer Google Sheet: {str(e)}")
        return None

def crear_lookup_solicitantes_areas(df_solicitantes: pd.DataFrame) -> dict:
    """Crear diccionario de lookup para Solicitante -> Área"""
    if df_solicitantes is None or df_solicitantes.empty:
        print("   ⚠️ No hay datos de solicitantes para crear lookup")
        return {}
    
    lookup = {}
    
    for idx, row in df_solicitantes.iterrows():
        try:
            solicitante = str(row['SOLICITANTE']).strip().upper()
            # Buscar AREA o ÁREA
            area = None
            if 'AREA' in row:
                area = str(row['AREA']).strip()
            elif 'ÁREA' in row:
                area = str(row['ÁREA']).strip()
            
            if solicitante and area and solicitante != 'NAN' and area != 'nan':
                lookup[solicitante] = area
                
        except Exception as e:
            print(f"   ⚠️ Error procesando fila {idx}: {e}")
            continue
    
    print(f"   ✅ Lookup creado: {len(lookup)} solicitantes mapeados")
    
    # Mostrar algunas muestras
    if lookup:
        print(f"   💡 Muestras:")
        samples = list(lookup.items())[:3]
        for solicitante, area in samples:
            print(f"      '{solicitante}' → '{area}'")
    
    return lookup

def obtener_area_para_solicitante(solicitante, lookup_solicitantes: dict, proyecto: str = None) -> str:
    """
    Obtener área para solicitante - Implementación de la fórmula Excel:
    =+SI(W13=0;"SERVICIOS";BUSCARV(W13;Solicitantes!$A$1:$B$116;2;FALSO))
    
    EXCEPCIONES:
    - Si proyecto = "A048" → retorna "AUTOPAGO"
    - Si solicitante es de TI y proyecto = "VENE" → retorna "DIR CONSTRUCCIÓN Y PROYECTOS"
    
    Args:
        solicitante: Nombre del solicitante
        lookup_solicitantes: Diccionario con lookup de solicitantes -> áreas
        proyecto: Código del proyecto (opcional)
    
    Returns:
        str: Área correspondiente
    """
    # EXCEPCIÓN 1: Si proyecto es A048, retornar AUTOPAGO directamente
    if proyecto and str(proyecto).strip().upper() == "A048":
        return "AUTOPAGO"
    
    # Si está vacío o es 0, devolver "SERVICIOS"
    if not solicitante or str(solicitante).strip() in ['', '0', 'nan', 'None', 'NAN']:
        return "SERVICIOS"
    
    # Si no hay lookup disponible
    if not lookup_solicitantes:
        return "SIN_GOOGLE_SHEET"
    
    # Limpiar y buscar
    solicitante_clean = str(solicitante).strip().upper()
    proyecto_clean = str(proyecto).strip().upper() if proyecto else ""
    
    # Búsqueda exacta
    area_encontrada = None
    if solicitante_clean in lookup_solicitantes:
        area_encontrada = lookup_solicitantes[solicitante_clean]
    else:
        # Búsqueda parcial por palabras clave (apellidos)
        for sol_ref, area in lookup_solicitantes.items():
            # Buscar por coincidencia parcial
            if solicitante_clean in sol_ref or sol_ref in solicitante_clean:
                area_encontrada = area
                break
            
            # Buscar por apellidos (última palabra de cada nombre)
            palabras_ref = sol_ref.split()
            palabras_buscar = solicitante_clean.split()
            
            if len(palabras_ref) > 0 and len(palabras_buscar) > 0:
                # Comparar último apellido
                if palabras_ref[-1] in palabras_buscar or palabras_buscar[-1] in palabras_ref:
                    area_encontrada = area
                    break
    
    # EXCEPCIÓN 2: Si solicitante es de TI y proyecto es VENE, asignar DIR CONSTRUCCIÓN Y PROYECTOS
    if area_encontrada:
        area_clean = str(area_encontrada).strip().upper()
        # Verificar si el área contiene "TI" (Tecnología de Información)
        if "TI" in area_clean or "TECNOLOGIA" in area_clean or "TECNOLOGÍA" in area_clean or "INFORMACION" in area_clean or "INFORMACIÓN" in area_clean:
            if proyecto_clean == "VENE":
                return "DIR CONSTRUCCIÓN Y PROYECTOS"
    
    # Retornar área encontrada o área no encontrada
    if area_encontrada:
        return area_encontrada
    
    # No encontrado en Google Sheet
    return "AREA_NO_ENCONTRADA"


def diagnosticar_archivo_colombia(archivo):
    """Función específica para diagnosticar archivos de Colombia"""
    print(f"\n🇻🇪 DIAGNÓSTICO ESPECÍFICO Colombia")
    print("=" * 50)
    
    # Análisis completo
    skip_recomendado = analizar_estructura_archivo(archivo)
    
    # Probar lectura con la recomendación
    print(f"\n🧪 PROBANDO LECTURA CON SKIPROWS={skip_recomendado}")
    print("-" * 50)
    
    try:
        df = pd.read_excel(archivo, skiprows=skip_recomendado)
        
        headers_esperados = [
            "Numero de Factura", "Numero de OC", "Tipo Factura", "Nombre Lote",
            "Proveedor", "RIF", "Fecha Documento", "Tienda", "Sucursal",
            "Monto", "Moneda", "Fecha Vencimiento", "Cuenta", "Id Cta",
            "Método de Pago", "Pago Independiente", "Prioridad",
            "Monto CAPEX EXT", "Monto CAPEX ORD", "Monto CADM",
            "Fecha Creación", "Solicitante", "Proveedor Remito"
        ]
        
        # Comparar headers
        coincidencias = 0
        for header_esp in headers_esperados:
            for col_real in df.columns:
                if header_esp.lower() in str(col_real).lower():
                    coincidencias += 1
                    break
        
        porcentaje = (coincidencias / len(headers_esperados)) * 100
        
        print(f"✅ Headers coincidentes: {coincidencias}/{len(headers_esperados)} ({porcentaje:.1f}%)")
        
        if porcentaje >= 80:
            print(f"🎯 ARCHIVO COMPATIBLE - Usar skiprows={skip_recomendado}")
            return skip_recomendado
        else:
            print(f"⚠️ BAJA COMPATIBILIDAD - Revisar estructura manualmente")
            return None
            
    except Exception as e:
        print(f"❌ Error en prueba: {e}")
        return None

def procesar_datos_colombia_especifico(df):
    """Procesamiento específico de datos de Colombia con ajuste de CAPEX según prioridad"""
    print("co Aplicando procesamiento específico de Colombia...")
    
    df_procesado = df.copy()
    
    print(f"🔍 Columnas antes del procesamiento: {len(df_procesado.columns)}")

    
    # 2. Limpiar datos
    filas_antes = len(df_procesado)
    df_procesado = df_procesado.dropna(how='all')
    filas_despues = len(df_procesado)
    
    if filas_antes != filas_despues:
        print(f"🧹 Removidas {filas_antes - filas_despues} filas vacías")
    
    # 3. Validar montos
    col_monto = None
    for col in df_procesado.columns:
        if "monto" in str(col).lower() and len(str(col)) < 20:
            col_monto = col
            break
    
    if col_monto:
        print(f"💰 Procesando columna de monto: '{col_monto}'")
        
        valores_ejemplo = df_procesado[col_monto].head(3).tolist()
        print(f"💰 Valores ejemplo: {valores_ejemplo}")
        
        df_procesado[col_monto] = pd.to_numeric(df_procesado[col_monto], errors='coerce')
        
        montos_validos = df_procesado[col_monto].dropna().count()
        montos_total = len(df_procesado)
        print(f"💰 Montos válidos: {montos_validos} de {montos_total}")
    
    # ===================================================================
    # 4. AJUSTE DE CAPEX SEGÚN PRIORIDAD (NUEVO)
    # ===================================================================
    print(f"\n🔧 AJUSTANDO CAPEX SEGÚN PRIORIDADES...")
    
    # Diccionario de prioridades CAPEX
    prioridades_capex = {
        60: 'ORD',   # BNC USD RETIRO ORD
        70: 'EXT',   # BNC USD RETIRO EXT
        71: 'ORD',   # PANAMARICANO EUR ORD
        72: 'EXT',   # PANAMERICANO EUR EXT
        73: 'ORD',   # PANAMERICANO USD ORD
        74: 'EXT',   # PANAMERICANO USD EXT
        75: 'ORD',   # EXTRANJERO USD ORD
        76: 'EXT',   # EXTRANJERO USD EXT
        77: 'ORD',   # EXTRANJERO EUR ORD
        78: 'ORD',   # PAGOS BS CAPEX ORD
        79: 'EXT',    # PAGOS BS CAPEX EXT
        91: 'ORD',   # POR CONFIRMAR CON ANDREA
    }
    
    # Buscar columnas necesarias
    col_capex_ext = None
    col_capex_ord = None
    col_prio = None
    
    for col in df_procesado.columns:
        col_lower = str(col).lower()
        if 'capex' in col_lower and 'ext' in col_lower:
            col_capex_ext = col
        elif 'capex' in col_lower and 'ord' in col_lower:
            col_capex_ord = col
        elif 'prioridad' in col_lower:
            col_prio = col
    
    print(f"   📋 Monto CAPEX EXT: '{col_capex_ext}'" if col_capex_ext else "   ⚠️ Monto CAPEX EXT no encontrado")
    print(f"   📋 Monto CAPEX ORD: '{col_capex_ord}'" if col_capex_ord else "   ⚠️ Monto CAPEX ORD no encontrado")
    print(f"   📋 Prioridad: '{col_prio}'" if col_prio else "   ⚠️ Prioridad no encontrado")
    
    if col_capex_ext and col_capex_ord and col_prio and col_monto:
        ajustes_realizados = 0
        
        for idx, row in df_procesado.iterrows():
            # Verificar si ambas columnas CAPEX están vacías
            capex_ext_val = row[col_capex_ext]
            capex_ord_val = row[col_capex_ord]
            
            capex_ext_vacio = pd.isna(capex_ext_val) or capex_ext_val == 0 or capex_ext_val == ''
            capex_ord_vacio = pd.isna(capex_ord_val) or capex_ord_val == 0 or capex_ord_val == ''
            
            if capex_ext_vacio and capex_ord_vacio:
                # Ambas columnas vacías, revisar prioridad
                prioridad = row[col_prio]
                monto = row[col_monto]
                
                # Convertir prioridad a int si es posible
                try:
                    prioridad_int = int(prioridad) if pd.notna(prioridad) else None
                except:
                    prioridad_int = None
                
                if prioridad_int in prioridades_capex and pd.notna(monto):
                    tipo = prioridades_capex[prioridad_int]
                    
                    if tipo == 'EXT':
                        df_procesado.at[idx, col_capex_ext] = monto
                        df_procesado.at[idx, col_capex_ord] = 0
                    else:  # ORD
                        df_procesado.at[idx, col_capex_ord] = monto
                        df_procesado.at[idx, col_capex_ext] = 0
                    
                    ajustes_realizados += 1
        
        print(f"   ✅ Ajustes realizados: {ajustes_realizados} registros")
    else:
        print(f"   ⚠️ No se pudieron ajustar CAPEX (faltan columnas necesarias)")
    
    # ===================================================================
    # 5. Remover Proveedor Remito si existe
    # ===================================================================
    if "Proveedor Remito" in df_procesado.columns:
        print("🗑️ Removiendo columna 'Proveedor Remito' del consolidado")
        df_procesado = df_procesado.drop(columns=["Proveedor Remito"])
    else:
        print("ℹ️ Columna 'Proveedor Remito' no encontrada (ya removida o no existe)")
    
    print(f"\n✅ Procesamiento específico completado: {len(df_procesado)} filas, {len(df_procesado.columns)} columnas")
    
    # DEBUG: Mostrar columnas finales
    print(f"📋 Columnas finales para consolidado:")
    for i, col in enumerate(df_procesado.columns, 1):
        print(f"  {i:2d}. {col}")
    
    return df_procesado


def procesar_colombia(archivo_reporte_pago, archivo_reporte_absoluto=None):
    """Procesar consolidado CAPEX para Colombia con TIENDA desde Reporte Absoluto"""
    print("🇻🇪 PROCESANDO CONSOLIDADO CAPEX Colombia + REPORTE ABSOLUTO")
    print("=" * 70)
    
    try:
        # Diagnóstico previo del archivo principal
        print(f"\n🔍 DIAGNÓSTICO PREVIO DEL ARCHIVO...")
        skip_recomendado = diagnosticar_archivo_colombia(archivo_reporte_pago)
        
        if skip_recomendado is None:
            print("❌ Archivo Reporte Pago no compatible")
            return None
        
        # Validar Reporte Absoluto si se proporciona
        if archivo_reporte_absoluto:
            if not validar_reporte_absoluto(archivo_reporte_absoluto):
                print("⚠️ Continuando sin Reporte Absoluto")
                archivo_reporte_absoluto = None
        else:
            print("ℹ️ No se proporcionó Reporte Absoluto - columna TIENDA_BUSCARV quedará vacía")
        
        # NUEVO: Cargar Google Sheets para Solicitantes-Áreas
        print(f"\n📊 CARGANDO DATOS DE GOOGLE SHEETS...")
        print("-" * 50)
        
        df_solicitantes = leer_google_sheet_con_configuracion()
        lookup_solicitantes_areas = crear_lookup_solicitantes_areas(df_solicitantes)
        
        # DEBUG: Verificar que el lookup tiene datos
        print(f"🔍 DEBUG: lookup_solicitantes_areas contiene {len(lookup_solicitantes_areas)} entradas")

        if lookup_solicitantes_areas:
            print(f"✅ Google Sheets cargado: {len(lookup_solicitantes_areas)} solicitantes mapeados")
        else:
            print(f"⚠️ Google Sheets no disponible, columna AREA usará valores por defecto")
            
        # NUEVA LÓGICA: Obtener tasa del viernes de la semana pasada
        print(f"\n💰 OBTENIENDO TASA DEL VIERNES ANTERIOR...")
        print("-" * 50)

        # 1. Obtener tasa de cambio
        api_helper = APIHelper()
        tasa_dolar, fecha_tasa = api_helper.obtener_tasa_colombia()
        
        if not tasa_dolar or tasa_dolar <= 0:
            print("❌ No se pudo obtener tasa de cambio del viernes anterior")
            return None
        
        print(f"✅ Tasa seleccionada: {tasa_dolar:.2f} COP/USD (fecha: {fecha_tasa})")
        
        # 2. Leer archivo principal
        print(f"\n📂 CARGANDO ARCHIVO CON ESTRUCTURA DETECTADA...")
        print("-" * 30)
        df_reporte = leer_excel_safe(archivo_reporte_pago)
        if df_reporte is None:
            return None
        
        # ELIMINAR columna "Banco" si existe (ANTES de cualquier validación o procesamiento)
        # Esta columna no es necesaria y desbarata el mapeo de columnas
        if 'Banco' in df_reporte.columns:
            print(f"\n⚠️  Columna 'Banco' detectada. Eliminándola para evitar problemas de mapeo...")
            df_reporte = df_reporte.drop(columns=['Banco'])
            print(f"✅ Columna 'Banco' eliminada. Columnas restantes: {len(df_reporte.columns)}")
        
        # 3. Validaciones específicas de Colombia
        print(f"\n🔍 INICIANDO VALIDACIONES...")
        print("-" * 30)
        
        if not validar_columnas_colombia(df_reporte):
            print(f"\n❌ VALIDACIÓN FALLÓ - Estructura incorrecta")
            return None
        
        validar_monedas_colombia(df_reporte)
        
        # 4. Procesamiento específico de Colombia
        print(f"\n🔧 PROCESANDO DATOS...")
        print("-" * 30)
        df_procesado = procesar_datos_colombia_especifico(df_reporte)
        
        # 5. Crear archivo consolidado CON TIENDA
        print(f"\n📝 CREANDO CONSOLIDADO CON TIENDA...")
        print("-" * 50)
        
        excel_processor = ExcelProcessor('COLOMBIA', 'COP', tasa_dolar, archivo_reporte_absoluto, lookup_solicitantes_areas)
        nombre_salida = "ConsolidadoCapexCOLOMBIA.xlsx"
        
        if excel_processor.crear_archivo_consolidado(df_procesado, nombre_salida):
            print(f"\n✅ CONSOLIDADO COLOMBIA CON TIENDA COMPLETADO")
            print("=" * 70)
            
            tienda_info = "INCLUIDA (con BUSCARV)" if archivo_reporte_absoluto else "VACÍA (sin Reporte Absoluto)"
            area_info = "DESDE GOOGLE SHEETS" if lookup_solicitantes_areas else "VALORES POR DEFECTO"

            print(f"💰 TASA UTILIZADA:")
            print(f"   📅 Fecha: {fecha_tasa}")
            print(f"   💵 Valor: {tasa_dolar:.2f} COP/USD")
            print(f"   🗓️ Lógica: Viernes de semana pasada")

            print(f"📊 ESTRUCTURA FINAL:")
            print(f"   📄 Originales: 23 columnas")
            print(f"   💱 Calculadas: 19 columnas")  # ACTUALIZADO: 18 → 19
            print(f"   📋 TOTAL: 42 columnas")       # ACTUALIZADO: 41 → 42

            print(f"\n📊 NUEVAS COLUMNAS:")
            print(f"   36. AJ - TIENDA_LOOKUP ({tienda_info})")
            print(f"   37. AK - CECO")
            print(f"   38. AL - PROYECTO") 
            print(f"   39. AM - AREA ({area_info})")
            print(f"   40. AN - FECHA RECIBO")
            print(f"   41. AO - DESCRIPCIÓN")
            print(f"   42. AP - AÑO FISCAL ✨ (Agosto-Julio)")  # NUEVO
            
            resultado = {
                'archivo_salida': nombre_salida,
                'filas_procesadas': len(df_procesado),
                'tasa_utilizada': tasa_dolar,
                'fecha_tasa': fecha_tasa,  # NUEVO
                'logica_tasa': 'Viernes anterior',  # NUEVO
                'pais': 'COLOMBIA',
                'moneda': 'COP',
                'columnas_consolidado': 42,
                'columnas_calculadas': 19,
                'estructura': 'Consolidado CAPEX + Reporte Absoluto + Google Sheets + Tasa Viernes Anterior'
            }
            return resultado, excel_processor
        else:
            return None
            
    except Exception as e:
        print(f"❌ Error procesando Colombia: {e}")
        import traceback
        traceback.print_exc()
        return None


def generar_excel_colombia_con_detalle(df_bosqueto_original: pd.DataFrame, 
                                         df_detalle_corregido: pd.DataFrame) -> str:
    """
    Generar archivo Excel para Colombia con dos hojas: BOSQUETO y DETALLE CORREGIDO
    
    Args:
        df_bosqueto_original: DataFrame con datos del BOSQUETO original
        df_detalle_corregido: DataFrame con registros que fueron cargados a BQ
    
    Returns:
        str: Ruta del archivo temporal generado
    """
    print(f"📝 Generando Excel Colombia con BOSQUETO + DETALLE CORREGIDO...")
    
    # Crear archivo temporal
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_path = temp_file.name
    temp_file.close()
    
    try:
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            # Hoja 1: BOSQUETO (original con todas las facturas)
            df_bosqueto_original.to_excel(writer, sheet_name='BOSQUETO', index=False)
            print(f"   ✅ Hoja 'BOSQUETO' creada: {len(df_bosqueto_original)} filas")
            
            # Hoja 2: DETALLE CORREGIDO (solo las que se cargaron)
            if not df_detalle_corregido.empty:
                df_detalle_corregido.to_excel(writer, sheet_name='DETALLE CORREGIDO', index=False)
                print(f"   ✅ Hoja 'DETALLE CORREGIDO' creada: {len(df_detalle_corregido)} filas")
            else:
                # Crear hoja vacía con headers si no hay datos
                df_vacio = pd.DataFrame(columns=df_bosqueto_original.columns)
                df_vacio.to_excel(writer, sheet_name='DETALLE CORREGIDO', index=False)
                print(f"   ⚠️ Hoja 'DETALLE CORREGIDO' vacía (sin registros nuevos)")
        
        # Opcional: Aplicar estilos con openpyxl
        aplicar_estilos_excel_colombia(temp_path)
        
        print(f"✅ Excel Colombia generado: {temp_path}")
        return temp_path
        
    except Exception as e:
        print(f"❌ Error generando Excel: {e}")
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise

def limpiar_valor_para_excel(valor):
    """
    Limpiar valor para escribir en Excel: convertir NaN/NaT a None o string vacío
    Maneja todos los tipos de datos de pandas correctamente
    """
    # Verificar si es NaN o NaT (funciona con cualquier tipo de NaN de pandas)
    if pd.isna(valor):
        return ""  # None se escribe como celda vacía en Excel
    
    # Si es un tipo que Excel puede manejar directamente, retornarlo
    if isinstance(valor, (int, float, str, bool)):
        return valor
    
    # Manejar pd.Timestamp
    if isinstance(valor, pd.Timestamp):
        return valor
    
    # Para otros tipos (numpy, etc.), convertir a string o el tipo base
    try:
        # Intentar convertir a Python nativo
        if hasattr(valor, 'item'):  # numpy scalar
            return valor.item()
        return str(valor) if valor is not None else ""
    except:
        return ""

def agregar_hoja_detalle_al_excel(archivo_excel: str, df_detalle: pd.DataFrame):
    """
    Agregar hoja DETALLE CORREGIDO a un Excel existente sin borrar las fórmulas
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    
    print(f"📝 Agregando hoja DETALLE CORREGIDO al Excel existente...")
    
    # NO usar fillna en todo el DataFrame (causa problemas con tipos Int64, etc.)
    # En su lugar, limpiar valores individualmente al escribir
    
    # Cargar Excel existente (con fórmulas)
    wb = load_workbook(archivo_excel)
    
    # Eliminar hoja si existe
    if 'DETALLE CORREGIDO' in wb.sheetnames:
        del wb['DETALLE CORREGIDO']
    
    # Crear nueva hoja
    ws = wb.create_sheet('DETALLE CORREGIDO')
    
    # Escribir headers
    for col_idx, header in enumerate(df_detalle.columns, 1):
        header_limpio = limpiar_valor_para_excel(header)
        cell = ws.cell(row=1, column=col_idx, value=header_limpio)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Escribir datos - limpiar cada valor individualmente
    for row_idx, row in df_detalle.iterrows():
        for col_idx, value in enumerate(row, 1):
            # Limpiar valor antes de escribir (maneja NaN correctamente)
            valor_limpio = limpiar_valor_para_excel(value)
            ws.cell(row=row_idx + 2, column=col_idx, value=valor_limpio)
    
    # Guardar
    wb.save(archivo_excel)
    print(f"✅ Hoja DETALLE CORREGIDO agregada: {len(df_detalle)} filas")

def crear_hoja_capex_pagado_por_recibo(archivo_excel: str, df_detalle: pd.DataFrame):
    """
    Crear hoja 'CAPEX PAGADO POR RECIBO' con 5 tablas dinámicas
    
    Args:
        archivo_excel: Ruta del archivo Excel
        df_detalle: DataFrame con DETALLE CORREGIDO
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import pandas as pd
    
    print(f"\n📊 Creando hoja CAPEX PAGADO POR RECIBO...")

        # DIAGNÓSTICO INICIAL
    print(f"\n🔍 DIAGNÓSTICO DEL DATAFRAME:")
    print(f"   Total filas: {len(df_detalle)}")
    print(f"   Total columnas: {len(df_detalle.columns)}")
    print(f"\n📋 Columnas disponibles:")
    for i, col in enumerate(df_detalle.columns, 1):
        print(f"   {i:2d}. {col}")
    
    # Verificar columnas clave
    columnas_necesarias = [
        'AÑO FISCAL', 'MES DE PAGO', 'MONTO A PAGAR CAPEX',
        'AREA', 'MONTO ORD', 'MONTO EXT',
        'METODO DE PAGO', 'SEMANA', 'Monto USD', 'DIA DE PAGO'
    ]
    
    print(f"\n✅ Verificación de columnas necesarias:")
    for col in columnas_necesarias:
        existe = col in df_detalle.columns
        emoji = "✅" if existe else "❌"
        print(f"   {emoji} {col}: {'ENCONTRADA' if existe else 'NO ENCONTRADA'}")
        
        if existe and len(df_detalle) > 0:
            # Mostrar valores únicos
            valores_unicos = df_detalle[col].dropna().unique()
            print(f"      Valores únicos: {len(valores_unicos)}")
            if len(valores_unicos) <= 5:
                print(f"      Ejemplos: {list(valores_unicos)}")
            else:
                print(f"      Ejemplos: {list(valores_unicos[:5])}...")
    
    # Obtener mes actual
    meses_en_espanol = {
        'JANUARY': 'ENERO',
        'FEBRUARY': 'FEBRERO',
        'MARCH': 'MARZO',
        'APRIL': 'ABRIL',
        'MAY': 'MAYO',
        'JUNE': 'JUNIO',
        'JULY': 'JULIO',
        'AUGUST': 'AGOSTO',
        'SEPTEMBER': 'SEPTIEMBRE',
        'OCTOBER': 'OCTUBRE',
        'NOVEMBER': 'NOVIEMBRE',
        'DECEMBER': 'DICIEMBRE'
    }

    # Obtener mes basado en el viernes de la semana pasada (igual que la columna SEMANA)
    # Usa la misma lógica que APIHelper._obtener_viernes_pasado()
    import datetime as dt
    
    # Obtener el viernes de la semana pasada (misma lógica que en utils.py)
    hoy = dt.date.today()
    dia_semana_actual = hoy.weekday()  # lunes=0, viernes=4, domingo=6
    
    # Calcular días hasta el viernes de esta semana
    dias_hasta_viernes_esta_semana = (4 - dia_semana_actual) % 7
    
    # Si hoy es viernes (dias_hasta_viernes_esta_semana = 0), el viernes pasado fue hace 7 días
    # Si no, el viernes pasado fue hace (dias_hasta_viernes_esta_semana + 7) días
    if dias_hasta_viernes_esta_semana == 0:
        dias_retroceso = 7
    else:
        dias_retroceso = dias_hasta_viernes_esta_semana + 7
    
    viernes_pasado = hoy - dt.timedelta(days=dias_retroceso)
    
    # Obtener el mes del viernes pasado
    meses = {
        1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
        5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
        9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
    }
    mes_actual = meses[viernes_pasado.month]
    print(f"\n📅 Mes actual para filtros (basado en viernes pasado): {mes_actual}")
    
    # Verificar filtro de mes
    print(f"\n🔍 Valores de 'MES DE PAGO' en el DataFrame:")
    if 'MES DE PAGO' in df_detalle.columns:
        meses_disponibles = df_detalle['MES DE PAGO'].value_counts()
        print(meses_disponibles)
        
        filas_mes_actual = len(df_detalle[df_detalle['MES DE PAGO'] == mes_actual])
        print(f"\n📊 Filas que coinciden con mes actual ({mes_actual}): {filas_mes_actual}")
        
        if filas_mes_actual == 0:
            print(f"⚠️ ADVERTENCIA: No hay registros para el mes {mes_actual}")
            print(f"   Meses disponibles: {list(meses_disponibles.index)}")
    
    # Cargar workbook
    wb = load_workbook(archivo_excel)
    
    # Eliminar si existe
    if 'CAPEX PAGADO POR RECIBO' in wb.sheetnames:
        del wb['CAPEX PAGADO POR RECIBO']
    
    ws = wb.create_sheet('CAPEX PAGADO POR RECIBO')
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
        
    fila_actual = 1
    
    # ===================================================================
    # TABLA 1: AÑO FISCAL vs MES DE PAGO (Sumatoria: MONTO A PAGAR CAPEX)
    # ===================================================================
    print(f"   📋 Tabla 1: AÑO FISCAL vs MES DE PAGO")
    
    ws[f'A{fila_actual}'] = "TABLA 1: CAPEX POR AÑO FISCAL Y MES DE PAGO"
    ws[f'A{fila_actual}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{fila_actual}'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    
    fila_actual += 1
    
    # Crear tabla dinámica
    tabla1 = pd.pivot_table(
        df_detalle,
        values='MONTO A PAGAR CAPEX',
        index='MES DE PAGO',
        columns='AÑO FISCAL',
        aggfunc='sum',
        fill_value=0
    )
    print(f"✅ Tabla 1 creada: {tabla1.shape[0]} filas x {tabla1.shape[1]} columnas")
    print(f"   Total: {tabla1.values.sum():,.2f}")
    print(f"\n{tabla1}")
    
    # Escribir tabla 1
    for col_idx, col_name in enumerate(tabla1.columns, 1):
        ws.cell(row=fila_actual, column=col_idx + 1, value=col_name)
        ws.cell(row=fila_actual, column=col_idx + 1).fill = header_fill
        ws.cell(row=fila_actual, column=col_idx + 1).font = header_font
    
    ws.cell(row=fila_actual, column=1, value="MES").fill = header_fill
    ws.cell(row=fila_actual, column=1).font = header_font
    
    fila_actual += 1
    
    for idx, (mes, row) in enumerate(tabla1.iterrows()):
        ws.cell(row=fila_actual + idx, column=1, value=limpiar_valor_para_excel(mes))
        for col_idx, valor in enumerate(row, 1):
            ws.cell(row=fila_actual + idx, column=col_idx + 1, value=limpiar_valor_para_excel(valor))
    
    fila_actual += len(tabla1) + 3
    
    # ===================================================================
    # TABLA 2: MONTO A PAGAR CAPEX vs AREA (Con filtro mes actual)
    # ===================================================================
    print(f"   📋 Tabla 2: CAPEX vs AREA (Mes: {mes_actual})")
    
    ws[f'A{fila_actual}'] = f"TABLA 2: CAPEX POR ÁREA (MES: {mes_actual})"
    ws[f'A{fila_actual}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{fila_actual}'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    
    fila_actual += 1
    
    # Filtrar por mes actual
    df_mes_actual = df_detalle[df_detalle['MES DE PAGO'] == mes_actual]
    
    tabla2 = pd.pivot_table(
        df_mes_actual,
        values=['MONTO ORD', 'MONTO EXT'],
        index='AREA',
        aggfunc='sum',
        fill_value=0
    )
    print(f"✅ Tabla 2 creada: {tabla2.shape[0]} áreas")
    print(f"   Total MONTO ORD: {tabla2['MONTO ORD'].sum():,.2f}")
    print(f"   Total MONTO EXT: {tabla2['MONTO EXT'].sum():,.2f}")
    print(f"\n{tabla2}")
    
    # Escribir tabla 2
    ws.cell(row=fila_actual, column=1, value="AREA").fill = header_fill
    ws.cell(row=fila_actual, column=1).font = header_font
    ws.cell(row=fila_actual, column=2, value="MONTO ORD").fill = header_fill
    ws.cell(row=fila_actual, column=2).font = header_font
    ws.cell(row=fila_actual, column=3, value="MONTO EXT").fill = header_fill
    ws.cell(row=fila_actual, column=3).font = header_font
    
    fila_actual += 1
    
    for idx, (area, row) in enumerate(tabla2.iterrows()):
        ws.cell(row=fila_actual + idx, column=1, value=limpiar_valor_para_excel(area))
        ws.cell(row=fila_actual + idx, column=2, value=limpiar_valor_para_excel(row.get('MONTO ORD', 0)))
        ws.cell(row=fila_actual + idx, column=3, value=limpiar_valor_para_excel(row.get('MONTO EXT', 0)))
    
    fila_actual += len(tabla2) + 3
    
    # ===================================================================
    # TABLA 3: METODO DE PAGO vs SEMANA (Sumatoria: MONTO USD)
    # ===================================================================
    print(f"   📋 Tabla 3: METODO DE PAGO vs SEMANA (Mes: {mes_actual})")
    
    ws[f'A{fila_actual}'] = f"TABLA 3: MONTO USD POR MÉTODO DE PAGO Y SEMANA (MES: {mes_actual})"
    ws[f'A{fila_actual}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{fila_actual}'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    
    fila_actual += 1
    
    tabla3 = pd.pivot_table(
        df_mes_actual,
        values='Monto USD',
        index='SEMANA',
        columns='METODO DE PAGO',
        aggfunc='sum',
        fill_value=0
    )
    print(f"✅ Tabla 3 creada: {tabla3.shape[0]} semanas x {tabla3.shape[1]} métodos")
    print(f"   Total: {tabla3.values.sum():,.2f}")
    print(f"\n{tabla3}")
    
    # Escribir tabla 3
    for col_idx, col_name in enumerate(tabla3.columns, 1):
        ws.cell(row=fila_actual, column=col_idx + 1, value=col_name)
        ws.cell(row=fila_actual, column=col_idx + 1).fill = header_fill
        ws.cell(row=fila_actual, column=col_idx + 1).font = header_font
    
    ws.cell(row=fila_actual, column=1, value="SEMANA").fill = header_fill
    ws.cell(row=fila_actual, column=1).font = header_font
    
    fila_actual += 1
    
    for idx, (semana, row) in enumerate(tabla3.iterrows()):
        ws.cell(row=fila_actual + idx, column=1, value=limpiar_valor_para_excel(semana))
        for col_idx, valor in enumerate(row, 1):
            ws.cell(row=fila_actual + idx, column=col_idx + 1, value=limpiar_valor_para_excel(valor))
    
    fila_actual += len(tabla3) + 3
    
    # ===================================================================
    # TABLA 4: METODO DE PAGO + DIA DE PAGO vs SEMANA
    # ===================================================================
    print(f"   📋 Tabla 4: METODO Y DÍA DE PAGO vs SEMANA (Mes: {mes_actual})")
    
    ws[f'A{fila_actual}'] = f"TABLA 4: MONTO ORD + EXT POR MÉTODO, DÍA Y SEMANA (MES: {mes_actual})"
    ws[f'A{fila_actual}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{fila_actual}'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    
    fila_actual += 1
    
    # Crear columna combinada
    df_mes_actual_copy = df_mes_actual.copy()
    df_mes_actual_copy['METODO_DIA'] = df_mes_actual_copy['METODO DE PAGO'].astype(str) + ' - ' + df_mes_actual_copy['DIA DE PAGO'].astype(str)
    
    tabla4 = pd.pivot_table(
        df_mes_actual_copy,
        values=['MONTO ORD', 'MONTO EXT'],
        index='SEMANA',
        columns='METODO_DIA',
        aggfunc='sum',
        fill_value=0
    )
    print(f"✅ Tabla 4 creada: {tabla4.shape[0]} semanas x {len(tabla4.columns)} combinaciones método-día")
    print(f"   Total MONTO ORD: {tabla4['MONTO ORD'].sum().sum():,.2f}")
    print(f"   Total MONTO EXT: {tabla4['MONTO EXT'].sum().sum():,.2f}")
    print(f"\n{tabla4}")
    
    # Escribir tabla 4 (formato más compacto)
    ws.cell(row=fila_actual, column=1, value="SEMANA").fill = header_fill
    ws.cell(row=fila_actual, column=1).font = header_font
    
    for col_idx, col_name in enumerate(tabla4.columns, 1):
        ws.cell(row=fila_actual, column=col_idx + 1, value=col_name[1])  # Solo el método-día
        ws.cell(row=fila_actual, column=col_idx + 1).fill = header_fill
        ws.cell(row=fila_actual, column=col_idx + 1).font = header_font
    
    fila_actual += 1
    
    for idx, (semana, row) in enumerate(tabla4.iterrows()):
        ws.cell(row=fila_actual + idx, column=1, value=limpiar_valor_para_excel(semana))
        for col_idx, valor in enumerate(row, 1):
            ws.cell(row=fila_actual + idx, column=col_idx + 1, value=limpiar_valor_para_excel(valor))
    
    fila_actual += len(tabla4) + 3
    
    # ===================================================================
    # TABLA 5: METODO DE PAGO vs SEMANA (Sumatoria: MONTO A PAGAR CAPEX)
    # ===================================================================
    print(f"   📋 Tabla 5: METODO DE PAGO vs SEMANA - CAPEX A PAGAR (Mes: {mes_actual})")
    
    ws[f'A{fila_actual}'] = f"TABLA 5: CAPEX A PAGAR POR MÉTODO Y SEMANA (MES: {mes_actual})"
    ws[f'A{fila_actual}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{fila_actual}'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    
    fila_actual += 1
    
    tabla5 = pd.pivot_table(
        df_mes_actual,
        values='MONTO A PAGAR CAPEX',
        index='SEMANA',
        columns='METODO DE PAGO',
        aggfunc='sum',
        fill_value=0
    )
    print(f"✅ Tabla 5 creada: {tabla5.shape[0]} semanas x {tabla5.shape[1]} métodos")
    print(f"   Total: {tabla5.values.sum():,.2f}")
    print(f"\n{tabla5}")
    
    # Escribir tabla 5
    for col_idx, col_name in enumerate(tabla5.columns, 1):
        ws.cell(row=fila_actual, column=col_idx + 1, value=col_name)
        ws.cell(row=fila_actual, column=col_idx + 1).fill = header_fill
        ws.cell(row=fila_actual, column=col_idx + 1).font = header_font
    
    ws.cell(row=fila_actual, column=1, value="SEMANA").fill = header_fill
    ws.cell(row=fila_actual, column=1).font = header_font
    
    fila_actual += 1
    
    for idx, (semana, row) in enumerate(tabla5.iterrows()):
        ws.cell(row=fila_actual + idx, column=1, value=limpiar_valor_para_excel(semana))
        for col_idx, valor in enumerate(row, 1):
            ws.cell(row=fila_actual + idx, column=col_idx + 1, value=limpiar_valor_para_excel(valor))
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col].width = 15
    
    # Guardar
    wb.save(archivo_excel)
    print(f"✅ Hoja 'CAPEX PAGADO POR RECIBO' creada con 5 tablas dinámicas")

def crear_hoja_presupuesto_mensual(archivo_excel: str, df_responsables: pd.DataFrame):
    """
    Crear hoja 'Presupuesto Mensual' con tabla de responsables por mes
    Columnas ordenadas por fecha ASCENDENTE (fecha más vieja primero)
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import pandas as pd
    from datetime import datetime
    
    print(f"\n" + "="*70)
    print(f"💰 CREANDO HOJA PRESUPUESTO MENSUAL")
    print(f"="*70)
    
    # Diagnóstico
    print(f"\n🔍 DIAGNÓSTICO DEL DATAFRAME:")
    print(f"   Filas: {len(df_responsables)}")
    print(f"   Columnas: {list(df_responsables.columns)}")
    
    if df_responsables.empty:
        print(f"⚠️ DataFrame vacío - abortando")
        return
    
    # Renombrar columnas para facilitar manejo
    df = df_responsables.copy()
    df.columns = ['anio_fiscal', 'fecha', 'tipo_capex', 'area', 'monto']
    
    print(f"\n📊 Muestra de datos:")
    print(df.head(10))
    
    # Crear columna de fecha para la tabla (formato: ago-25, sep-25, etc)
    df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')
    df['fecha_mes'] = df['fecha'].dt.strftime('%b-%y').str.upper()
    
    # Crear columna de fecha ordenable
    df['fecha_sortable'] = df['fecha']
    
    print(f"\n📅 Fechas disponibles:")
    fechas_unicas = df[['fecha', 'fecha_mes']].drop_duplicates().sort_values('fecha')
    print(fechas_unicas)
    
    # Crear tabla dinámica: TIPO + AREA (filas) vs FECHA (columnas)
    print(f"\n⚙️ Creando tabla dinámica...")
    tabla = pd.pivot_table(
        df,
        values='monto',
        index=['tipo_capex', 'area'],
        columns='fecha_mes',
        aggfunc='sum',
        fill_value=0,
        margins=True,  # Agregar totales
        margins_name='Total general'
    )
    
    # ===================================================================
    # ORDENAR COLUMNAS POR FECHA ASCENDENTE
    # ===================================================================
    # Crear un diccionario con el orden de fechas
    fecha_orden = (
        df[['fecha_mes', 'fecha_sortable']]
        .drop_duplicates()
        .sort_values('fecha_sortable')
    )
    
    # Mapeo de mes_año a posición ordenada
    orden_fechas = {mes: idx for idx, mes in enumerate(fecha_orden['fecha_mes'].values)}
    
    # Reordenar columnas (excepto Total general que va al final)
    columnas_ordenadas = sorted(
        [col for col in tabla.columns if col != 'Total general'],
        key=lambda x: orden_fechas.get(x, 999)
    )
    columnas_ordenadas.append('Total general')
    
    tabla = tabla[columnas_ordenadas]
    
    print(f"✅ Tabla creada: {tabla.shape[0]} filas (tipos+áreas) x {tabla.shape[1]} meses")
    print(f"📅 Orden de columnas (fecha ascendente):")
    for i, col in enumerate(tabla.columns, 1):
        print(f"   {i}. {col}")
    print(f"\n📊 Tipos únicos en datos: {df['tipo_capex'].unique()}")
    print(f"\n{tabla.head(20)}")
    
    # Cargar workbook
    wb = load_workbook(archivo_excel)
    
    # Eliminar si existe
    if 'Presupuesto Mensual' in wb.sheetnames:
        del wb['Presupuesto Mensual']
    
    ws = wb.create_sheet('Presupuesto Mensual')
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True, size=11)
    tipo_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Naranja para tipos
    tipo_font = Font(bold=True, size=11, color="000000")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    fila_actual = 1
    
    # Título
    ws[f'A{fila_actual}'] = "TABLA 1: PRESUPUESTO CAPEX POR RESPONSABLE Y MES"
    ws[f'A{fila_actual}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{fila_actual}'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws.merge_cells(f'A{fila_actual}:{chr(65 + len(tabla.columns))}{fila_actual}')
    fila_actual += 2
    
    # Encabezados de columnas (meses, ORDENADOS POR FECHA ASCENDENTE)
    ws.cell(row=fila_actual, column=1, value="Responsable").fill = header_fill
    ws.cell(row=fila_actual, column=1).font = header_font
    ws.cell(row=fila_actual, column=1).border = border
    ws.cell(row=fila_actual, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for col_idx, mes in enumerate(tabla.columns, 1):
        celda = ws.cell(row=fila_actual, column=col_idx + 1, value=mes)
        celda.fill = header_fill
        celda.font = header_font
        celda.border = border
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    fila_actual += 1
    
    # Organizar datos por tipo: primero CAPEX EXTRAORDINARIO, luego CAPEX ORDINARIO
    tipos_orden = ['CAPEX EXTRAORDINARIO', 'CAPEX ORDINARIO']
    tipo_anterior = None
    row_idx = 0
    
    # Separar el total general si existe
    filas_tabla = []
    fila_total = None
    
    for idx, (multi_index, fila_datos) in enumerate(tabla.iterrows()):
        # Verificar si es el total general (puede ser string o tupla con 'Total general')
        if isinstance(multi_index, str) and multi_index == 'Total general':
            fila_total = (multi_index, fila_datos)
        elif isinstance(multi_index, tuple) and len(multi_index) > 0 and multi_index[0] == 'Total general':
            fila_total = (multi_index, fila_datos)
        else:
            # multi_index es una tupla (tipo_capex, area)
            if isinstance(multi_index, tuple) and len(multi_index) == 2:
                tipo_capex, area = multi_index
                filas_tabla.append((tipo_capex, area, fila_datos))
            else:
                # Caso especial: si no es tupla, intentar extraer tipo y área de otra forma
                print(f"⚠️ Formato de índice inesperado: {multi_index} (tipo: {type(multi_index)})")
    
    # Ordenar filas: primero por tipo (EXTRAORDINARIO, luego ORDINARIO), luego por área
    def ordenar_filas(fila):
        tipo, area, _ = fila
        tipo_orden = tipos_orden.index(tipo) if tipo in tipos_orden else 999
        return (tipo_orden, area)
    
    filas_tabla.sort(key=ordenar_filas)
    
    # Escribir filas organizadas por tipo
    for tipo_capex, area, fila_datos in filas_tabla:
        # Si cambió el tipo, agregar fila de encabezado de tipo
        if tipo_capex != tipo_anterior:
            # Fila de encabezado de tipo
            celda_tipo = ws.cell(row=fila_actual + row_idx, column=1, value=tipo_capex)
            celda_tipo.fill = tipo_fill
            celda_tipo.font = tipo_font
            celda_tipo.border = border
            celda_tipo.alignment = Alignment(horizontal='left', vertical='center')
            
            # Calcular totales por tipo para cada columna
            tipo_filas = [(t, a, fd) for t, a, fd in filas_tabla if t == tipo_capex]
            for col_idx, mes in enumerate(tabla.columns, 1):
                total_tipo = sum(fd[mes] for _, _, fd in tipo_filas)
                celda = ws.cell(row=fila_actual + row_idx, column=col_idx + 1, value=limpiar_valor_para_excel(total_tipo))
                celda.fill = tipo_fill
                celda.font = tipo_font
                celda.border = border
                celda.number_format = '#,##0.00'
                celda.alignment = Alignment(horizontal='right')
            
            row_idx += 1
            tipo_anterior = tipo_capex
        
        # Fila de área
        celda_area = ws.cell(row=fila_actual + row_idx, column=1, value=limpiar_valor_para_excel(area))
        celda_area.border = border
        celda_area.alignment = Alignment(horizontal='left', vertical='center')
        
        # Datos de montos
        for col_idx, valor in enumerate(fila_datos, 1):
            celda = ws.cell(row=fila_actual + row_idx, column=col_idx + 1, value=limpiar_valor_para_excel(valor))
            celda.border = border
            celda.number_format = '#,##0.00'
            celda.alignment = Alignment(horizontal='right')
        
        row_idx += 1
    
    # Agregar fila de total general al final si existe
    if fila_total:
        total_idx, total_datos = fila_total
        celda_total = ws.cell(row=fila_actual + row_idx, column=1, value=limpiar_valor_para_excel(total_idx))
        celda_total.fill = total_fill
        celda_total.font = total_font
        celda_total.border = border
        celda_total.alignment = Alignment(horizontal='left', vertical='center')
        
        for col_idx, valor in enumerate(total_datos, 1):
            celda = ws.cell(row=fila_actual + row_idx, column=col_idx + 1, value=limpiar_valor_para_excel(valor))
            celda.border = border
            celda.number_format = '#,##0.00'
            celda.alignment = Alignment(horizontal='right')
            celda.fill = total_fill
            celda.font = total_font
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 25
    for col_idx in range(len(tabla.columns)):
        col_letter = chr(66 + col_idx)  # B, C, D, etc.
        ws.column_dimensions[col_letter].width = 14
    
    # Guardar
    wb.save(archivo_excel)
    print(f"\n✅ Hoja 'Presupuesto Mensual' creada exitosamente")
    print(f"   ✅ Columnas ordenadas por FECHA ASCENDENTE (más vieja → más nueva)")
    print(f"="*70)

def extraer_tabla2_capex_pagado_recibo(archivo_excel: str) -> pd.DataFrame:
    """
    Extraer la tabla 2 (CAPEX vs AREA) de la hoja CAPEX PAGADO POR RECIBO
    La tabla comienza en B8 aproximadamente
    """
    import openpyxl
    import pandas as pd
    
    print(f"\n📊 Extrayendo tabla 2 de CAPEX PAGADO POR RECIBO...")
    
    try:
        # Cargar el workbook
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
        ws = wb['CAPEX PAGADO POR RECIBO']
        
        # Buscar la tabla 2 (comienza alrededor de fila 8)
        # Estructura: AREA | MONTO ORD | MONTO EXT
        datos = []
        
        # Escanear desde fila 8 hasta encontrar "Total"
        for row_idx in range(8, ws.max_row + 1):
            col_a = ws[f'A{row_idx}'].value
            col_b = ws[f'B{row_idx}'].value
            col_c = ws[f'C{row_idx}'].value
            
            if col_a is None or col_a == '':
                continue
            
            # Convertir a string para comparaciones
            col_a_str = str(col_a).upper()
            
            # Si es "Total", detener
            if col_a_str == 'TOTAL':
                break
            
            # Si es header, saltar
            if col_a_str in ['AREA', 'SEMANA', 'MES', 'RESPONSABLE']:
                continue
            
            # Filtrar títulos de otras tablas (TABLA 3, TABLA 4, TABLA 5, etc.)
            if 'TABLA' in col_a_str and ('MONTO USD' in col_a_str or 'MONTO ORD' in col_a_str or 'CAPEX A PAGAR' in col_a_str):
                continue
            
            # Filtrar cualquier fila que contenga "TABLA" seguido de un número
            if 'TABLA' in col_a_str:
                continue
            
            # Validar que col_b o col_c tengan valores numéricos (para asegurar que es una fila de datos)
            # Si ambas columnas están vacías o no son numéricas, probablemente no es una fila de datos válida
            try:
                monto_ord = float(col_b) if pd.notna(col_b) and col_b != '' else 0
                monto_ext = float(col_c) if pd.notna(col_c) and col_c != '' else 0
            except (ValueError, TypeError):
                # Si no se puede convertir a número, probablemente no es una fila de datos
                continue
            
            # Filtrar valores numéricos simples que no son áreas válidas
            # (como "3", "4", "5" que pueden ser números de tabla o índices)
            # Un área válida generalmente tiene al menos 3 caracteres o contiene letras
            col_a_clean = str(col_a).strip()
            if col_a_clean.isdigit() and len(col_a_clean) <= 2:
                # Es un número simple (1-99), probablemente no es un área
                continue
            
            # Filtrar si el área es solo un número sin contexto
            # (pero permitir áreas que sean números con contexto, como "AREA 1")
            if col_a_clean.isdigit():
                continue
            
            # Agregar fila
            datos.append({
                'area': col_a,
                'monto_ord': monto_ord,
                'monto_ext': monto_ext
            })
        
        df_tabla2 = pd.DataFrame(datos)
        
        print(f"✅ Tabla 2 extraída: {len(df_tabla2)} áreas")
        print(f"\n{df_tabla2}")
        
        return df_tabla2
        
    except Exception as e:
        print(f"❌ Error extrayendo tabla 2: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()

def crear_tabla2_presupuesto_mensual(archivo_excel: str, df_diferencia: pd.DataFrame, df_ejecutado: pd.DataFrame):
    """
    Crear TABLA 2: Presupuesto vs Ejecutado vs Diferencia
    Con nombres de columnas DINÁMICOS según el mes actual
    """
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    import pandas as pd
    from datetime import datetime, timedelta
    from dateutil.relativedelta import relativedelta
    
    print(f"\n" + "="*70)
    print(f"📊 CREANDO TABLA 2: PRESUPUESTO vs EJECUTADO vs DIFERENCIA")
    print(f"="*70)
    
    if df_diferencia.empty or df_ejecutado.empty:
        print(f"⚠️ DataFrames vacíos - abortando")
        return None
    
    # ===================================================================
    # CALCULAR NOMBRES DE COLUMNAS DINÁMICOS
    # Basado en el viernes de la semana pasada (igual que la columna SEMANA)
    # Usa la misma lógica que APIHelper._obtener_viernes_pasado()
    # ===================================================================
    import datetime as dt
    
    # Obtener el viernes de la semana pasada (misma lógica que en utils.py)
    hoy = dt.date.today()
    dia_semana_actual = hoy.weekday()  # lunes=0, viernes=4, domingo=6
    
    # Calcular días hasta el viernes de esta semana
    dias_hasta_viernes_esta_semana = (4 - dia_semana_actual) % 7
    
    # Si hoy es viernes (dias_hasta_viernes_esta_semana = 0), el viernes pasado fue hace 7 días
    # Si no, el viernes pasado fue hace (dias_hasta_viernes_esta_semana + 7) días
    if dias_hasta_viernes_esta_semana == 0:
        dias_retroceso = 7
    else:
        dias_retroceso = dias_hasta_viernes_esta_semana + 7
    
    viernes_pasado = hoy - dt.timedelta(days=dias_retroceso)
    
    # Mes actual es el mes del viernes pasado
    mes_actual = viernes_pasado
    # Mes anterior es el mes anterior al del viernes pasado
    mes_anterior = viernes_pasado - relativedelta(months=1)
    
    # Traducción de meses
    meses_espanol = {
        'JANUARY': 'ENE', 'FEBRUARY': 'FEB', 'MARCH': 'MAR', 'APRIL': 'ABR',
        'MAY': 'MAY', 'JUNE': 'JUN', 'JULY': 'JUL', 'AUGUST': 'AGO',
        'SEPTEMBER': 'SEP', 'OCTOBER': 'OCT', 'NOVEMBER': 'NOV', 'DECEMBER': 'DIC'
    }
    
    # Obtener nombres de meses en español con año
    mes_actual_str_en = mes_actual.strftime('%B').upper()
    mes_anterior_str_en = mes_anterior.strftime('%B').upper()
    
    mes_actual_str = meses_espanol.get(mes_actual_str_en, mes_actual_str_en[:3])
    mes_anterior_str = meses_espanol.get(mes_anterior_str_en, mes_anterior_str_en[:3])
    
    anio_actual = mes_actual.strftime('%y')
    anio_anterior = mes_anterior.strftime('%y')
    
    # Nombres dinámicos de columnas
    col_remanente = f"Remanente {mes_anterior_str}-{anio_anterior}"
    col_presupuesto = f"Presupuesto {mes_actual_str}-{anio_actual}"
    col_ejecutado = f"Ejecutado {mes_actual_str}-{anio_actual}"
    col_diferencia = "Diferencia"
    
    print(f"\n📅 NOMBRES DE COLUMNAS DINÁMICOS:")
    print(f"   Mes anterior: {mes_anterior_str}-{anio_anterior}")
    print(f"   Mes actual: {mes_actual_str}-{anio_actual}")
    print(f"   Columnas:")
    print(f"     1. {col_remanente}")
    print(f"     2. {col_presupuesto}")
    print(f"     3. {col_ejecutado}")
    print(f"     4. {col_diferencia}")
    
    # Renombrar columnas para facilitar
    df_dif = df_diferencia.copy()
    df_dif.columns = ['mes', 'tipo', 'area', 'remanente', 'presupuesto', 'fecha']
    
    df_ej = df_ejecutado.copy()
    df_ej.columns = ['area', 'monto_ord', 'monto_ext']
    
    print(f"\n🔍 DIAGNÓSTICO INICIAL:")
    print(f"   Diferencia - Áreas: {df_dif['area'].nunique()}")
    print(f"   Ejecutado - Áreas: {df_ej['area'].nunique()}")
    
    # ===================================================================
    # NORMALIZAR NOMBRES DE ÁREAS
    # ===================================================================
    def normalizar_area(area):
        """Normalizar nombres de áreas según reglas de negocio"""
        if pd.isna(area):
            return area
        area_str = str(area).strip()
        area_upper = area_str.upper()
        
        # Presidencia → Dirección de Retail (case insensitive)
        if 'PRESIDENCIA' in area_upper:
            return 'Dirección de Retail'
        
        # Dirección de Retail / DIRECCIÓN DE RETAIL → Dirección de Retail (normalizar)
        if 'DIRECCION' in area_upper and 'RETAIL' in area_upper:
            return 'Dirección de Retail'
        
        # TI → VP Tecnología de la Información (case insensitive)
        if area_upper in ['TI', 'T.I.', 'T.I']:
            return 'VP Tecnología de la Información'
        
        # VP Tecnología de la Información / VP TECNOLOGIA DE LA INFORMACION → VP Tecnología de la Información (normalizar)
        if 'TECNOLOGIA' in area_upper and 'INFORMACION' in area_upper:
            return 'VP Tecnología de la Información'
        
        # Eliminar Importación y Servicios
        if 'IMPORTACION' in area_upper or 'IMPORTACIÓN' in area_upper:
            return None  # Marcar para eliminar
        if 'SERVICIOS' in area_upper:
            return None  # Marcar para eliminar
        
        return area_str
    
    # Aplicar normalización
    df_dif['area_normalizada'] = df_dif['area'].apply(normalizar_area)
    df_ej['area_normalizada'] = df_ej['area'].apply(normalizar_area)
    
    # Eliminar áreas marcadas como None
    df_dif = df_dif[df_dif['area_normalizada'].notna()].copy()
    df_ej = df_ej[df_ej['area_normalizada'].notna()].copy()
    
    # Reemplazar columna area con area_normalizada
    df_dif['area'] = df_dif['area_normalizada']
    df_ej['area'] = df_ej['area_normalizada']
    df_dif = df_dif.drop(columns=['area_normalizada'])
    df_ej = df_ej.drop(columns=['area_normalizada'])
    
    print(f"\n🔍 DIAGNÓSTICO DESPUÉS DE NORMALIZACIÓN:")
    print(f"   Diferencia - Áreas: {df_dif['area'].nunique()}")
    print(f"   Ejecutado - Áreas: {df_ej['area'].nunique()}")
    
    # ===================================================================
    # MANEJAR DIR CONSTRUCCIÓN Y PROYECTOS (SEPARAR ORD Y EXT)
    # ===================================================================
    # Identificar DIR CONSTRUCCIÓN Y PROYECTOS
    def es_construccion_proyectos(area):
        """Verificar si es DIR CONSTRUCCIÓN Y PROYECTOS"""
        if pd.isna(area):
            return False
        area_str = str(area).strip().upper()
        return 'DIR CONSTRUCCIÓN' in area_str and 'PROYECTOS' in area_str
    
    # IMPORTANTE: Para CONSTRUCCIÓN, necesitamos separar según acento ANTES de pivotar
    # Crear una función para identificar si tiene acento
    def tiene_acento_construccion(area):
        """Verificar si CONSTRUCCIÓN tiene acento (EXTRAORDINARIO)"""
        if pd.isna(area):
            return False
        area_str = str(area).strip()
        # Buscar "CONSTRUCCIÓN" con acento (ó) vs "CONSTRUCCION" sin acento
        # Si tiene acento (ó), es EXTRAORDINARIO
        # Verificar si contiene el carácter con acento (ó) en cualquier variación
        return 'CONSTRUCCIÓN' in area_str or 'CONSTRUCCIÓN' in area_str.upper() or 'ó' in area_str or 'Ó' in area_str
    
    # Separar df_dif en CONSTRUCCIÓN con acento y sin acento
    df_dif_construccion = df_dif[df_dif['area'].apply(es_construccion_proyectos)].copy()
    df_dif_otros = df_dif[~df_dif['area'].apply(es_construccion_proyectos)].copy()
    
    # Para CONSTRUCCIÓN, separar según acento
    df_dif_construccion_extra = df_dif_construccion[df_dif_construccion['area'].apply(tiene_acento_construccion)].copy()
    df_dif_construccion_ord = df_dif_construccion[~df_dif_construccion['area'].apply(tiene_acento_construccion)].copy()
    
    # Agregar columna temporal para identificar tipo
    if not df_dif_construccion_extra.empty:
        df_dif_construccion_extra['tipo_temp'] = 'EXTRAORDINARIO'
    if not df_dif_construccion_ord.empty:
        df_dif_construccion_ord['tipo_temp'] = 'ORDINARIO'
    
    # Normalizar nombre de área para CONSTRUCCIÓN (usar mismo nombre para ambas)
    area_construccion_nombre = 'DIR CONSTRUCCIÓN Y PROYECTOS'  # Nombre estándar
    if not df_dif_construccion_extra.empty:
        df_dif_construccion_extra['area'] = area_construccion_nombre
    if not df_dif_construccion_ord.empty:
        df_dif_construccion_ord['area'] = area_construccion_nombre
    
    # Recombinar
    df_dif_procesado = pd.concat([df_dif_otros, df_dif_construccion_extra, df_dif_construccion_ord], ignore_index=True)
    
    # Pivotar diferencia para obtener Presupuesto y Remanente
    # Para CONSTRUCCIÓN, necesitamos pivotar por área Y tipo
    if 'tipo_temp' in df_dif_procesado.columns:
        # Para CONSTRUCCIÓN, pivotar con área y tipo
        tabla_dif_construccion = pd.pivot_table(
            df_dif_procesado[df_dif_procesado['area'] == area_construccion_nombre],
            values=['remanente', 'presupuesto'],
            index=['area', 'tipo_temp'],
            aggfunc='sum',
            fill_value=0
        )
        
        # Para otras áreas, pivotar normal
        tabla_dif_otros = pd.pivot_table(
            df_dif_procesado[df_dif_procesado['area'] != area_construccion_nombre],
            values=['remanente', 'presupuesto'],
            index='area',
            aggfunc='sum',
            fill_value=0
        )
    else:
        # Si no hay CONSTRUCCIÓN, pivotar normal
        tabla_dif = pd.pivot_table(
            df_dif_procesado,
            values=['remanente', 'presupuesto'],
            index='area',
            aggfunc='sum',
            fill_value=0
        )
        tabla_dif_construccion = pd.DataFrame()
        tabla_dif_otros = tabla_dif
    
    # Separar ejecutado: para CONSTRUCCIÓN separar ORD y EXT, para otros sumar
    df_ej_construccion = df_ej[df_ej['area'].apply(es_construccion_proyectos)].copy()
    df_ej_otros = df_ej[~df_ej['area'].apply(es_construccion_proyectos)].copy()
    
    # Crear filas de ejecutado procesadas
    ejecutado_rows = []
    
    # Para CONSTRUCCIÓN: crear dos filas separadas (ORD y EXT)
    if not df_ej_construccion.empty:
        for area_original, group in df_ej_construccion.groupby('area'):
            monto_ord_total = group['monto_ord'].sum()
            monto_ext_total = group['monto_ext'].sum()
            
            # Normalizar nombre de área
            area_nombre_normalizado = area_construccion_nombre
            
            # Fila ORDINARIO (solo si hay monto)
            if monto_ord_total > 0:
                ejecutado_rows.append({
                    'area': area_nombre_normalizado,
                    'ejecutado': monto_ord_total,
                    'es_extraordinario': False
                })
            
            # Fila EXTRAORDINARIO (solo si hay monto)
            if monto_ext_total > 0:
                ejecutado_rows.append({
                    'area': area_nombre_normalizado,
                    'ejecutado': monto_ext_total,
                    'es_extraordinario': True
                })
    
    # Para otras áreas: sumar ORD + EXT
    for idx, row in df_ej_otros.iterrows():
        ejecutado_rows.append({
            'area': row['area'],
            'ejecutado': row['monto_ord'] + row['monto_ext'],
            'es_extraordinario': False  # Se asignará después según el área
        })
    
    # Crear DataFrame de ejecutado procesado
    df_ejecutado_procesado = pd.DataFrame(ejecutado_rows)
    
    # Construir tabla final combinando diferencia y ejecutado
    tabla_final_rows = []
    
    # Procesar cada área con ejecutado
    for area_name in df_ejecutado_procesado['area'].unique():
        # Obtener todas las filas de ejecutado para esta área
        ejecutados_area = df_ejecutado_procesado[df_ejecutado_procesado['area'] == area_name]
        
        # Para CONSTRUCCIÓN, buscar en tabla_dif_construccion por tipo
        if area_name == area_construccion_nombre and not tabla_dif_construccion.empty:
            # Buscar remanente y presupuesto según tipo (ORD o EXT)
            for _, ej_row in ejecutados_area.iterrows():
                tipo_buscar = 'EXTRAORDINARIO' if ej_row['es_extraordinario'] else 'ORDINARIO'
                
                # Buscar en tabla_dif_construccion
                try:
                    if (area_name, tipo_buscar) in tabla_dif_construccion.index:
                        remanente = tabla_dif_construccion.loc[(area_name, tipo_buscar), 'remanente']
                        presupuesto = tabla_dif_construccion.loc[(area_name, tipo_buscar), 'presupuesto']
                    else:
                        remanente = 0
                        presupuesto = 0
                except:
                    remanente = 0
                    presupuesto = 0
                
                ejecutado = ej_row['ejecutado']
                diferencia = ejecutado - presupuesto + remanente
                
                tabla_final_rows.append({
                    'area': area_name,
                    'remanente': remanente,
                    'presupuesto': presupuesto,
                    'ejecutado': ejecutado,
                    'diferencia': diferencia,
                    'es_extraordinario': ej_row['es_extraordinario']
                })
        else:
            # Para otras áreas, buscar en tabla_dif_otros
            if area_name in tabla_dif_otros.index:
                remanente = tabla_dif_otros.loc[area_name, 'remanente']
                presupuesto = tabla_dif_otros.loc[area_name, 'presupuesto']
            else:
                remanente = 0
                presupuesto = 0
            
            # Crear una fila en tabla_final por cada fila de ejecutado
            for _, ej_row in ejecutados_area.iterrows():
                ejecutado = ej_row['ejecutado']
                diferencia = ejecutado - presupuesto + remanente
                
                tabla_final_rows.append({
                    'area': area_name,
                    'remanente': remanente,
                    'presupuesto': presupuesto,
                    'ejecutado': ejecutado,
                    'diferencia': diferencia,
                    'es_extraordinario': ej_row['es_extraordinario']
                })
    
    # Agregar áreas que están en diferencia pero no en ejecutado
    # Para tabla_dif_otros
    for area_name in tabla_dif_otros.index:
        if area_name not in df_ejecutado_procesado['area'].values:
            remanente = tabla_dif_otros.loc[area_name, 'remanente']
            presupuesto = tabla_dif_otros.loc[area_name, 'presupuesto']
            ejecutado = 0
            diferencia = ejecutado - presupuesto + remanente
            
            tabla_final_rows.append({
                'area': area_name,
                'remanente': remanente,
                'presupuesto': presupuesto,
                'ejecutado': ejecutado,
                'diferencia': diferencia,
                'es_extraordinario': False
            })
    
    # Para tabla_dif_construccion (si existe)
    if not tabla_dif_construccion.empty:
        for (area_name, tipo), row in tabla_dif_construccion.iterrows():
            # Verificar si ya existe esta combinación en ejecutado
            existe = False
            for ej_row in ejecutado_rows:
                if ej_row['area'] == area_name and ej_row['es_extraordinario'] == (tipo == 'EXTRAORDINARIO'):
                    existe = True
                    break
            
            if not existe:
                remanente = row['remanente']
                presupuesto = row['presupuesto']
                ejecutado = 0
                diferencia = ejecutado - presupuesto + remanente
                
                tabla_final_rows.append({
                    'area': area_name,
                    'remanente': remanente,
                    'presupuesto': presupuesto,
                    'ejecutado': ejecutado,
                    'diferencia': diferencia,
                    'es_extraordinario': (tipo == 'EXTRAORDINARIO')
                })
    
    # Crear tabla final
    tabla_final = pd.DataFrame(tabla_final_rows)
    
    # NO usar set_index porque puede haber áreas duplicadas (CONSTRUCCIÓN con ORD y EXT)
    # Mantener 'area' como columna normal
    
    # Renombrar columnas con nombres dinámicos
    tabla_final.columns = ['area', col_remanente, col_presupuesto, col_ejecutado, col_diferencia, 'es_extraordinario']
    
    # ===================================================================
    # ASIGNAR TIPO CAPEX A CADA ÁREA Y AGRUPAR
    # ===================================================================
    def asignar_tipo_capex_area(row):
        """Asignar tipo CAPEX según el área y si es extraordinario"""
        # Obtener el nombre del área desde la columna
        area = row['area'] if 'area' in row.index else None
        if pd.isna(area):
            return 'CAPEX ORDINARIO'
        
        area_str = str(area).strip()
        
        # Para DIR CONSTRUCCIÓN Y PROYECTOS, usar el flag es_extraordinario
        if 'DIR CONSTRUCCIÓN' in area_str and 'PROYECTOS' in area_str:
            if 'es_extraordinario' in row.index:
                return 'CAPEX EXTRAORDINARIO' if row['es_extraordinario'] else 'CAPEX ORDINARIO'
            # Por defecto, si no hay flag, usar la lógica antigua (pero esto no debería pasar)
            return 'CAPEX EXTRAORDINARIO'
        
        return 'CAPEX ORDINARIO'
    
    # Agregar columna de tipo CAPEX (antes de eliminar es_extraordinario)
    tabla_final['tipo_capex'] = tabla_final.apply(asignar_tipo_capex_area, axis=1)
    
    # Eliminar columna es_extraordinario (ya no es necesaria después de asignar tipo_capex)
    if 'es_extraordinario' in tabla_final.columns:
        tabla_final = tabla_final.drop(columns=['es_extraordinario'])
    
    # ===================================================================
    # UNIFICAR ÁREAS DUPLICADAS DESPUÉS DE ASIGNAR TIPO CAPEX
    # ===================================================================
    def unificar_areas_duplicadas(df):
        """Unificar áreas duplicadas dentro del mismo tipo CAPEX"""
        # Agrupar por tipo_capex y procesar cada grupo
        grupos_unificados = []
        
        for tipo_capex in df['tipo_capex'].unique():
            grupo = df[df['tipo_capex'] == tipo_capex].copy()
            
            # Normalizar nombres de áreas para identificar duplicados
            def normalizar_para_unificacion(area):
                """Normalizar nombre de área para identificar duplicados (ignorar acentos)"""
                if pd.isna(area):
                    return area
                area_str = str(area).strip()
                area_upper = area_str.upper()
                
                # Quitar acentos para comparación (normalizar)
                import unicodedata
                area_sin_acentos = ''.join(
                    c for c in unicodedata.normalize('NFD', area_upper)
                    if unicodedata.category(c) != 'Mn'
                )
                
                # DIR CONSTRUCCIÓN Y PROYECTOS / DIR CONSTRUCCION Y PROYECTOS → mismo grupo
                if 'CONSTRUCCION' in area_sin_acentos and 'PROYECTOS' in area_sin_acentos:
                    return 'DIR CONSTRUCCION Y PROYECTOS'
                
                # Dirección de Retail / DIRECCIÓN DE RETAIL → mismo grupo
                if 'DIRECCION' in area_sin_acentos and 'RETAIL' in area_sin_acentos:
                    return 'DIRECCION DE RETAIL'
                
                return area_sin_acentos
            
            # Agregar columna temporal para agrupar
            grupo['area_normalizada'] = grupo['area'].apply(normalizar_para_unificacion)
            
            # Debug: mostrar áreas antes de unificar
            print(f"\n🔍 DEBUG - Antes de unificar en {tipo_capex}:")
            print(f"   Total filas: {len(grupo)}")
            print(f"   Áreas únicas normalizadas: {grupo['area_normalizada'].value_counts().to_dict()}")
            
            # Crear diccionario de nombres estándar antes de agrupar
            nombres_estandar = {}
            for area_norm in grupo['area_normalizada'].unique():
                if area_norm == 'DIR CONSTRUCCION Y PROYECTOS':
                    nombres_estandar[area_norm] = 'DIR CONSTRUCCIÓN Y PROYECTOS'
                elif area_norm == 'DIRECCION DE RETAIL':
                    nombres_estandar[area_norm] = 'Dirección de Retail'
                else:
                    # Para otras áreas, tomar el primer nombre del grupo original (no normalizado)
                    nombres_originales = grupo[grupo['area_normalizada'] == area_norm]['area'].unique()
                    if len(nombres_originales) > 0:
                        # Preferir nombre con acento si está disponible
                        nombre_con_acento = None
                        for nombre in nombres_originales:
                            if 'ó' in str(nombre) or 'Ó' in str(nombre) or 'CONSTRUCCIÓN' in str(nombre):
                                nombre_con_acento = nombre
                                break
                        nombres_estandar[area_norm] = nombre_con_acento if nombre_con_acento else nombres_originales[0]
                    else:
                        nombres_estandar[area_norm] = area_norm
            
            # Debug: mostrar qué se va a unificar
            areas_duplicadas = grupo['area_normalizada'].value_counts()
            areas_duplicadas = areas_duplicadas[areas_duplicadas > 1]
            if len(areas_duplicadas) > 0:
                print(f"   Áreas a unificar: {areas_duplicadas.to_dict()}")
                for area_norm, count in areas_duplicadas.items():
                    areas_originales = grupo[grupo['area_normalizada'] == area_norm]['area'].unique()
                    print(f"      {area_norm} ({count} filas): {list(areas_originales)}")
            
            # Agrupar y sumar valores numéricos
            grupo_unificado = grupo.groupby('area_normalizada', as_index=False).agg({
                col_remanente: 'sum',
                col_presupuesto: 'sum',
                col_ejecutado: 'sum',
                col_diferencia: 'sum',
                'tipo_capex': 'first'
            })
            
            # Agregar columna de área con nombre estándar
            grupo_unificado['area'] = grupo_unificado['area_normalizada'].map(nombres_estandar)
            
            # Eliminar columna temporal antes de reordenar
            grupo_unificado = grupo_unificado.drop(columns=['area_normalizada'])
            
            # Reordenar columnas para que 'area' esté primero
            columnas_ordenadas = ['area', col_remanente, col_presupuesto, col_ejecutado, col_diferencia, 'tipo_capex']
            grupo_unificado = grupo_unificado[columnas_ordenadas]
            
            # Debug: mostrar resultado después de unificar
            print(f"\n✅ DEBUG - Después de unificar en {tipo_capex}:")
            print(f"   Total filas: {len(grupo_unificado)}")
            construccion_filas = grupo_unificado[grupo_unificado['area'].str.contains('CONSTRUCCION', case=False, na=False)]
            retail_filas = grupo_unificado[grupo_unificado['area'].str.contains('RETAIL', case=False, na=False)]
            if len(construccion_filas) > 0:
                print(f"   CONSTRUCCIÓN: {len(construccion_filas)} fila(s) - {construccion_filas['area'].tolist()}")
            if len(retail_filas) > 0:
                print(f"   RETAIL: {len(retail_filas)} fila(s) - {retail_filas['area'].tolist()}")
            
            grupos_unificados.append(grupo_unificado)
        
        # Recombinar todos los grupos
        df_unificado = pd.concat(grupos_unificados, ignore_index=True)
        
        print(f"\n✅ DEBUG - Tabla final unificada:")
        print(f"   Total filas: {len(df_unificado)}")
        construccion_final = df_unificado[df_unificado['area'].str.contains('CONSTRUCCION', case=False, na=False)]
        retail_final = df_unificado[df_unificado['area'].str.contains('RETAIL', case=False, na=False)]
        if len(construccion_final) > 0:
            print(f"   CONSTRUCCIÓN: {len(construccion_final)} fila(s)")
            for idx, row in construccion_final.iterrows():
                print(f"      - {row['area']}: Rem={row[col_remanente]}, Pres={row[col_presupuesto]}, Ejec={row[col_ejecutado]}")
        if len(retail_final) > 0:
            print(f"   RETAIL: {len(retail_final)} fila(s)")
            for idx, row in retail_final.iterrows():
                print(f"      - {row['area']}: Rem={row[col_remanente]}, Pres={row[col_presupuesto]}, Ejec={row[col_ejecutado]}")
        
        return df_unificado
    
    # Aplicar unificación después de asignar tipo CAPEX
    tabla_final = unificar_areas_duplicadas(tabla_final)
    
    # Ordenar: primero EXTRAORDINARIO, luego ORDINARIO
    tabla_final = tabla_final.sort_values('tipo_capex', ascending=False)  # False = EXTRAORDINARIO primero
    
    print(f"\n📊 Tabla final (agrupada por tipo CAPEX):")
    print(tabla_final)
    
    # ===================================================================
    # ESCRIBIR EN EXCEL
    # ===================================================================
    wb = load_workbook(archivo_excel)
    ws = wb['Presupuesto Mensual']
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_font = Font(bold=True, size=11)
    separador_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    separador_font = Font(bold=True, size=11, color="000000")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    fila_actual = 25  # Dejar espacio para tabla 1
    
    # Título
    ws[f'A{fila_actual}'] = "TABLA 2: PRESUPUESTO vs EJECUTADO vs DIFERENCIA"
    ws[f'A{fila_actual}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{fila_actual}'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws.merge_cells(f'A{fila_actual}:E{fila_actual}')
    fila_actual += 2
    
    # Encabezados (con nombres dinámicos)
    headers = ['RESPONSABLE', col_remanente, col_presupuesto, col_ejecutado, col_diferencia]
    for col_idx, header in enumerate(headers, 1):
        celda = ws.cell(row=fila_actual, column=col_idx, value=header)
        celda.fill = header_fill
        celda.font = header_font
        celda.border = border
        celda.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    fila_actual += 1
    
    # Datos agrupados por tipo CAPEX
    tipo_actual = None
    fila_inicio_datos = fila_actual
    
    for idx, row in tabla_final.iterrows():
        area = row['area']
        tipo_capex = row['tipo_capex']
        
        # Si cambió el tipo, agregar separador
        if tipo_actual != tipo_capex:
            if tipo_actual is not None:
                # Agregar fila en blanco entre tipos
                fila_actual += 1
            
            # Fila separadora con el tipo CAPEX
            ws.cell(row=fila_actual, column=1, value=f"--- {tipo_capex} ---").fill = separador_fill
            ws.cell(row=fila_actual, column=1).font = separador_font
            ws.cell(row=fila_actual, column=1).border = border
            ws.merge_cells(f'A{fila_actual}:E{fila_actual}')
            fila_actual += 1
            tipo_actual = tipo_capex
        
        # Escribir fila de datos (sin la columna tipo_capex)
        ws.cell(row=fila_actual, column=1, value=limpiar_valor_para_excel(area)).border = border
        
        # Escribir valores de las columnas (excluyendo tipo_capex)
        valores = [row[col_remanente], row[col_presupuesto], row[col_ejecutado], row[col_diferencia]]
        for col_idx, valor in enumerate(valores, 2):
            celda = ws.cell(row=fila_actual, column=col_idx, value=limpiar_valor_para_excel(valor))
            celda.border = border
            celda.number_format = '#,##0.00'
            celda.alignment = Alignment(horizontal='right')
        
        fila_actual += 1
    
    # Fila de Total
    fila_total = fila_actual
    ws.cell(row=fila_total, column=1, value='TOTAL').fill = total_fill
    ws.cell(row=fila_total, column=1).font = total_font
    ws.cell(row=fila_total, column=1).border = border
    
    for col_idx in range(2, 6):
        celda = ws.cell(row=fila_total, column=col_idx)
        celda.value = f"=SUM({chr(64+col_idx)}{fila_actual}:{chr(64+col_idx)}{fila_total-1})"
        celda.fill = total_fill
        celda.font = total_font
        celda.border = border
        celda.number_format = '#,##0.00'
    
    # Ajustar ancho
    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20
    
    wb.save(archivo_excel)
    print(f"\n✅ Tabla 2 creada en Presupuesto Mensual")
    print(f"="*70)
    
    return tabla_final



def aplicar_estilos_excel_colombia(archivo_path: str):
    """
    Aplicar estilos específicos de Colombia al Excel generado
    (Opcional: puedes personalizar colores, formatos, etc.)
    """
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(archivo_path)
        
        # Aplicar color verde a la hoja BOSQUETO
        if 'BOSQUETO' in wb.sheetnames:
            ws_bosqueto = wb['BOSQUETO']
            ws_bosqueto.sheet_properties.tabColor = "00FF00"  # Verde
            
            # Header con fondo gris
            for cell in ws_bosqueto[1]:
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Aplicar color azul a la hoja DETALLE CORREGIDO
        if 'DETALLE CORREGIDO' in wb.sheetnames:
            ws_detalle = wb['DETALLE CORREGIDO']
            ws_detalle.sheet_properties.tabColor = "0000FF"  # Azul
            
            # Header con fondo gris
            for cell in ws_detalle[1]:
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        wb.save(archivo_path)
        print(f"   ✅ Estilos aplicados al Excel")
        
    except Exception as e:
        print(f"   ⚠️ No se pudieron aplicar estilos: {e}")
        # No es crítico, continuar

def obtener_info_colombia():
    """Obtener información COMPLETA sobre el procesador de Colombia"""
    return {
        'pais': 'Colombia',
        'moneda_principal': 'COP',
        'api_fuente': 'TRM-Colombia jrfajardo',
        'logica_tasa': 'Viernes de semana anterior',  # NUEVO
        'archivo_salida': 'ConsolidadoCapexCOLOMBIA.xlsx',
        'hoja_destino': 'BOSQUETO',
        'columnas_input_principal': 24,
        'columnas_input_adicional': 35,
        'columnas_output': 42,  # ACTUALIZADO
        'estructura': 'Consolidado CAPEX + Análisis + Integración Reporte Absoluto + Tasa Histórica',
        'conversion_moneda': {
            'logica': 'Tasa del viernes de la semana pasada',
            'respaldo': 'Jueves o Miércoles si no hay datos del viernes',
            'ultimo_recurso': 'Tasa actual'
        }
    }

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1:
        archivo = sys.argv[1]
        archivo_abs = sys.argv[2] if len(sys.argv) > 2 else None
        
        print(f"🧪 PROCESANDO: {archivo}")
        if archivo_abs:
            print(f"🧪 CON REPORTE ABSOLUTO: {archivo_abs}")
        
        resultado = procesar_colombia(archivo, archivo_abs)
        
        if resultado:
            print(f"✅ ÉXITO: {resultado}")
        else:
            print(f"❌ FALLÓ EL PROCESAMIENTO")
    else:
        print("🧪 Uso: python colombia.py archivo_pago.xlsx [archivo_absoluto.xlsx]")
