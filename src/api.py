"""
API para cargar datos de Consolidado CAPEX Venezuela a BigQuery
Versión con verificación de duplicados embebida
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
from google.cloud import bigquery
from google.oauth2 import service_account
from google.cloud import storage
import pandas as pd
import hashlib
import os
from datetime import datetime
from zoneinfo import ZoneInfo
from typing import Dict, List
import traceback
import tempfile



# Intentar importar módulos por país
VENEZUELA_MODULE_AVAILABLE = False
COLOMBIA_MODULE_AVAILABLE = False
ARGENTINA_MODULE_AVAILABLE = False

BATCH_SIZE = 10000

GCS_BUCKET_NAME = os.getenv('GCS_BUCKET_NAME')

try:
    from countries.venezuela import (
        procesar_venezuela,
        generar_excel_venezuela_con_detalle,
        agregar_hoja_detalle_al_excel as agregar_hoja_detalle_venezuela,
        crear_hoja_capex_pagado_por_recibo as crear_hoja_capex_venezuela,
        crear_hoja_presupuesto_mensual as crear_hoja_presupuesto_venezuela,
        extraer_tabla2_capex_pagado_recibo as extraer_tabla2_venezuela,
        crear_tabla2_presupuesto_mensual as crear_tabla2_venezuela,
    )
    VENEZUELA_MODULE_AVAILABLE = True
    print("✅ Módulo venezuela.py importado correctamente")
except ImportError as e:
    print(f"⚠️ No se pudo importar venezuela: {e}")

try:
    from countries.colombia import (
        procesar_colombia,
        generar_excel_colombia_con_detalle,
        agregar_hoja_detalle_al_excel as agregar_hoja_detalle_colombia,
        crear_hoja_capex_pagado_por_recibo as crear_hoja_capex_colombia,
        crear_hoja_presupuesto_mensual as crear_hoja_presupuesto_colombia,
        extraer_tabla2_capex_pagado_recibo as extraer_tabla2_colombia,
        crear_tabla2_presupuesto_mensual as crear_tabla2_colombia,
    )
    COLOMBIA_MODULE_AVAILABLE = True
    print("✅ Módulo colombia.py importado correctamente")
except ImportError as e:
    print(f"⚠️ No se pudo importar colombia: {e}")

# try:
#     from countries.argentina import generar_excel_argentina_con_detalle
#     ARGENTINA_MODULE_AVAILABLE = True
#     print("✅ Módulo argentina.py importado correctamente")
# except ImportError as e:
#     print(f"⚠️ No se pudo importar argentina: {e}")

app = Flask(__name__)

# Configurar CORS para permitir solicitudes desde el frontend
CORS(app, resources={
    r"/api/*": {
        "origins": "*",  # En producción, especifica los orígenes permitidos
        "methods": ["GET", "POST", "PUT", "DELETE", "OPTIONS"],
        "allow_headers": ["Content-Type", "Authorization"]
    }
})

# =================== CONFIGURACIÓN ===================

GCP_PROJECT_ID = os.getenv('GCP_PROJECT_ID')
BIGQUERY_DATASET = os.getenv('BIGQUERY_DATASET')
BIGQUERY_DATASET_COP = os.getenv('BIGQUERY_DATASET_COP')
BIGQUERY_TABLE = os.getenv('BIGQUERY_TABLE')
BIGQUERY_TABLE_COP = os.getenv('BIGQUERY_TABLE_COP')
BIGQUERY_TABLE_RESPONSABLE = os.getenv('BIGQUERY_TABLE_RESPONSABLE')
BIGQUERY_TABLE_RESPONSABLE_COP = os.getenv('BIGQUERY_TABLE_RESPONSABLE_COP')
BIGQUERY_TABLE_DIFERENCIA = os.getenv('BIGQUERY_TABLE_DIFERENCIA')
BIGQUERY_TABLE_DIFERENCIA_COP = os.getenv('BIGQUERY_TABLE_DIFERENCIA_COP')
CREDENTIALS_FILE = os.getenv('GOOGLE_APPLICATION_CREDENTIALS')

# =================== CLIENTE BIGQUERY ===================

def crear_cliente_bigquery():
    """Crear cliente de BigQuery con credenciales"""
    try:
        # Si hay un archivo de credenciales especificado y existe, usarlo
        if CREDENTIALS_FILE and os.path.exists(CREDENTIALS_FILE):
            credentials = service_account.Credentials.from_service_account_file(
                CREDENTIALS_FILE,
                scopes=["https://www.googleapis.com/auth/bigquery"]
            )
            client = bigquery.Client(
                credentials=credentials,
                project=GCP_PROJECT_ID
            )
            print(f"✅ Cliente BigQuery creado para proyecto: {GCP_PROJECT_ID} (usando archivo de credenciales)")
            return client
        else:
            # Usar Application Default Credentials (ADC) - funciona en Cloud Run, GCE, etc.
            client = bigquery.Client(project=GCP_PROJECT_ID)
            print(f"✅ Cliente BigQuery creado con ADC para proyecto: {GCP_PROJECT_ID}")
            return client
            
    except Exception as e:
        print(f"❌ Error creando cliente BigQuery: {e}")
        raise

# =================== FUNCIONES DE MAPEO ===================

def generar_id_unico(numero_factura: str, proveedor: str) -> str:
    """
    Generar ID único usando SHA256(numero_factura + proveedor)
    """
    concatenado = f"{str(numero_factura).strip()}{str(proveedor).strip()}"
    hash_obj = hashlib.sha256(concatenado.encode('utf-8'))
    return hash_obj.hexdigest()

def mapear_columnas_bosqueto_a_bigquery_venezuela(df_bosqueto: pd.DataFrame) -> pd.DataFrame:
    """
    Mapear columnas del BOSQUETO Excel a esquema de BigQuery
    """
    print(f"📋 Mapeando {len(df_bosqueto)} filas del BOSQUETO a BigQuery...")
    
    columnas_mapeo = {
        'Numero de Factura': 'vzla_capex_pago_numero_factura',
        'Numero de OC': 'vzla_capex_pago_orden_compra',
        'Tipo Factura': 'vzla_capex_pago_tipo_documento',
        'Nombre Lote': 'vzla_capex_pago_nombre_lote',
        'Proveedor': 'vzla_capex_pago_proveedor',
        'RIF': 'vzla_capex_pago_rif',
        'Fecha Documento': 'vzla_capex_pago_fecha_documento',
        'Tienda': 'vzla_capex_pago_tienda',
        'Sucursal': 'vzla_capex_pago_sucursal',
        'Monto': 'vzla_capex_pago_monto',
        'Moneda': 'vzla_capex_pago_moneda',
        'Fecha Vencimiento': 'vzla_capex_pago_fecha_vencimiento',
        'Cuenta': 'vzla_capex_pago_cuenta',
        'Id Cta': 'vzla_capex_pago_id_cuenta',
        'Método de Pago': 'vzla_capex_pago_metodo_pago',
        'Pago Independiente': 'vzla_capex_pago_es_independiente',
        'Prioridad': 'vzla_capex_pago_prioridad',
        'Monto CAPEX EXT': 'vzla_capex_pago_monto_ext',
        'Monto CAPEX ORD': 'vzla_capex_pago_monto_ord',
        'Monto CADM': 'vzla_capex_pago_monto_cadm',
        'Fecha Creación': 'vzla_capex_pago_fecha_creacion',
        'Solicitante': 'vzla_capex_pago_solicitante',
        'Monto USD': 'vzla_capex_pago_monto_usd',
        'CATEGORIA': 'vzla_capex_pago_categoria',
        'MONTO A PAGAR CAPEX': 'vzla_capex_pago_monto_pagar_capex',
        'MONTO A PAGAR OPEX': 'vzla_capex_pago_monto_pagar_opex',
        'VALIDACION': 'vzla_capex_pago_validacion',
        'METODO DE PAGO': 'vzla_capex_pago_calcu_moneda',
        'SEMANA': 'vzla_capex_pago_semana_pago',
        'MES DE PAGO': 'vzla_capex_pago_mes_pago',
        'TIPO DE CAPEX': 'vzla_capex_pago_tipo_capex',
        'MONTO ORD': 'vzla_capex_pago_calcu_monto_ord',
        'MONTO EXT': 'vzla_capex_pago_calcu_monto_ext',
        'DIA DE PAGO': 'vzla_capex_pago_dia_pago',
        'TIENDA_LOOKUP': 'vzla_capex_pago_calcu_tienda',
        'CECO': 'vzla_capex_pago_ceco',
        'PROYECTO': 'vzla_capex_pago_proyecto',
        'AREA': 'vzla_capex_pago_area',
        'FECHA RECIBO': 'vzla_capex_pago_fecha_recibo',
        'DESCRIPCIÓN': 'vzla_capex_pago_descripcion',
    }
    
    df_mapped = pd.DataFrame()
    
    for col_excel, col_bq in columnas_mapeo.items():
        if col_excel in df_bosqueto.columns:
            df_mapped[col_bq] = df_bosqueto[col_excel]
        else:
            print(f"⚠️ Columna '{col_excel}' no encontrada")
            df_mapped[col_bq] = None
    
    # Generar ID único
    print("🔐 Generando IDs únicos con SHA256...")
    df_mapped['vzla_capex_pago_id'] = df_mapped.apply(
        lambda row: generar_id_unico(
            row['vzla_capex_pago_numero_factura'],
            row['vzla_capex_pago_proveedor']
        ),
        axis=1
    )
    
    # Procesar AÑO FISCAL
    if 'AÑO FISCAL' in df_bosqueto.columns:
        anio_fiscal_str = df_bosqueto['AÑO FISCAL'].iloc[0] if len(df_bosqueto) > 0 else "2025-2026"
        try:
            if '-' in str(anio_fiscal_str):
                partes = str(anio_fiscal_str).split('-')
                df_mapped['vzla_capex_pago_current_fiscal_year'] = int(partes[0])
                df_mapped['vzla_capex_pago_next_fiscal_year'] = int(partes[1])
            else:
                anio_actual = datetime.now().year
                df_mapped['vzla_capex_pago_current_fiscal_year'] = anio_actual
                df_mapped['vzla_capex_pago_next_fiscal_year'] = anio_actual + 1
        except:
            anio_actual = datetime.now().year
            df_mapped['vzla_capex_pago_current_fiscal_year'] = anio_actual
            df_mapped['vzla_capex_pago_next_fiscal_year'] = anio_actual + 1
    else:
        anio_actual = datetime.now().year
        df_mapped['vzla_capex_pago_current_fiscal_year'] = anio_actual
        df_mapped['vzla_capex_pago_next_fiscal_year'] = anio_actual + 1
    
    # Agregar columna de país
    df_mapped['vzla_capex_pago_pais'] = 'Venezuela'
    
    # Convertir fechas a datetime
    columnas_fecha = [
        'vzla_capex_pago_fecha_documento',
        'vzla_capex_pago_fecha_vencimiento',
        'vzla_capex_pago_fecha_creacion',
        'vzla_capex_pago_fecha_recibo'
    ]
    
    for col_fecha in columnas_fecha:
        if col_fecha in df_mapped.columns:
            # Especificar formato y usar utc=False para evitar warnings
            df_mapped[col_fecha] = pd.to_datetime(
            df_mapped[col_fecha],
            format='mixed', # Permite múltiples formatos
            errors='coerce',
            utc=False
            )

    # Convertir prioridad a INTEGER
    if 'vzla_capex_pago_prioridad' in df_mapped.columns:
        df_mapped['vzla_capex_pago_prioridad'] = pd.to_numeric(
            df_mapped['vzla_capex_pago_prioridad'], 
            errors='coerce'
        ).fillna(0).astype(int)
    
    # Convertir columnas numéricas a FLOAT
    columnas_float = [
        'vzla_capex_pago_monto',
        'vzla_capex_pago_monto_ext',
        'vzla_capex_pago_monto_ord',
        'vzla_capex_pago_monto_cadm',
        'vzla_capex_pago_monto_usd',
        'vzla_capex_pago_monto_pagar_capex',
        'vzla_capex_pago_monto_pagar_opex',
        'vzla_capex_pago_calcu_monto_ord',
        'vzla_capex_pago_calcu_monto_ext'
    ]
    
    for col_float in columnas_float:
        if col_float in df_mapped.columns:
            df_mapped[col_float] = pd.to_numeric(df_mapped[col_float], errors='coerce')
    
    print(f"✅ Mapeo completado: {len(df_mapped)} filas, {len(df_mapped.columns)} columnas")
    
    return df_mapped


def mapear_columnas_bosqueto_a_bigquery_colombia(df_bosqueto: pd.DataFrame) -> pd.DataFrame:
    """
    Mapear columnas del BOSQUETO Excel a esquema de BigQuery
    """
    print(f"📋 Mapeando {len(df_bosqueto)} filas del BOSQUETO a BigQuery...")
    
    columnas_mapeo = {
        'Numero de Factura': 'col_capex_pago_numero_factura',
        'Numero de OC': 'col_capex_pago_orden_compra',
        'Tipo Factura': 'col_capex_pago_tipo_documento',
        'Nombre Lote': 'col_capex_pago_nombre_lote',
        'Proveedor': 'col_capex_pago_proveedor',
        'RIF': 'col_capex_pago_rif',
        'Fecha Documento': 'col_capex_pago_fecha_documento',
        'Tienda': 'col_capex_pago_tienda',
        'Sucursal': 'col_capex_pago_sucursal',
        'Monto': 'col_capex_pago_monto',
        'Moneda': 'col_capex_pago_moneda',
        'Fecha Vencimiento': 'col_capex_pago_fecha_vencimiento',
        'Cuenta': 'col_capex_pago_cuenta',
        'Id Cta': 'col_capex_pago_id_cuenta',
        'Método de Pago': 'col_capex_pago_metodo_pago',
        'Pago Independiente': 'col_capex_pago_es_independiente',
        'Prioridad': 'col_capex_pago_prioridad',
        'Monto CAPEX EXT': 'col_capex_pago_monto_ext',
        'Monto CAPEX ORD': 'col_capex_pago_monto_ord',
        'Monto CADM': 'col_capex_pago_monto_cadm',
        'Fecha Creación': 'col_capex_pago_fecha_creacion',
        'Solicitante': 'col_capex_pago_solicitante',
        'Monto USD': 'col_capex_pago_monto_usd',
        'CATEGORIA': 'col_capex_pago_categoria',
        'MONTO A PAGAR CAPEX': 'col_capex_pago_monto_pagar_capex',
        'MONTO A PAGAR OPEX': 'col_capex_pago_monto_pagar_opex',
        'VALIDACION': 'col_capex_pago_validacion',
        'METODO DE PAGO': 'col_capex_pago_calcu_moneda',
        'SEMANA': 'col_capex_pago_semana_pago',
        'MES DE PAGO': 'col_capex_pago_mes_pago',
        'TIPO DE CAPEX': 'col_capex_pago_tipo_capex',
        'MONTO ORD': 'col_capex_pago_calcu_monto_ord',
        'MONTO EXT': 'col_capex_pago_calcu_monto_ext',
        'DIA DE PAGO': 'col_capex_pago_dia_pago',
        'TIENDA_LOOKUP': 'col_capex_pago_calcu_tienda',
        'CECO': 'col_capex_pago_ceco',
        'PROYECTO': 'col_capex_pago_proyecto',
        'AREA': 'col_capex_pago_area',
        'FECHA RECIBO': 'col_capex_pago_fecha_recibo',
        'DESCRIPCIÓN': 'col_capex_pago_descripcion',
    }
    
    df_mapped = pd.DataFrame()
    
    for col_excel, col_bq in columnas_mapeo.items():
        if col_excel in df_bosqueto.columns:
            df_mapped[col_bq] = df_bosqueto[col_excel]
        else:
            print(f"⚠️ Columna '{col_excel}' no encontrada")
            df_mapped[col_bq] = None
    
    # Generar ID único
    print("🔐 Generando IDs únicos con SHA256...")
    df_mapped['col_capex_pago_id'] = df_mapped.apply(
        lambda row: generar_id_unico(
            row['col_capex_pago_numero_factura'],
            row['col_capex_pago_proveedor']
        ),
        axis=1
    )
    
    # Procesar AÑO FISCAL
    if 'AÑO FISCAL' in df_bosqueto.columns:
        anio_fiscal_str = df_bosqueto['AÑO FISCAL'].iloc[0] if len(df_bosqueto) > 0 else "2025-2026"
        try:
            if '-' in str(anio_fiscal_str):
                partes = str(anio_fiscal_str).split('-')
                df_mapped['col_capex_pago_current_fiscal_year'] = int(partes[0])
                df_mapped['col_capex_pago_next_fiscal_year'] = int(partes[1])
            else:
                anio_actual = datetime.now().year
                df_mapped['col_capex_pago_current_fiscal_year'] = anio_actual
                df_mapped['col_capex_pago_next_fiscal_year'] = anio_actual + 1
        except:
            anio_actual = datetime.now().year
            df_mapped['col_capex_pago_current_fiscal_year'] = anio_actual
            df_mapped['col_capex_pago_next_fiscal_year'] = anio_actual + 1
    else:
        anio_actual = datetime.now().year
        df_mapped['col_capex_pago_current_fiscal_year'] = anio_actual
        df_mapped['col_capex_pago_next_fiscal_year'] = anio_actual + 1
    
    # Agregar columna de país
    df_mapped['col_capex_pago_pais'] = 'Colombia'
    
    # Convertir fechas a datetime
    columnas_fecha = [
        'col_capex_pago_fecha_documento',
        'col_capex_pago_fecha_vencimiento',
        'col_capex_pago_fecha_creacion',
        'col_capex_pago_fecha_recibo'
    ]
    
    for col_fecha in columnas_fecha:
        if col_fecha in df_mapped.columns:
            # Especificar formato y usar utc=False para evitar warnings
            df_mapped[col_fecha] = pd.to_datetime(
            df_mapped[col_fecha],
            format='mixed', # Permite múltiples formatos
            errors='coerce',
            utc=False
            )

    # Convertir prioridad a INTEGER
    if 'col_capex_pago_prioridad' in df_mapped.columns:
        df_mapped['col_capex_pago_prioridad'] = pd.to_numeric(
            df_mapped['col_capex_pago_prioridad'], 
            errors='coerce'
        ).fillna(0).astype(int)
    
    # Convertir columnas numéricas a FLOAT
    columnas_float = [
        'col_capex_pago_monto',
        'col_capex_pago_monto_ext',
        'col_capex_pago_monto_ord',
        'col_capex_pago_monto_cadm',
        'col_capex_pago_monto_usd',
        'col_capex_pago_monto_pagar_capex',
        'col_capex_pago_monto_pagar_opex',
        'col_capex_pago_calcu_monto_ord',
        'col_capex_pago_calcu_monto_ext'
    ]
    
    for col_float in columnas_float:
        if col_float in df_mapped.columns:
            df_mapped[col_float] = pd.to_numeric(df_mapped[col_float], errors='coerce')
    
    print(f"✅ Mapeo completado: {len(df_mapped)} filas, {len(df_mapped.columns)} columnas")
    
    return df_mapped

# =================== VERIFICACIÓN DE DUPLICADOS ===================

def verificar_duplicados_batch_venezuela(client: bigquery.Client, ids_a_verificar: List[str]) -> Dict[str, bool]:
    """
    Verificar duplicados en batch usando query
    """
    if not ids_a_verificar:
        return {}
    
    print(f"🔍 Verificando {len(ids_a_verificar)} IDs en BigQuery...")
    
    batch_size = 1000
    resultados = {}
    
    for i in range(0, len(ids_a_verificar), batch_size):
        batch = ids_a_verificar[i:i+batch_size]
        ids_str = "', '".join(batch)
        
        query = f"""
        SELECT vzla_capex_pago_id
        FROM `{GCP_PROJECT_ID}.{BIGQUERY_DATASET}.{BIGQUERY_TABLE}`
        WHERE vzla_capex_pago_id IN ('{ids_str}')
        """
        
        try:
            query_job = client.query(query)
            results = query_job.result()
            
            for row in results:
                resultados[row.vzla_capex_pago_id] = True
            
        except Exception as e:
            print(f"⚠️ Error en batch {i//batch_size + 1}: {e}")
    
    for id_check in ids_a_verificar:
        if id_check not in resultados:
            resultados[id_check] = False
    
    duplicados_count = sum(1 for existe in resultados.values() if existe)
    print(f"📊 Resultado: {duplicados_count} duplicados, {len(ids_a_verificar) - duplicados_count} nuevos")
    
    return resultados

def verificar_duplicados_batch_colombia(client: bigquery.Client, ids_a_verificar: List[str]) -> Dict[str, bool]:
    """
    Verificar duplicados en batch usando query
    """
    if not ids_a_verificar:
        return {}
    
    print(f"🔍 Verificando {len(ids_a_verificar)} IDs en BigQuery...")
    
    batch_size = 1000
    resultados = {}
    
    for i in range(0, len(ids_a_verificar), batch_size):
        batch = ids_a_verificar[i:i+batch_size]
        ids_str = "', '".join(batch)
        
        query = f"""
        SELECT col_capex_pago_id
        FROM `{GCP_PROJECT_ID}.{BIGQUERY_DATASET_COP}.{BIGQUERY_TABLE_COP}`
        WHERE col_capex_pago_id IN ('{ids_str}')
        """
        
        try:
            query_job = client.query(query)
            results = query_job.result()
            
            for row in results:
                resultados[row.col_capex_pago_id] = True
            
        except Exception as e:
            print(f"⚠️ Error en batch {i//batch_size + 1}: {e}")
    
    for id_check in ids_a_verificar:
        if id_check not in resultados:
            resultados[id_check] = False
    
    duplicados_count = sum(1 for existe in resultados.values() if existe)
    print(f"📊 Resultado: {duplicados_count} duplicados, {len(ids_a_verificar) - duplicados_count} nuevos")
    
    return resultados

# =================== ROUTER DE GENERACIÓN DE EXCEL POR PAÍS ===================

def generar_excel_generico(df_bosqueto: pd.DataFrame, df_detalle: pd.DataFrame) -> str:
    """
    Generador genérico de Excel (fallback cuando no hay módulo específico del país)
    """
    print(f"📝 Generando Excel genérico (fallback)...")
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_path = temp_file.name
    temp_file.close()
    
    try:
        with pd.ExcelWriter(temp_path, engine='openpyxl') as writer:
            df_bosqueto.to_excel(writer, sheet_name='BOSQUETO', index=False)
            print(f"   ✅ Hoja 'BOSQUETO' creada: {len(df_bosqueto)} filas")
            
            if not df_detalle.empty:
                df_detalle.to_excel(writer, sheet_name='DETALLE CORREGIDO', index=False)
                print(f"   ✅ Hoja 'DETALLE CORREGIDO' creada: {len(df_detalle)} filas")
            else:
                pd.DataFrame(columns=df_bosqueto.columns).to_excel(
                    writer, sheet_name='DETALLE CORREGIDO', index=False
                )
                print(f"   ⚠️ Hoja 'DETALLE CORREGIDO' vacía")
        
        print(f"✅ Excel genérico generado: {temp_path}")
        return temp_path
        
    except Exception as e:
        print(f"❌ Error generando Excel genérico: {e}")
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise

def generar_excel_consolidado(df_bosqueto: pd.DataFrame, 
                               df_detalle: pd.DataFrame,
                               pais: str = 'venezuela') -> str:
    """
    Router principal para generar Excel según el país
    
    Args:
        df_bosqueto: DataFrame con BOSQUETO original
        df_detalle: DataFrame con DETALLE CORREGIDO (solo registros cargados)
        pais: País del consolidado ('venezuela', 'colombia', 'argentina')
    
    Returns:
        str: Ruta del archivo Excel generado
    """
    pais_lower = pais.lower().strip()
    
    print(f"🌎 Generando Excel para país: {pais_lower.upper()}")
    
    # Venezuela
    if pais_lower == 'venezuela':
        if VENEZUELA_MODULE_AVAILABLE:
            print(f"   ✅ Usando generador específico de Venezuela")
            return agregar_hoja_detalle_venezuela(df_bosqueto, df_detalle)
        else:
            print(f"   ⚠️ Módulo Venezuela no disponible, usando genérico")
            return generar_excel_generico(df_bosqueto, df_detalle)

    # Colombia
    if pais_lower == 'colombia':
        if COLOMBIA_MODULE_AVAILABLE:
            print(f"   ✅ Usando generador específico de Colombia")
            return agregar_hoja_detalle_colombia(df_bosqueto, df_detalle)
        else:
            print(f"   ⚠️ Módulo Venezuela no disponible, usando genérico")
            return generar_excel_generico(df_bosqueto, df_detalle)

    # # Argentina
    # elif pais_lower == 'argentina':
    #     if ARGENTINA_MODULE_AVAILABLE:
    #         print(f"   ✅ Usando generador específico de Argentina")
    #         return generar_excel_argentina_con_detalle(df_bosqueto, df_detalle)
    #     else:
    #         print(f"   ⚠️ Módulo Argentina no disponible, usando genérico")
    #         return generar_excel_generico(df_bosqueto, df_detalle)
    
    # País no reconocido o genérico
    else:
        print(f"   ⚠️ País '{pais}' no reconocido, usando generador genérico")
        return generar_excel_generico(df_bosqueto, df_detalle)

# =================== CARGA A BIGQUERY ===================
def cargar_datos_a_bigquery_venezuela(client: bigquery.Client, df: pd.DataFrame) -> dict:
    """
    Cargar datos a BigQuery con verificación automática de duplicados.
    Retorna un dict con todo lo necesario.
    """
    table_id = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET}.{BIGQUERY_TABLE}"
    print(f"📤 Cargando datos a BigQuery: {table_id}")
    print(f"   Total de registros: {len(df)}")

    # PASO 1: Verificar duplicados (embebido)
    ids_a_verificar = df['vzla_capex_pago_id'].tolist()
    duplicados_map = verificar_duplicados_batch_venezuela(client, ids_a_verificar)
    df_nuevos = df[df['vzla_capex_pago_id'].apply(lambda x: not duplicados_map.get(x, False))]
    registros_duplicados = len(df) - len(df_nuevos)

    print(f"   🔄 Duplicados omitidos: {registros_duplicados}")
    print(f"   ✅ Registros nuevos a cargar: {len(df_nuevos)}")

    result = {
        'success': True,
        'total_rows': len(df),
        'rows_loaded': len(df_nuevos),
        'rows_duplicated': registros_duplicados,
        'table_id': table_id,
        'message': '',
        'df_cargados': None
    }

    if len(df_nuevos) == 0:
        result['message'] = 'No hay registros nuevos para cargar (todos son duplicados)'
        result['df_cargados'] = pd.DataFrame()
        return result

    job_config = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
        schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION]
    )

    try:
        df_nuevos = ajustar_df_a_schema_bigquery_venezuela(df_nuevos, client, BIGQUERY_DATASET, BIGQUERY_TABLE)
        
        print(f"⏳ Iniciando carga a BigQuery ({len(df_nuevos)} filas)...", flush=True)
        job = client.load_table_from_dataframe(df_nuevos, table_id, job_config=job_config)
        print(f"⏳ Job creado, esperando resultado (timeout: 300s)...", flush=True)
        job.result(timeout=300)  # Timeout de 5 minutos

        print(f"✅ Carga completada exitosamente", flush=True)
        print(f"   📊 Filas cargadas: {len(df_nuevos)}", flush=True)

        result['message'] = f'Carga exitosa: {len(df_nuevos)} registros nuevos, {registros_duplicados} duplicados omitidos'
        result['df_cargados'] = df_nuevos
        return result

    except Exception as e:
        print(f"❌ Error cargando datos: {e}")
        traceback.print_exc()
        result['success'] = False
        result['error'] = str(e)
        result['message'] = f'Error en carga: {str(e)}'
        result['df_cargados'] = pd.DataFrame()
        return result

def cargar_datos_a_bigquery_colombia(client: bigquery.Client, df: pd.DataFrame) -> dict:
    """
    Cargar datos a BigQuery con verificación automática de duplicados.
    Retorna un dict con todo lo necesario.
    """
    table_id = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET_COP}.{BIGQUERY_TABLE_COP}"
    print(f"📤 Cargando datos a BigQuery: {table_id}")
    print(f"   Total de registros: {len(df)}")

    # PASO 1: Verificar duplicados (embebido)
    ids_a_verificar = df['col_capex_pago_id'].tolist()
    duplicados_map = verificar_duplicados_batch_colombia(client, ids_a_verificar)
    df_nuevos = df[df['col_capex_pago_id'].apply(lambda x: not duplicados_map.get(x, False))]
    registros_duplicados = len(df) - len(df_nuevos)

    print(f"   🔄 Duplicados omitidos: {registros_duplicados}")
    print(f"   ✅ Registros nuevos a cargar: {len(df_nuevos)}")

    result = {
        'success': True,
        'total_rows': len(df),
        'rows_loaded': len(df_nuevos),
        'rows_duplicated': registros_duplicados,
        'table_id': table_id,
        'message': '',
        'df_cargados': None
    }

    if len(df_nuevos) == 0:
        result['message'] = 'No hay registros nuevos para cargar (todos son duplicados)'
        result['df_cargados'] = pd.DataFrame()
        return result

    job_config = bigquery.LoadJobConfig(
        write_disposition=bigquery.WriteDisposition.WRITE_APPEND,
        schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION]
    )

    try:
        df_nuevos = ajustar_df_a_schema_bigquery_colombia(df_nuevos, client, BIGQUERY_DATASET_COP, BIGQUERY_TABLE_COP)
        
        print(f"⏳ Iniciando carga a BigQuery ({len(df_nuevos)} filas)...", flush=True)
        job = client.load_table_from_dataframe(df_nuevos, table_id, job_config=job_config)
        print(f"⏳ Job creado, esperando resultado (timeout: 300s)...", flush=True)
        job.result(timeout=300)  # Timeout de 5 minutos

        print(f"✅ Carga completada exitosamente", flush=True)
        print(f"   📊 Filas cargadas: {len(df_nuevos)}", flush=True)

        result['message'] = f'Carga exitosa: {len(df_nuevos)} registros nuevos, {registros_duplicados} duplicados omitidos'
        result['df_cargados'] = df_nuevos
        return result

    except Exception as e:
        print(f"❌ Error cargando datos: {e}")
        traceback.print_exc()
        result['success'] = False
        result['error'] = str(e)
        result['message'] = f'Error en carga: {str(e)}'
        result['df_cargados'] = pd.DataFrame()
        return result

# =================== EXTRACCIÓN DESDE BIGQUERY POR LOTES ===================

def extraer_tabla_completa_por_lotes_venezuela(client: bigquery.Client) -> pd.DataFrame:
    """
    Extraer toda la tabla de BigQuery por lotes para evitar timeouts
    """
    table_id = f"`{GCP_PROJECT_ID}.{BIGQUERY_DATASET}.{BIGQUERY_TABLE}`"
    
    print(f"📊 Extrayendo datos de BigQuery por lotes...", flush=True)
    
    # Primero, contar cuántas filas hay
    print(f"   Contando filas...", flush=True)
    count_query = f"SELECT COUNT(*) as total FROM {table_id}"
    count_job = client.query(count_query)
    count_result = count_job.result(timeout=120)  # Timeout 2 min
    total_rows = list(count_result)[0].total
    
    print(f"   Total de filas en la tabla: {total_rows}", flush=True)
    
    if total_rows == 0:
        print("⚠️ La tabla está vacía", flush=True)
        return pd.DataFrame()
    
    # Extraer por lotes usando LIMIT y OFFSET
    all_dataframes = []
    offset = 0
    
    while offset < total_rows:
        print(f"   Extrayendo lote: filas {offset} a {offset + BATCH_SIZE}...", flush=True)
        
        query = f"""
        SELECT *
        FROM {table_id}
        ORDER BY vzla_capex_pago_id
        LIMIT {BATCH_SIZE}
        OFFSET {offset}
        """
        
        try:
            query_job = client.query(query)
            df_batch = query_job.result(timeout=300).to_dataframe()  # Timeout 5 min por lote
            all_dataframes.append(df_batch)
            offset += BATCH_SIZE
            
            print(f"   ✅ Lote extraído: {len(df_batch)} filas", flush=True)
            
        except Exception as e:
            print(f"   ❌ Error extrayendo lote: {e}", flush=True)
            traceback.print_exc()
            break
    
    if not all_dataframes:
        return pd.DataFrame()
    
    # Combinar todos los lotes
    df_completo = pd.concat(all_dataframes, ignore_index=True)
    print(f"✅ Extracción completa: {len(df_completo)} filas", flush=True)
    
    return df_completo


def extraer_tabla_completa_por_lotes_colombia(client: bigquery.Client) -> pd.DataFrame:
    """
    Extraer toda la tabla de BigQuery por lotes para evitar timeouts
    """
    table_id = f"`{GCP_PROJECT_ID}.{BIGQUERY_DATASET_COP}.{BIGQUERY_TABLE_COP}`"
    
    print(f"📊 Extrayendo datos de BigQuery por lotes...", flush=True)
    
    # Primero, contar cuántas filas hay
    print(f"   Contando filas...", flush=True)
    count_query = f"SELECT COUNT(*) as total FROM {table_id}"
    count_job = client.query(count_query)
    count_result = count_job.result(timeout=120)  # Timeout 2 min
    total_rows = list(count_result)[0].total
    
    print(f"   Total de filas en la tabla: {total_rows}", flush=True)
    
    if total_rows == 0:
        print("⚠️ La tabla está vacía", flush=True)
        return pd.DataFrame()
    
    # Extraer por lotes usando LIMIT y OFFSET
    all_dataframes = []
    offset = 0
    
    while offset < total_rows:
        print(f"   Extrayendo lote: filas {offset} a {offset + BATCH_SIZE}...", flush=True)
        
        query = f"""
        SELECT *
        FROM {table_id}
        ORDER BY col_capex_pago_id
        LIMIT {BATCH_SIZE}
        OFFSET {offset}
        """
        
        try:
            query_job = client.query(query)
            df_batch = query_job.result(timeout=300).to_dataframe()  # Timeout 5 min por lote
            all_dataframes.append(df_batch)
            offset += BATCH_SIZE
            
            print(f"   ✅ Lote extraído: {len(df_batch)} filas", flush=True)
            
        except Exception as e:
            print(f"   ❌ Error extrayendo lote: {e}", flush=True)
            traceback.print_exc()
            break
    
    if not all_dataframes:
        return pd.DataFrame()
    
    # Combinar todos los lotes
    df_completo = pd.concat(all_dataframes, ignore_index=True)
    print(f"✅ Extracción completa: {len(df_completo)} filas", flush=True)
    
    return df_completo

def extraer_responsables_capex_venezuela(bq_client, anio_fiscal: str = None) -> pd.DataFrame:
    """
    Extraer datos de la tabla vzla_capex_pago_responsable de BigQuery
    
    Args:
        anio_fiscal: Ej: "2025-2026" (si es None, usa el actual)
    """
    from google.cloud import bigquery
    from datetime import datetime
    
    print(f"\n📊 Extrayendo datos de vzla_capex_pago_responsable...")
    
    # Calcular año fiscal actual si no se proporciona
    if not anio_fiscal:
        hoy = datetime.now()
        if hoy.month >= 8:  # Agosto o después
            anio_inicio = hoy.year
            anio_fin = hoy.year + 1
        else:
            anio_inicio = hoy.year - 1
            anio_fin = hoy.year
        anio_fiscal = f"{anio_inicio}-{anio_fin}"
    
    print(f"   📅 Año fiscal: {anio_fiscal}")
    
    # Calcular rango de fechas del año fiscal (Agosto a Julio)
    anio_inicio_int = int(anio_fiscal.split('-')[0])
    fecha_inicio = f"{anio_inicio_int}-08-01"
    fecha_fin = f"{anio_inicio_int + 1}-07-31"
    
    print(f"   📅 Rango de fechas: {fecha_inicio} a {fecha_fin}")
    
    table_id_responsable = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET}.{BIGQUERY_TABLE_RESPONSABLE}"
    query = f"""
    SELECT
        vzla_capex_responsable_anio_fiscal,
        vzla_capex_responsable_fecha,
        vzla_capex_responsable_tipo,
        vzla_capex_responsable_area,
        vzla_capex_responsable_monto
    FROM `{table_id_responsable}`
    WHERE vzla_capex_responsable_fecha BETWEEN '{fecha_inicio}' AND '{fecha_fin}'
      AND vzla_capex_responsable_anio_fiscal = '{anio_fiscal}'
    ORDER BY vzla_capex_responsable_fecha
    """
    
    try:
        df_responsables = bq_client.query(query).to_dataframe()
        print(f"✅ {len(df_responsables)} registros extraídos")
        
        if not df_responsables.empty:
            print(f"\n📋 Columnas: {list(df_responsables.columns)}")
            print(f"📋 Áreas únicas: {df_responsables['vzla_capex_responsable_area'].nunique()}")
            print(f"📋 Tipos CAPEX: {df_responsables['vzla_capex_responsable_tipo'].unique()}")
            print(f"\n📊 Muestra de datos:")
            print(df_responsables.head())
        
        return df_responsables
        
    except Exception as e:
        print(f"❌ Error extrayendo datos: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()

def extraer_diferencia_capex_venezuela(bq_client, anio_fiscal: str = None) -> pd.DataFrame:
    """
    Extraer datos de vzla_capex_pago_diferencia (Presupuesto + Remanente)
    Particionada por vzla_capex_diferencia_fecha_ejecucion
    Clustered by vzla_capex_diferencia_area
    """
    
    print(f"\n📊 Extrayendo datos de vzla_capex_pago_diferencia...")
    
    # Calcular año fiscal actual si no se proporciona
    if not anio_fiscal:
        hoy = datetime.now()
        if hoy.month >= 8:
            anio_inicio = hoy.year
            anio_fin = hoy.year + 1
        else:
            anio_inicio = hoy.year - 1
            anio_fin = hoy.year
        anio_fiscal = f"{anio_inicio}-{anio_fin}"
    
    print(f"   📅 Año fiscal: {anio_fiscal}")
    
    # Calcular rango de fechas del año fiscal (Agosto a Julio)
    anio_inicio_int = int(anio_fiscal.split('-')[0])
    fecha_inicio = f"{anio_inicio_int}-08-01"
    fecha_fin = f"{anio_inicio_int + 1}-07-31"
    
    print(f"   📅 Rango de fechas: {fecha_inicio} a {fecha_fin}")
    
    table_id_diferencia = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET}.{BIGQUERY_TABLE_DIFERENCIA}"
    query = f"""
    WITH datos_recientes AS (
        SELECT
            vzla_capex_diferencia_mes,
            vzla_capex_diferencia_tipo,
            vzla_capex_diferencia_area,
            vzla_capex_diferencia_remanente,
            vzla_capex_diferencia_presupuesto,
            vzla_capex_diferencia_fecha_ejecucion,
            ROW_NUMBER() OVER (
                PARTITION BY vzla_capex_diferencia_area, vzla_capex_diferencia_tipo, vzla_capex_diferencia_mes
                ORDER BY vzla_capex_diferencia_fecha_ejecucion DESC
            ) as rn
        FROM `{table_id_diferencia}`
        WHERE vzla_capex_diferencia_fecha_ejecucion BETWEEN '{fecha_inicio}' AND '{fecha_fin}'
    )
    SELECT
        vzla_capex_diferencia_mes,
        vzla_capex_diferencia_tipo,
        vzla_capex_diferencia_area,
        vzla_capex_diferencia_remanente,
        vzla_capex_diferencia_presupuesto,
        vzla_capex_diferencia_fecha_ejecucion
    FROM datos_recientes
    WHERE rn = 1
    ORDER BY vzla_capex_diferencia_area, vzla_capex_diferencia_tipo
    """
    
    try:
        df_diferencia = bq_client.query(query).to_dataframe()
        print(f"✅ {len(df_diferencia)} registros extraídos")
        
        if not df_diferencia.empty:
            print(f"\n📋 Áreas: {df_diferencia['vzla_capex_diferencia_area'].nunique()}")
            print(f"📋 Tipos: {df_diferencia['vzla_capex_diferencia_tipo'].unique()}")
            print(f"\n📊 Muestra:")
            print(df_diferencia.head(10))
        
        return df_diferencia
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()


def extraer_responsables_capex_colombia(bq_client, anio_fiscal: str = None) -> pd.DataFrame:
    """
    Extraer datos de la tabla col_capex_pago_responsable de BigQuery
    
    Args:
        anio_fiscal: Ej: "2025-2026" (si es None, usa el actual)
    """
    from google.cloud import bigquery
    from datetime import datetime
    
    print(f"\n📊 Extrayendo datos de col_capex_pago_responsable...")
    
    # Calcular año fiscal actual si no se proporciona
    if not anio_fiscal:
        hoy = datetime.now()
        if hoy.month >= 8:  # Agosto o después
            anio_inicio = hoy.year
            anio_fin = hoy.year + 1
        else:
            anio_inicio = hoy.year - 1
            anio_fin = hoy.year
        anio_fiscal = f"{anio_inicio}-{anio_fin}"
    
    print(f"   📅 Año fiscal: {anio_fiscal}")
    
    # Calcular rango de fechas del año fiscal (Agosto a Julio)
    anio_inicio_int = int(anio_fiscal.split('-')[0])
    fecha_inicio = f"{anio_inicio_int}-08-01"
    fecha_fin = f"{anio_inicio_int + 1}-07-31"
    
    print(f"   📅 Rango de fechas: {fecha_inicio} a {fecha_fin}")
    
    table_id_responsable = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET_COP}.{BIGQUERY_TABLE_RESPONSABLE_COP}"
    query = f"""
    SELECT
        col_capex_responsable_anio_fiscal,
        col_capex_responsable_fecha,
        col_capex_responsable_tipo,
        col_capex_responsable_area,
        col_capex_responsable_monto
    FROM `{table_id_responsable}`
    WHERE col_capex_responsable_fecha BETWEEN '{fecha_inicio}' AND '{fecha_fin}'
      AND col_capex_responsable_anio_fiscal = '{anio_fiscal}'
    ORDER BY col_capex_responsable_fecha
    """
    
    try:
        df_responsables = bq_client.query(query).to_dataframe()
        print(f"✅ {len(df_responsables)} registros extraídos")
        
        if not df_responsables.empty:
            print(f"\n📋 Columnas: {list(df_responsables.columns)}")
            print(f"📋 Áreas únicas: {df_responsables['col_capex_responsable_area'].nunique()}")
            print(f"📋 Tipos CAPEX: {df_responsables['col_capex_responsable_tipo'].unique()}")
            print(f"\n📊 Muestra de datos:")
            print(df_responsables.head())
        
        return df_responsables
        
    except Exception as e:
        print(f"❌ Error extrayendo datos: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()

def extraer_diferencia_capex_colombia(bq_client, anio_fiscal: str = None) -> pd.DataFrame:
    """
    Extraer datos de col_capex_pago_diferencia (Presupuesto + Remanente)
    Particionada por col_capex_diferencia_fecha_ejecucion
    Clustered by col_capex_diferencia_area
    """
    
    print(f"\n📊 Extrayendo datos de col_capex_pago_diferencia...")
    
    # Calcular año fiscal actual si no se proporciona
    if not anio_fiscal:
        hoy = datetime.now()
        if hoy.month >= 8:
            anio_inicio = hoy.year
            anio_fin = hoy.year + 1
        else:
            anio_inicio = hoy.year - 1
            anio_fin = hoy.year
        anio_fiscal = f"{anio_inicio}-{anio_fin}"
    
    print(f"   📅 Año fiscal: {anio_fiscal}")
    
    # Calcular rango de fechas del año fiscal (Agosto a Julio)
    anio_inicio_int = int(anio_fiscal.split('-')[0])
    fecha_inicio = f"{anio_inicio_int}-08-01"
    fecha_fin = f"{anio_inicio_int + 1}-07-31"
    
    print(f"   📅 Rango de fechas: {fecha_inicio} a {fecha_fin}")
    
    table_id_diferencia = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET_COP}.{BIGQUERY_TABLE_DIFERENCIA_COP}"
    query = f"""
    WITH datos_recientes AS (
        SELECT
            col_capex_diferencia_mes,
            col_capex_diferencia_tipo,
            col_capex_diferencia_area,
            col_capex_diferencia_remanente,
            col_capex_diferencia_presupuesto,
            col_capex_diferencia_fecha_ejecucion,
            ROW_NUMBER() OVER (
                PARTITION BY col_capex_diferencia_area, col_capex_diferencia_tipo, col_capex_diferencia_mes
                ORDER BY col_capex_diferencia_fecha_ejecucion DESC
            ) as rn
        FROM `{table_id_diferencia}`
        WHERE col_capex_diferencia_fecha_ejecucion BETWEEN '{fecha_inicio}' AND '{fecha_fin}'
    )
    SELECT
        col_capex_diferencia_mes,
        col_capex_diferencia_tipo,
        col_capex_diferencia_area,
        col_capex_diferencia_remanente,
        col_capex_diferencia_presupuesto,
        col_capex_diferencia_fecha_ejecucion
    FROM datos_recientes
    WHERE rn = 1
    ORDER BY col_capex_diferencia_area, col_capex_diferencia_tipo
    """
    
    try:
        df_diferencia = bq_client.query(query).to_dataframe()
        print(f"✅ {len(df_diferencia)} registros extraídos")
        
        if not df_diferencia.empty:
            print(f"\n📋 Áreas: {df_diferencia['col_capex_diferencia_area'].nunique()}")
            print(f"📋 Tipos: {df_diferencia['col_capex_diferencia_tipo'].unique()}")
            print(f"\n📊 Muestra:")
            print(df_diferencia.head(10))
        
        return df_diferencia
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return pd.DataFrame()

def generar_id_diferencia(remanente, presupuesto, ejecutado):
    """
    Generar ID único para diferencia usando SHA256(remanente + presupuesto + ejecutado)
    
    Args:
        remanente: Valor del remanente
        presupuesto: Valor del presupuesto
        ejecutado: Valor del ejecutado
    
    Returns:
        str: ID único en hexadecimal
    """
    # Convertir valores a string y normalizar (manejar NaN, None, etc.)
    remanente_str = str(remanente) if pd.notna(remanente) else "0"
    presupuesto_str = str(presupuesto) if pd.notna(presupuesto) else "0"
    ejecutado_str = str(ejecutado) if pd.notna(ejecutado) else "0"
    
    # Concatenar valores
    concatenado = f"{remanente_str}|{presupuesto_str}|{ejecutado_str}"
    
    # Generar hash SHA256
    hash_obj = hashlib.sha256(concatenado.encode('utf-8'))
    return hash_obj.hexdigest()


def verificar_duplicados_diferencia_venezuela(bq_client, ids_a_verificar: List[str]) -> set:
    """
    Verificar qué IDs ya existen en BigQuery para evitar duplicados
    Procesa en lotes para evitar queries muy largas
    
    Args:
        bq_client: Cliente de BigQuery
        ids_a_verificar: Lista de IDs a verificar
    
    Returns:
        set: Conjunto de IDs que ya existen en BigQuery
    """
    if not ids_a_verificar:
        return set()
    
    try:
        table_id_diferencia = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET}.{BIGQUERY_TABLE_DIFERENCIA}"
        ids_existentes = set()
        batch_size = 1000  # Procesar en lotes de 1000 IDs
        
        print(f"🔍 Verificando {len(ids_a_verificar)} IDs en BigQuery (en lotes de {batch_size})...")
        
        # Procesar en lotes
        for i in range(0, len(ids_a_verificar), batch_size):
            batch = ids_a_verificar[i:i + batch_size]
            
            # Escapar IDs para SQL (usar comillas simples y escapar comillas internas)
            ids_escaped = [f"'{id_val.replace(chr(39), chr(39)+chr(39))}'" for id_val in batch]
            ids_list = ",".join(ids_escaped)
            
            query = f"""
            SELECT DISTINCT vzla_capex_diferencia_id
            FROM `{table_id_diferencia}`
            WHERE vzla_capex_diferencia_id IN ({ids_list})
            """
            
            query_job = bq_client.query(query)
            resultados = query_job.result()
            
            batch_existentes = {row.vzla_capex_diferencia_id for row in resultados}
            ids_existentes.update(batch_existentes)
            
            print(f"   Lote {i//batch_size + 1}: {len(batch_existentes)} duplicados encontrados")
        
        print(f"   ✅ Total: {len(ids_existentes)} IDs duplicados encontrados")
        print(f"   ✅ Total: {len(ids_a_verificar) - len(ids_existentes)} IDs nuevos para cargar")
        
        return ids_existentes
        
    except Exception as e:
        print(f"⚠️ Error verificando duplicados: {e}")
        print(f"   → Continuando sin verificación de duplicados")
        import traceback
        traceback.print_exc()
        return set()


def cargar_diferencia_a_bigquery_venezuela(bq_client, df_tabla2: pd.DataFrame, anio_fiscal: str = None):
    """
    Cargar datos de diferencia a BigQuery (sin columna Diferencia)
    Respetando particiones y clustering
    """
    from google.cloud import bigquery
    from datetime import datetime
    
    print(f"\n📤 Cargando datos a BigQuery (vzla_capex_pago_diferencia)...")
    
    if df_tabla2.empty:
        print(f"⚠️ DataFrame vacío - abortando")
        return
    
    # Calcular año fiscal si no se proporciona
    if not anio_fiscal:
        hoy = datetime.now()
        if hoy.month >= 8:
            anio_inicio = hoy.year
            anio_fin = hoy.year + 1
        else:
            anio_inicio = hoy.year - 1
            anio_fin = hoy.year
        anio_fiscal = f"{anio_inicio}-{anio_fin}"
    
    # Preparar datos para BigQuery (sin columna Diferencia)
    # Obtener nombres de columnas dinámicamente del DataFrame
    columnas_df = list(df_tabla2.columns)
    
    # Identificar columnas por su prefijo (son dinámicas según el mes)
    col_remanente = [col for col in columnas_df if col.startswith('Remanente')]
    col_presupuesto = [col for col in columnas_df if col.startswith('Presupuesto')]
    col_ejecutado = [col for col in columnas_df if col.startswith('Ejecutado')]
    
    # Validar que existan las columnas necesarias
    if not col_remanente or not col_presupuesto or not col_ejecutado:
        print(f"❌ Error: No se encontraron las columnas esperadas en df_tabla2")
        print(f"   Columnas disponibles: {columnas_df}")
        raise ValueError(f"Columnas requeridas no encontradas. Disponibles: {columnas_df}")
    
    # Obtener los nombres de las columnas (tomar el primero si hay múltiples)
    nombre_remanente = col_remanente[0]
    nombre_presupuesto = col_presupuesto[0]
    nombre_ejecutado = col_ejecutado[0]
    
    print(f"\n📋 Columnas detectadas dinámicamente:")
    print(f"   Remanente: {nombre_remanente}")
    print(f"   Presupuesto: {nombre_presupuesto}")
    print(f"   Ejecutado: {nombre_ejecutado}")
    
    # IMPORTANTE: df_tabla2 ahora tiene 'area' como columna, no como índice
    # Verificar si 'area' está en las columnas
    if 'area' not in df_tabla2.columns:
        print(f"❌ Error: La columna 'area' no está en df_tabla2")
        print(f"   Columnas disponibles: {list(df_tabla2.columns)}")
        raise ValueError("La columna 'area' es requerida en df_tabla2")
    
    # Crear DataFrame con las columnas dinámicas
    df_bq = df_tabla2[[nombre_remanente, nombre_presupuesto, nombre_ejecutado]].copy()
    
    # Obtener las áreas de la columna 'area', no del índice
    areas_list = [str(area) for area in df_tabla2['area']]
    
    # Calcular mes actual para vzla_capex_diferencia_mes basado en el viernes de la semana pasada
    # Formato: 'NOV-25' (mes abreviado - año de 2 dígitos)
    # Usa la misma lógica que en utils.py (viernes de la semana pasada)
    import datetime as dt
    
    # Obtener el viernes de la semana pasada (misma lógica que en utils.py)
    hoy_date = dt.date.today()
    dia_semana_actual = hoy_date.weekday()  # lunes=0, viernes=4, domingo=6
    
    # Calcular días hasta el viernes de esta semana
    dias_hasta_viernes_esta_semana = (4 - dia_semana_actual) % 7
    
    # Si hoy es viernes (dias_hasta_viernes_esta_semana = 0), el viernes pasado fue hace 7 días
    # Si no, el viernes pasado fue hace (dias_hasta_viernes_esta_semana + 7) días
    if dias_hasta_viernes_esta_semana == 0:
        dias_retroceso = 7
    else:
        dias_retroceso = dias_hasta_viernes_esta_semana + 7
    
    viernes_pasado = hoy_date - dt.timedelta(days=dias_retroceso)
    
    # Usar el viernes pasado solo para el mes
    df_bq['vzla_capex_diferencia_mes'] = viernes_pasado.strftime('%b-%y').upper()
    
    # La fecha de ejecución es la fecha actual (momento en que se ejecuta el proceso)
    hoy = datetime.now()
    
    # Asignar tipo CAPEX según el área
    # IMPORTANTE: Si df_tabla2 ya tiene la columna 'tipo_capex', usarla directamente
    # Esto es crítico porque después de la unificación, todas las filas de CONSTRUCCIÓN
    # tienen el mismo nombre pero diferentes tipos (ORDINARIO vs EXTRAORDINARIO)
    if 'tipo_capex' in df_tabla2.columns:
        print(f"✅ Usando columna 'tipo_capex' existente de df_tabla2")
        print(f"   Tipos únicos: {df_tabla2['tipo_capex'].unique()}")
        # Mostrar distribución de tipos para CONSTRUCCIÓN
        construccion_mask = df_tabla2['area'].str.contains('CONSTRUCCION', case=False, na=False)
        if construccion_mask.any():
            construccion_tipos = df_tabla2[construccion_mask][['area', 'tipo_capex']]
            print(f"   CONSTRUCCIÓN - Distribución de tipos:")
            for idx, row in construccion_tipos.iterrows():
                print(f"      {row['area']}: {row['tipo_capex']}")
        # Asegurarse de que el orden coincida con las filas de df_bq
        df_bq['vzla_capex_diferencia_tipo'] = df_tabla2['tipo_capex'].values
    else:
        print(f"⚠️  Columna 'tipo_capex' no encontrada, calculando basándose en nombre del área")
        print(f"   ⚠️  ADVERTENCIA: Esto puede causar que todas las filas de CONSTRUCCIÓN se clasifiquen como EXTRAORDINARIO")
        # CAPEX EXTRAORDINARIO para "DIR CONSTRUCCIÓN Y PROYECTOS" (con acento en la O)
        # CAPEX ORDINARIO para todas las demás áreas
        def asignar_tipo_capex(area):
            if pd.isna(area):
                return 'CAPEX ORDINARIO'
            area_str = str(area).strip()
            # Verificar específicamente que tenga "CONSTRUCCIÓN" con acento en la O
            # Buscar "DIR CONSTRUCCIÓN" (con acento) y "PROYECTOS"
            if 'DIR CONSTRUCCIÓN' in area_str and 'PROYECTOS' in area_str:
                return 'CAPEX EXTRAORDINARIO'
            # Si tiene "CONSTRUCCION" sin acento, es ORDINARIO
            return 'CAPEX ORDINARIO'
        
        # Aplicar la función a cada área para asignar el tipo CAPEX
        df_bq['vzla_capex_diferencia_tipo'] = [asignar_tipo_capex(area) for area in areas_list]
    
    # Asignar el área con el nombre correcto de BigQuery directamente desde la lista
    df_bq['vzla_capex_diferencia_area'] = areas_list
    
    # Asegurarse de que el índice no tenga nombre que cause problemas
    df_bq.index.name = None
    
    # ===================================================================
    # CALCULAR REMANENTE DEL MES ACTUAL BASADO EN DIFERENCIA DEL MES ANTERIOR
    # Si hay cambio de mes, el remanente = presupuesto_mes_anterior - ejecutado_mes_anterior
    # ===================================================================
    print(f"\n🔄 Verificando cambio de mes para calcular remanente...")
    
    # Obtener mes actual y mes anterior (basado en viernes pasado)
    from dateutil.relativedelta import relativedelta
    
    mes_actual = viernes_pasado
    mes_anterior = viernes_pasado - relativedelta(months=1)
    
    # Formatear meses para comparación (formato: 'NOV-25')
    meses_espanol = {
        'JANUARY': 'ENE', 'FEBRUARY': 'FEB', 'MARCH': 'MAR', 'APRIL': 'ABR',
        'MAY': 'MAY', 'JUNE': 'JUN', 'JULY': 'JUL', 'AUGUST': 'AGO',
        'SEPTEMBER': 'SEP', 'OCTOBER': 'OCT', 'NOVEMBER': 'NOV', 'DECEMBER': 'DIC'
    }
    mes_actual_str = meses_espanol.get(mes_actual.strftime('%B').upper(), mes_actual.strftime('%b').upper())
    mes_anterior_str = meses_espanol.get(mes_anterior.strftime('%B').upper(), mes_anterior.strftime('%b').upper())
    mes_actual_formato = f"{mes_actual_str}-{mes_actual.strftime('%y')}"
    mes_anterior_formato = f"{mes_anterior_str}-{mes_anterior.strftime('%y')}"
    
    print(f"   Mes actual: {mes_actual_formato}")
    print(f"   Mes anterior: {mes_anterior_formato}")
    
    # Calcular rango de fechas del año fiscal para la query
    anio_inicio_int = int(anio_fiscal.split('-')[0])
    fecha_inicio_query = f"{anio_inicio_int}-08-01"
    fecha_fin_query = f"{anio_inicio_int + 1}-07-31"
    
    # Consultar BigQuery para obtener datos del mes anterior
    table_id_diferencia = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET}.{BIGQUERY_TABLE_DIFERENCIA}"
    query_mes_anterior = f"""
    WITH datos_recientes AS (
        SELECT
            vzla_capex_diferencia_mes,
            vzla_capex_diferencia_tipo,
            vzla_capex_diferencia_area,
            vzla_capex_diferencia_presupuesto,
            vzla_capex_diferencia_ejecutado,
            ROW_NUMBER() OVER (
                PARTITION BY vzla_capex_diferencia_area, vzla_capex_diferencia_tipo
                ORDER BY vzla_capex_diferencia_fecha_ejecucion DESC
            ) as rn
        FROM `{table_id_diferencia}`
        WHERE vzla_capex_diferencia_mes = '{mes_anterior_formato}'
          AND vzla_capex_diferencia_fecha_ejecucion BETWEEN '{fecha_inicio_query}' AND '{fecha_fin_query}'
    )
    SELECT
        vzla_capex_diferencia_tipo,
        vzla_capex_diferencia_area,
        vzla_capex_diferencia_presupuesto,
        vzla_capex_diferencia_ejecutado
    FROM datos_recientes
    WHERE rn = 1
    """
    
    try:
        df_mes_anterior = bq_client.query(query_mes_anterior).to_dataframe()
        print(f"   ✅ Datos del mes anterior encontrados: {len(df_mes_anterior)} registros")
        
        if not df_mes_anterior.empty:
            # Calcular nuevo remanente: presupuesto - ejecutado (diferencia del mes anterior)
            df_mes_anterior['nuevo_remanente'] = (
                df_mes_anterior['vzla_capex_diferencia_presupuesto'] - 
                df_mes_anterior['vzla_capex_diferencia_ejecutado']
            )
            
            # Crear diccionario de lookup: (tipo, area) -> nuevo_remanente
            lookup_remanente = {}
            for _, row in df_mes_anterior.iterrows():
                key = (str(row['vzla_capex_diferencia_tipo']), str(row['vzla_capex_diferencia_area']))
                lookup_remanente[key] = row['nuevo_remanente']
            
            print(f"   📊 Calculando remanente del mes actual basado en diferencia del mes anterior...")
            print(f"   → Remanente = Presupuesto({mes_anterior_formato}) - Ejecutado({mes_anterior_formato})")
            
            # Aplicar nuevo remanente a df_bq
            remanentes_actualizados = 0
            for idx in df_bq.index:
                tipo = df_bq.loc[idx, 'vzla_capex_diferencia_tipo']
                area = df_bq.loc[idx, 'vzla_capex_diferencia_area']
                key = (str(tipo), str(area))
                
                if key in lookup_remanente:
                    nuevo_remanente = lookup_remanente[key]
                    # Solo actualizar si el remanente actual es diferente (evitar sobrescribir si ya está correcto)
                    remanente_actual = df_bq.loc[idx, nombre_remanente]
                    if abs(float(remanente_actual) - float(nuevo_remanente)) > 0.01:  # Tolerancia para comparación de floats
                        df_bq.loc[idx, nombre_remanente] = nuevo_remanente
                        remanentes_actualizados += 1
                        print(f"      {area} ({tipo}): {remanente_actual:.2f} → {nuevo_remanente:.2f}")
            
            print(f"   ✅ Remanentes actualizados: {remanentes_actualizados} de {len(df_bq)}")
        else:
            print(f"   ⚠️ No se encontraron datos del mes anterior ({mes_anterior_formato})")
            print(f"   → Usando remanente del DataFrame original")
    except Exception as e:
        print(f"   ⚠️ Error consultando mes anterior: {e}")
        print(f"   → Usando remanente del DataFrame original")
        import traceback
        traceback.print_exc()
    
    # Asignar valores a las columnas de BigQuery
    df_bq['vzla_capex_diferencia_remanente'] = df_bq[nombre_remanente]
    df_bq['vzla_capex_diferencia_presupuesto'] = df_bq[nombre_presupuesto]
    df_bq['vzla_capex_diferencia_ejecutado'] = df_bq[nombre_ejecutado]
    df_bq['vzla_capex_diferencia_fecha_ejecucion'] = hoy
    
    # Generar IDs únicos para cada fila usando SHA256(remanente + presupuesto + ejecutado)
    print(f"\n🔑 Generando IDs únicos para cada fila...")
    df_bq['vzla_capex_diferencia_id'] = df_bq.apply(
        lambda row: generar_id_diferencia(
            row['vzla_capex_diferencia_remanente'],
            row['vzla_capex_diferencia_presupuesto'],
            row['vzla_capex_diferencia_ejecutado']
        ),
        axis=1
    )
    
    # Verificar duplicados en BigQuery
    ids_a_verificar = df_bq['vzla_capex_diferencia_id'].tolist()
    ids_existentes = verificar_duplicados_diferencia_venezuela(bq_client, ids_a_verificar)
    
    # Filtrar filas que no son duplicados
    if ids_existentes:
        filas_antes = len(df_bq)
        df_bq = df_bq[~df_bq['vzla_capex_diferencia_id'].isin(ids_existentes)]
        filas_despues = len(df_bq)
        print(f"\n📊 Filtrado de duplicados:")
        print(f"   Filas antes: {filas_antes}")
        print(f"   Filas duplicadas eliminadas: {filas_antes - filas_despues}")
        print(f"   Filas nuevas a cargar: {filas_despues}")
    
    # Si no hay filas nuevas después de filtrar duplicados, abortar
    if df_bq.empty:
        print(f"\n✅ No hay datos nuevos para cargar (todos son duplicados)")
        return
    
    # Seleccionar solo columnas BigQuery (incluyendo el ID)
    df_bq = df_bq[[
        'vzla_capex_diferencia_id',
        'vzla_capex_diferencia_mes',
        'vzla_capex_diferencia_tipo',
        'vzla_capex_diferencia_area',
        'vzla_capex_diferencia_remanente',
        'vzla_capex_diferencia_presupuesto',
        'vzla_capex_diferencia_ejecutado',
        'vzla_capex_diferencia_fecha_ejecucion'
    ]]
    
    print(f"\n📊 Datos a cargar:")
    print(f"   Filas: {len(df_bq)}")
    print(f"   Columnas: {list(df_bq.columns)}")
    print(f"   Índice nombre: {df_bq.index.name}")
    
    # Verificar que no haya columnas con nombres incorrectos
    columnas_incorrectas = [col for col in df_bq.columns if col.lower() == 'area' and col != 'vzla_capex_diferencia_area']
    if columnas_incorrectas:
        print(f"⚠️  Advertencia: Se encontraron columnas con nombre 'area': {columnas_incorrectas}")
        df_bq = df_bq.drop(columns=columnas_incorrectas)
    
    try:
        # Ajustar DataFrame al schema de BigQuery antes de cargar
        print(f"\n🔧 Ajustando DataFrame al schema de BigQuery...")
        df_bq = ajustar_df_a_schema_bigquery_venezuela(df_bq, bq_client, BIGQUERY_DATASET, BIGQUERY_TABLE_DIFERENCIA)
        
        # Verificar nuevamente las columnas después del ajuste
        print(f"📋 Columnas después del ajuste: {list(df_bq.columns)}")
        columnas_incorrectas_post = [col for col in df_bq.columns if col.lower() == 'area' and col != 'vzla_capex_diferencia_area']
        if columnas_incorrectas_post:
            print(f"❌ Error: Columnas incorrectas después del ajuste: {columnas_incorrectas_post}")
            df_bq = df_bq.drop(columns=columnas_incorrectas_post)
        
        job_config = bigquery.LoadJobConfig(
            write_disposition="WRITE_APPEND",  # Agregar datos
        )
        
        table_id_diferencia = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET}.{BIGQUERY_TABLE_DIFERENCIA}"
        job = bq_client.load_table_from_dataframe(
            df_bq,
            table_id_diferencia,
            job_config=job_config
        )
        
        job.result()
        print(f"✅ {job.output_rows} filas cargadas a BigQuery")
        
    except Exception as e:
        print(f"❌ Error cargando: {e}")
        import traceback
        traceback.print_exc()

def verificar_duplicados_diferencia_colombia(bq_client, ids_a_verificar: List[str]) -> set:
    """
    Verificar qué IDs ya existen en BigQuery para evitar duplicados
    Procesa en lotes para evitar queries muy largas
    
    IMPORTANTE: La tabla está particionada por col_capex_diferencia_fecha_ejecucion,
    por lo que se requiere un filtro sobre esa columna.
    
    Args:
        bq_client: Cliente de BigQuery
        ids_a_verificar: Lista de IDs a verificar
    
    Returns:
        set: Conjunto de IDs que ya existen en BigQuery
    """
    if not ids_a_verificar:
        return set()
    
    try:
        from datetime import datetime, timedelta
        
        table_id_diferencia = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET_COP}.{BIGQUERY_TABLE_DIFERENCIA_COP}"
        ids_existentes = set()
        batch_size = 1000  # Procesar en lotes de 1000 IDs
        
        # Calcular rango de fechas del año fiscal actual para el filtro de partición
        hoy = datetime.now()
        if hoy.month >= 8:  # Agosto o después
            anio_inicio = hoy.year
            anio_fin = hoy.year + 1
        else:
            anio_inicio = hoy.year - 1
            anio_fin = hoy.year
        
        # Rango de fechas del año fiscal (Agosto a Julio)
        fecha_inicio = f"{anio_inicio}-08-01"
        fecha_fin = f"{anio_fin}-07-31"
        
        print(f"🔍 Verificando {len(ids_a_verificar)} IDs en BigQuery (en lotes de {batch_size})...")
        print(f"   📅 Filtro de partición: {fecha_inicio} a {fecha_fin}")
        
        # Procesar en lotes
        for i in range(0, len(ids_a_verificar), batch_size):
            batch = ids_a_verificar[i:i + batch_size]
            
            # Escapar IDs para SQL (usar comillas simples y escapar comillas internas)
            ids_escaped = [f"'{id_val.replace(chr(39), chr(39)+chr(39))}'" for id_val in batch]
            ids_list = ",".join(ids_escaped)
            
            # IMPORTANTE: Agregar filtro sobre col_capex_diferencia_fecha_ejecucion para partición
            query = f"""
            SELECT DISTINCT col_capex_diferencia_id
            FROM `{table_id_diferencia}`
            WHERE col_capex_diferencia_id IN ({ids_list})
              AND col_capex_diferencia_fecha_ejecucion BETWEEN '{fecha_inicio}' AND '{fecha_fin}'
            """
            
            query_job = bq_client.query(query)
            resultados = query_job.result()
            
            batch_existentes = {row.col_capex_diferencia_id for row in resultados}
            ids_existentes.update(batch_existentes)
            
            print(f"   Lote {i//batch_size + 1}: {len(batch_existentes)} duplicados encontrados")
        
        print(f"   ✅ Total: {len(ids_existentes)} IDs duplicados encontrados")
        print(f"   ✅ Total: {len(ids_a_verificar) - len(ids_existentes)} IDs nuevos para cargar")
        
        return ids_existentes
        
    except Exception as e:
        print(f"⚠️ Error verificando duplicados: {e}")
        print(f"   → Continuando sin verificación de duplicados")
        import traceback
        traceback.print_exc()
        return set()


def cargar_diferencia_a_bigquery_colombia(bq_client, df_tabla2: pd.DataFrame, anio_fiscal: str = None):
    """
    Cargar datos de diferencia a BigQuery (sin columna Diferencia)
    Respetando particiones y clustering
    """
    from google.cloud import bigquery
    from datetime import datetime
    
    print(f"\n📤 Cargando datos a BigQuery (col_capex_pago_diferencia)...")
    
    if df_tabla2.empty:
        print(f"⚠️ DataFrame vacío - abortando")
        return
    
    # Calcular año fiscal si no se proporciona
    if not anio_fiscal:
        hoy = datetime.now()
        if hoy.month >= 8:
            anio_inicio = hoy.year
            anio_fin = hoy.year + 1
        else:
            anio_inicio = hoy.year - 1
            anio_fin = hoy.year
        anio_fiscal = f"{anio_inicio}-{anio_fin}"
    
    # Preparar datos para BigQuery (sin columna Diferencia)
    # Obtener nombres de columnas dinámicamente del DataFrame
    columnas_df = list(df_tabla2.columns)
    
    # Identificar columnas por su prefijo (son dinámicas según el mes)
    col_remanente = [col for col in columnas_df if col.startswith('Remanente')]
    col_presupuesto = [col for col in columnas_df if col.startswith('Presupuesto')]
    col_ejecutado = [col for col in columnas_df if col.startswith('Ejecutado')]
    
    # Validar que existan las columnas necesarias
    if not col_remanente or not col_presupuesto or not col_ejecutado:
        print(f"❌ Error: No se encontraron las columnas esperadas en df_tabla2")
        print(f"   Columnas disponibles: {columnas_df}")
        raise ValueError(f"Columnas requeridas no encontradas. Disponibles: {columnas_df}")
    
    # Obtener los nombres de las columnas (tomar el primero si hay múltiples)
    nombre_remanente = col_remanente[0]
    nombre_presupuesto = col_presupuesto[0]
    nombre_ejecutado = col_ejecutado[0]
    
    print(f"\n📋 Columnas detectadas dinámicamente:")
    print(f"   Remanente: {nombre_remanente}")
    print(f"   Presupuesto: {nombre_presupuesto}")
    print(f"   Ejecutado: {nombre_ejecutado}")
    
    # IMPORTANTE: df_tabla2 ahora tiene 'area' como columna, no como índice
    # Verificar si 'area' está en las columnas
    if 'area' not in df_tabla2.columns:
        print(f"❌ Error: La columna 'area' no está en df_tabla2")
        print(f"   Columnas disponibles: {list(df_tabla2.columns)}")
        raise ValueError("La columna 'area' es requerida en df_tabla2")
    
    # Crear DataFrame con las columnas dinámicas
    df_bq = df_tabla2[[nombre_remanente, nombre_presupuesto, nombre_ejecutado]].copy()
    
    # Obtener las áreas de la columna 'area', no del índice
    areas_list = [str(area) for area in df_tabla2['area']]
    
    # Calcular mes actual para vzla_capex_diferencia_mes basado en el viernes de la semana pasada
    # Formato: 'NOV-25' (mes abreviado - año de 2 dígitos)
    # Usa la misma lógica que en utils.py (viernes de la semana pasada)
    import datetime as dt
    
    # Obtener el viernes de la semana pasada (misma lógica que en utils.py)
    hoy_date = dt.date.today()
    dia_semana_actual = hoy_date.weekday()  # lunes=0, viernes=4, domingo=6
    
    # Calcular días hasta el viernes de esta semana
    dias_hasta_viernes_esta_semana = (4 - dia_semana_actual) % 7
    
    # Si hoy es viernes (dias_hasta_viernes_esta_semana = 0), el viernes pasado fue hace 7 días
    # Si no, el viernes pasado fue hace (dias_hasta_viernes_esta_semana + 7) días
    if dias_hasta_viernes_esta_semana == 0:
        dias_retroceso = 7
    else:
        dias_retroceso = dias_hasta_viernes_esta_semana + 7
    
    viernes_pasado = hoy_date - dt.timedelta(days=dias_retroceso)
    
    # Usar el viernes pasado solo para el mes
    df_bq['col_capex_diferencia_mes'] = viernes_pasado.strftime('%b-%y').upper()
    
    # La fecha de ejecución es la fecha actual (momento en que se ejecuta el proceso)
    hoy = datetime.now()
    
    # Asignar tipo CAPEX según el área
    # IMPORTANTE: Si df_tabla2 ya tiene la columna 'tipo_capex', usarla directamente
    # Esto es crítico porque después de la unificación, todas las filas de CONSTRUCCIÓN
    # tienen el mismo nombre pero diferentes tipos (ORDINARIO vs EXTRAORDINARIO)
    if 'tipo_capex' in df_tabla2.columns:
        print(f"✅ Usando columna 'tipo_capex' existente de df_tabla2")
        print(f"   Tipos únicos: {df_tabla2['tipo_capex'].unique()}")
        # Mostrar distribución de tipos para CONSTRUCCIÓN
        construccion_mask = df_tabla2['area'].str.contains('CONSTRUCCION', case=False, na=False)
        if construccion_mask.any():
            construccion_tipos = df_tabla2[construccion_mask][['area', 'tipo_capex']]
            print(f"   CONSTRUCCIÓN - Distribución de tipos:")
            for idx, row in construccion_tipos.iterrows():
                print(f"      {row['area']}: {row['tipo_capex']}")
        # Asegurarse de que el orden coincida con las filas de df_bq
        df_bq['col_capex_diferencia_tipo'] = df_tabla2['tipo_capex'].values
    else:
        print(f"⚠️  Columna 'tipo_capex' no encontrada, calculando basándose en nombre del área")
        print(f"   ⚠️  ADVERTENCIA: Esto puede causar que todas las filas de CONSTRUCCIÓN se clasifiquen como EXTRAORDINARIO")
        # CAPEX EXTRAORDINARIO para "DIR CONSTRUCCIÓN Y PROYECTOS" (con acento en la O)
        # CAPEX ORDINARIO para todas las demás áreas
        def asignar_tipo_capex(area):
            if pd.isna(area):
                return 'CAPEX ORDINARIO'
            area_str = str(area).strip()
            # Verificar específicamente que tenga "CONSTRUCCIÓN" con acento en la O
            # Buscar "DIR CONSTRUCCIÓN" (con acento) y "PROYECTOS"
            if 'DIR CONSTRUCCIÓN' in area_str and 'PROYECTOS' in area_str:
                return 'CAPEX EXTRAORDINARIO'
            # Si tiene "CONSTRUCCION" sin acento, es ORDINARIO
            return 'CAPEX ORDINARIO'
        
        # Aplicar la función a cada área para asignar el tipo CAPEX
        df_bq['col_capex_diferencia_tipo'] = [asignar_tipo_capex(area) for area in areas_list]
    
    # Asignar el área con el nombre correcto de BigQuery directamente desde la lista
    df_bq['col_capex_diferencia_area'] = areas_list
    
    # Asegurarse de que el índice no tenga nombre que cause problemas
    df_bq.index.name = None
    
    # ===================================================================
    # CALCULAR REMANENTE DEL MES ACTUAL BASADO EN DIFERENCIA DEL MES ANTERIOR
    # Si hay cambio de mes, el remanente = presupuesto_mes_anterior - ejecutado_mes_anterior
    # ===================================================================
    print(f"\n🔄 Verificando cambio de mes para calcular remanente...")
    
    # Obtener mes actual y mes anterior (basado en viernes pasado)
    import datetime as dt
    from dateutil.relativedelta import relativedelta
    
    hoy_date = dt.date.today()
    dia_semana_actual = hoy_date.weekday()
    dias_hasta_viernes_esta_semana = (4 - dia_semana_actual) % 7
    if dias_hasta_viernes_esta_semana == 0:
        dias_retroceso = 7
    else:
        dias_retroceso = dias_hasta_viernes_esta_semana + 7
    viernes_pasado = hoy_date - dt.timedelta(days=dias_retroceso)
    mes_actual = viernes_pasado
    mes_anterior = viernes_pasado - relativedelta(months=1)
    
    # Formatear meses para comparación (formato: 'NOV-25')
    meses_espanol = {
        'JANUARY': 'ENE', 'FEBRUARY': 'FEB', 'MARCH': 'MAR', 'APRIL': 'ABR',
        'MAY': 'MAY', 'JUNE': 'JUN', 'JULY': 'JUL', 'AUGUST': 'AGO',
        'SEPTEMBER': 'SEP', 'OCTOBER': 'OCT', 'NOVEMBER': 'NOV', 'DECEMBER': 'DIC'
    }
    mes_actual_str = meses_espanol.get(mes_actual.strftime('%B').upper(), mes_actual.strftime('%b').upper())
    mes_anterior_str = meses_espanol.get(mes_anterior.strftime('%B').upper(), mes_anterior.strftime('%b').upper())
    mes_actual_formato = f"{mes_actual_str}-{mes_actual.strftime('%y')}"
    mes_anterior_formato = f"{mes_anterior_str}-{mes_anterior.strftime('%y')}"
    
    print(f"   Mes actual: {mes_actual_formato}")
    print(f"   Mes anterior: {mes_anterior_formato}")
    
    # Calcular rango de fechas del año fiscal para la query
    anio_inicio_int = int(anio_fiscal.split('-')[0])
    fecha_inicio_query = f"{anio_inicio_int}-08-01"
    fecha_fin_query = f"{anio_inicio_int + 1}-07-31"
    
    # Consultar BigQuery para obtener datos del mes anterior
    table_id_diferencia = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET_COP}.{BIGQUERY_TABLE_DIFERENCIA_COP}"
    query_mes_anterior = f"""
    WITH datos_recientes AS (
        SELECT
            col_capex_diferencia_mes,
            col_capex_diferencia_tipo,
            col_capex_diferencia_area,
            col_capex_diferencia_presupuesto,
            col_capex_diferencia_ejecutado,
            ROW_NUMBER() OVER (
                PARTITION BY col_capex_diferencia_area, col_capex_diferencia_tipo
                ORDER BY col_capex_diferencia_fecha_ejecucion DESC
            ) as rn
        FROM `{table_id_diferencia}`
        WHERE col_capex_diferencia_mes = '{mes_anterior_formato}'
          AND col_capex_diferencia_fecha_ejecucion BETWEEN '{fecha_inicio_query}' AND '{fecha_fin_query}'
    )
    SELECT
        col_capex_diferencia_tipo,
        col_capex_diferencia_area,
        col_capex_diferencia_presupuesto,
        col_capex_diferencia_ejecutado
    FROM datos_recientes
    WHERE rn = 1
    """
    
    try:
        df_mes_anterior = bq_client.query(query_mes_anterior).to_dataframe()
        print(f"   ✅ Datos del mes anterior encontrados: {len(df_mes_anterior)} registros")
        
        if not df_mes_anterior.empty:
            # Calcular nuevo remanente: presupuesto - ejecutado (diferencia del mes anterior)
            df_mes_anterior['nuevo_remanente'] = (
                df_mes_anterior['col_capex_diferencia_presupuesto'] - 
                df_mes_anterior['col_capex_diferencia_ejecutado']
            )
            
            # Crear diccionario de lookup: (tipo, area) -> nuevo_remanente
            lookup_remanente = {}
            for _, row in df_mes_anterior.iterrows():
                key = (str(row['col_capex_diferencia_tipo']), str(row['col_capex_diferencia_area']))
                lookup_remanente[key] = row['nuevo_remanente']
            
            print(f"   📊 Calculando remanente del mes actual basado en diferencia del mes anterior...")
            print(f"   → Remanente = Presupuesto({mes_anterior_formato}) - Ejecutado({mes_anterior_formato})")
            
            # Aplicar nuevo remanente a df_bq
            remanentes_actualizados = 0
            for idx in df_bq.index:
                tipo = df_bq.loc[idx, 'col_capex_diferencia_tipo']
                area = df_bq.loc[idx, 'col_capex_diferencia_area']
                key = (str(tipo), str(area))
                
                if key in lookup_remanente:
                    nuevo_remanente = lookup_remanente[key]
                    # Solo actualizar si el remanente actual es diferente (evitar sobrescribir si ya está correcto)
                    remanente_actual = df_bq.loc[idx, nombre_remanente]
                    if abs(float(remanente_actual) - float(nuevo_remanente)) > 0.01:  # Tolerancia para comparación de floats
                        df_bq.loc[idx, nombre_remanente] = nuevo_remanente
                        remanentes_actualizados += 1
                        print(f"      {area} ({tipo}): {remanente_actual:.2f} → {nuevo_remanente:.2f}")
            
            print(f"   ✅ Remanentes actualizados: {remanentes_actualizados} de {len(df_bq)}")
        else:
            print(f"   ⚠️ No se encontraron datos del mes anterior ({mes_anterior_formato})")
            print(f"   → Usando remanente del DataFrame original")
    except Exception as e:
        print(f"   ⚠️ Error consultando mes anterior: {e}")
        print(f"   → Usando remanente del DataFrame original")
        import traceback
        traceback.print_exc()
    
    # Asignar valores a las columnas de BigQuery
    df_bq['col_capex_diferencia_remanente'] = df_bq[nombre_remanente]
    df_bq['col_capex_diferencia_presupuesto'] = df_bq[nombre_presupuesto]
    df_bq['col_capex_diferencia_ejecutado'] = df_bq[nombre_ejecutado]
    df_bq['col_capex_diferencia_fecha_ejecucion'] = hoy
    
    # Generar IDs únicos para cada fila usando SHA256(remanente + presupuesto + ejecutado)
    print(f"\n🔑 Generando IDs únicos para cada fila...")
    df_bq['col_capex_diferencia_id'] = df_bq.apply(
        lambda row: generar_id_diferencia(
            row['col_capex_diferencia_remanente'],
            row['col_capex_diferencia_presupuesto'],
            row['col_capex_diferencia_ejecutado']
        ),
        axis=1
    )
    
    # Verificar duplicados en BigQuery
    ids_a_verificar = df_bq['col_capex_diferencia_id'].tolist()
    ids_existentes = verificar_duplicados_diferencia_colombia(bq_client, ids_a_verificar)
    
    # Filtrar filas que no son duplicados
    if ids_existentes:
        filas_antes = len(df_bq)
        df_bq = df_bq[~df_bq['col_capex_diferencia_id'].isin(ids_existentes)]
        filas_despues = len(df_bq)
        print(f"\n📊 Filtrado de duplicados:")
        print(f"   Filas antes: {filas_antes}")
        print(f"   Filas duplicadas eliminadas: {filas_antes - filas_despues}")
        print(f"   Filas nuevas a cargar: {filas_despues}")
    
    # Si no hay filas nuevas después de filtrar duplicados, abortar
    if df_bq.empty:
        print(f"\n✅ No hay datos nuevos para cargar (todos son duplicados)")
        return
    
    # Seleccionar solo columnas BigQuery (incluyendo el ID)
    df_bq = df_bq[[
        'col_capex_diferencia_id',
        'col_capex_diferencia_mes',
        'col_capex_diferencia_tipo',
        'col_capex_diferencia_area',
        'col_capex_diferencia_remanente',
        'col_capex_diferencia_presupuesto',
        'col_capex_diferencia_ejecutado',
        'col_capex_diferencia_fecha_ejecucion'
    ]]
    
    print(f"\n📊 Datos a cargar:")
    print(f"   Filas: {len(df_bq)}")
    print(f"   Columnas: {list(df_bq.columns)}")
    print(f"   Índice nombre: {df_bq.index.name}")
    
    # Verificar que no haya columnas con nombres incorrectos
    columnas_incorrectas = [col for col in df_bq.columns if col.lower() == 'area' and col != 'col_capex_diferencia_area']
    if columnas_incorrectas:
        print(f"⚠️  Advertencia: Se encontraron columnas con nombre 'area': {columnas_incorrectas}")
        df_bq = df_bq.drop(columns=columnas_incorrectas)
    
    try:
        # Ajustar DataFrame al schema de BigQuery antes de cargar
        print(f"\n🔧 Ajustando DataFrame al schema de BigQuery...")
        df_bq = ajustar_df_a_schema_bigquery_colombia(df_bq, bq_client, BIGQUERY_DATASET_COP, BIGQUERY_TABLE_DIFERENCIA_COP)
        
        # Verificar nuevamente las columnas después del ajuste
        print(f"📋 Columnas después del ajuste: {list(df_bq.columns)}")
        columnas_incorrectas_post = [col for col in df_bq.columns if col.lower() == 'area' and col != 'col_capex_diferencia_area']
        if columnas_incorrectas_post:
            print(f"❌ Error: Columnas incorrectas después del ajuste: {columnas_incorrectas_post}")
            df_bq = df_bq.drop(columns=columnas_incorrectas_post)
        
        job_config = bigquery.LoadJobConfig(
            write_disposition="WRITE_APPEND",  # Agregar datos
        )
        
        table_id_diferencia = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET_COP}.{BIGQUERY_TABLE_DIFERENCIA_COP}"
        job = bq_client.load_table_from_dataframe(
            df_bq,
            table_id_diferencia,
            job_config=job_config
        )
        
        job.result()
        print(f"✅ {job.output_rows} filas cargadas a BigQuery")
        
    except Exception as e:
        print(f"❌ Error cargando: {e}")
        import traceback
        traceback.print_exc()


# =================== SUBIDA A GCS ===================

def subir_archivo_a_gcs(storage_client: storage.Client, archivo_local: str) -> Dict[str, str]:
    """
    Subir archivo a Google Cloud Storage
    Returns: (url_publica, nombre_blob)
    """
    try:
        # Generar nombre con timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        nombre_blob = f"Consolidado_de_pago_{timestamp}.xlsx"
        
        print(f"📤 Subiendo archivo a GCS: {nombre_blob}")
        
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(nombre_blob)
        
        # Subir archivo
        blob.upload_from_filename(archivo_local, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Hacer el blob público para que sea accesible sin autenticación
        try:
            blob.make_public()
            print(f"   ✓ Archivo configurado como público")
        except Exception as public_error:
            print(f"   ⚠️ No se pudo hacer el archivo público: {str(public_error)}")
            print(f"   → Verifica que la cuenta de servicio tenga permisos 'storage.objects.setIamPolicy'")
            print(f"   → O configura el bucket para permitir acceso público")
        
        # Usar URL pública del blob (siempre disponible después de make_public)
        url_publica = blob.public_url
        if not url_publica or 'storage.googleapis.com' not in url_publica:
            # Construir URL pública manualmente si es necesario
            url_publica = f"https://storage.googleapis.com/{GCS_BUCKET_NAME}/{nombre_blob}"
        
        print(f"✅ Archivo subido exitosamente")
        print(f"   URL pública: {url_publica}")
        
        return url_publica, nombre_blob
        
    except Exception as e:
        print(f"❌ Error subiendo archivo a GCS: {e}")
        raise

# =================== MAPEO INVERSO BQ → EXCEL ===================

def mapear_bigquery_a_excel_columns_venezuela(df_bq: pd.DataFrame) -> pd.DataFrame:
    """Convertir nombres de columnas de BigQuery a nombres de Excel"""
    
    mapeo_inverso = {
        'vzla_capex_pago_numero_factura': 'Numero de Factura',
        'vzla_capex_pago_orden_compra': 'Numero de OC',
        'vzla_capex_pago_tipo_documento': 'Tipo Factura',
        'vzla_capex_pago_nombre_lote': 'Nombre Lote',
        'vzla_capex_pago_proveedor': 'Proveedor',
        'vzla_capex_pago_rif': 'RIF',
        'vzla_capex_pago_fecha_documento': 'Fecha Documento',
        'vzla_capex_pago_tienda': 'Tienda',
        'vzla_capex_pago_sucursal': 'Sucursal',
        'vzla_capex_pago_monto': 'Monto',
        'vzla_capex_pago_moneda': 'Moneda',
        'vzla_capex_pago_fecha_vencimiento': 'Fecha Vencimiento',
        'vzla_capex_pago_cuenta': 'Cuenta',
        'vzla_capex_pago_id_cuenta': 'Id Cta',
        'vzla_capex_pago_metodo_pago': 'Método de Pago',
        'vzla_capex_pago_es_independiente': 'Pago Independiente',
        'vzla_capex_pago_prioridad': 'Prioridad',
        'vzla_capex_pago_monto_ext': 'Monto CAPEX EXT',
        'vzla_capex_pago_monto_ord': 'Monto CAPEX ORD',
        'vzla_capex_pago_monto_cadm': 'Monto CADM',
        'vzla_capex_pago_fecha_creacion': 'Fecha Creación',
        'vzla_capex_pago_solicitante': 'Solicitante',
        'vzla_capex_pago_monto_usd': 'Monto USD',
        'vzla_capex_pago_categoria': 'CATEGORIA',
        'vzla_capex_pago_monto_pagar_capex': 'MONTO A PAGAR CAPEX',
        'vzla_capex_pago_monto_pagar_opex': 'MONTO A PAGAR OPEX',
        'vzla_capex_pago_validacion': 'VALIDACION',
        'vzla_capex_pago_calcu_moneda': 'METODO DE PAGO',
        'vzla_capex_pago_semana_pago': 'SEMANA',
        'vzla_capex_pago_mes_pago': 'MES DE PAGO',
        'vzla_capex_pago_tipo_capex': 'TIPO DE CAPEX',
        'vzla_capex_pago_calcu_monto_ord': 'MONTO ORD',
        'vzla_capex_pago_calcu_monto_ext': 'MONTO EXT',
        'vzla_capex_pago_dia_pago': 'DIA DE PAGO',
        'vzla_capex_pago_calcu_tienda': 'TIENDA_LOOKUP',
        'vzla_capex_pago_ceco': 'CECO',
        'vzla_capex_pago_proyecto': 'PROYECTO',
        'vzla_capex_pago_area': 'AREA',
        'vzla_capex_pago_fecha_recibo': 'FECHA RECIBO',
        'vzla_capex_pago_descripcion': 'DESCRIPCIÓN',
        'vzla_capex_pago_current_fiscal_year': '_año_inicio',  # Temporal
        'vzla_capex_pago_next_fiscal_year': '_año_fin'          # Temporal
    }
    
    # Renombrar columnas que existen
    columnas_renombrar = {col_bq: col_excel for col_bq, col_excel in mapeo_inverso.items() if col_bq in df_bq.columns}
    df_excel = df_bq.rename(columns=columnas_renombrar)
    
    # ===================================================================
    # COMBINAR AÑO FISCAL (current_fiscal_year + next_fiscal_year)
    # ===================================================================
    if '_año_inicio' in df_excel.columns and '_año_fin' in df_excel.columns:
        print(f"🔗 Combinando años fiscales en 'AÑO FISCAL'...")
        
        def combinar_años(row):
            año_inicio = row.get('_año_inicio', '')
            año_fin = row.get('_año_fin', '')
            
            # Si ambos existen y son válidos
            if pd.notna(año_inicio) and pd.notna(año_fin):
                try:
                    return f"{int(año_inicio)}-{int(año_fin)}"
                except:
                    return f"{año_inicio}-{año_fin}"
            # Si solo hay uno, intentar calcularlo
            elif pd.notna(año_inicio):
                try:
                    año_inicio_int = int(año_inicio)
                    return f"{año_inicio_int}-{año_inicio_int + 1}"
                except:
                    return str(año_inicio)
            elif pd.notna(año_fin):
                try:
                    año_fin_int = int(año_fin)
                    return f"{año_fin_int - 1}-{año_fin_int}"
                except:
                    return str(año_fin)
            else:
                return "SIN_AÑO_FISCAL"
        
        df_excel['AÑO FISCAL'] = df_excel.apply(combinar_años, axis=1)
        
        # Eliminar columnas temporales
        df_excel = df_excel.drop(columns=['_año_inicio', '_año_fin'])
        
        print(f"   ✅ Columna 'AÑO FISCAL' creada")
    elif '_año_inicio' in df_excel.columns:
        # Si solo existe current_fiscal_year, crear año fiscal con +1
        print(f"⚠️ Solo existe 'current_fiscal_year', calculando año fiscal...")
        df_excel['AÑO FISCAL'] = df_excel['_año_inicio'].apply(
            lambda x: f"{int(x)}-{int(x) + 1}" if pd.notna(x) else "SIN_AÑO_FISCAL"
        )
        df_excel = df_excel.drop(columns=['_año_inicio'])
    
    # Mantener solo las columnas que están en el mapeo + AÑO FISCAL
    columnas_finales = [col for col in df_excel.columns if col in mapeo_inverso.values() or col == 'AÑO FISCAL']
    df_excel = df_excel[columnas_finales]
    
    return df_excel

def mapear_bigquery_a_excel_columns_colombia(df_bq: pd.DataFrame) -> pd.DataFrame:
    """Convertir nombres de columnas de BigQuery a nombres de Excel"""
    
    mapeo_inverso = {
        'col_capex_pago_numero_factura': 'Numero de Factura',
        'col_capex_pago_orden_compra': 'Numero de OC',
        'col_capex_pago_tipo_documento': 'Tipo Factura',
        'col_capex_pago_nombre_lote': 'Nombre Lote',
        'col_capex_pago_proveedor': 'Proveedor',
        'col_capex_pago_rif': 'RIF',
        'col_capex_pago_fecha_documento': 'Fecha Documento',
        'col_capex_pago_tienda': 'Tienda',
        'col_capex_pago_sucursal': 'Sucursal',
        'col_capex_pago_monto': 'Monto',
        'col_capex_pago_moneda': 'Moneda',
        'col_capex_pago_fecha_vencimiento': 'Fecha Vencimiento',
        'col_capex_pago_cuenta': 'Cuenta',
        'col_capex_pago_id_cuenta': 'Id Cta',
        'col_capex_pago_metodo_pago': 'Método de Pago',
        'col_capex_pago_es_independiente': 'Pago Independiente',
        'col_capex_pago_prioridad': 'Prioridad',
        'col_capex_pago_monto_ext': 'Monto CAPEX EXT',
        'col_capex_pago_monto_ord': 'Monto CAPEX ORD',
        'col_capex_pago_monto_cadm': 'Monto CADM',
        'col_capex_pago_fecha_creacion': 'Fecha Creación',
        'col_capex_pago_solicitante': 'Solicitante',
        'col_capex_pago_monto_usd': 'Monto USD',
        'col_capex_pago_categoria': 'CATEGORIA',
        'col_capex_pago_monto_pagar_capex': 'MONTO A PAGAR CAPEX',
        'col_capex_pago_monto_pagar_opex': 'MONTO A PAGAR OPEX',
        'col_capex_pago_validacion': 'VALIDACION',
        'col_capex_pago_calcu_moneda': 'METODO DE PAGO',
        'col_capex_pago_semana_pago': 'SEMANA',
        'col_capex_pago_mes_pago': 'MES DE PAGO',
        'col_capex_pago_tipo_capex': 'TIPO DE CAPEX',
        'col_capex_pago_calcu_monto_ord': 'MONTO ORD',
        'col_capex_pago_calcu_monto_ext': 'MONTO EXT',
        'col_capex_pago_dia_pago': 'DIA DE PAGO',
        'col_capex_pago_calcu_tienda': 'TIENDA_LOOKUP',
        'col_capex_pago_ceco': 'CECO',
        'col_capex_pago_proyecto': 'PROYECTO',
        'col_capex_pago_area': 'AREA',
        'col_capex_pago_fecha_recibo': 'FECHA RECIBO',
        'col_capex_pago_descripcion': 'DESCRIPCIÓN',
        'col_capex_pago_current_fiscal_year': '_año_inicio',  # Temporal
        'col_capex_pago_next_fiscal_year': '_año_fin'          # Temporal
    }
    
    # Renombrar columnas que existen
    columnas_renombrar = {col_bq: col_excel for col_bq, col_excel in mapeo_inverso.items() if col_bq in df_bq.columns}
    df_excel = df_bq.rename(columns=columnas_renombrar)
    
    # ===================================================================
    # COMBINAR AÑO FISCAL (current_fiscal_year + next_fiscal_year)
    # ===================================================================
    if '_año_inicio' in df_excel.columns and '_año_fin' in df_excel.columns:
        print(f"🔗 Combinando años fiscales en 'AÑO FISCAL'...")
        
        def combinar_años(row):
            año_inicio = row.get('_año_inicio', '')
            año_fin = row.get('_año_fin', '')
            
            # Si ambos existen y son válidos
            if pd.notna(año_inicio) and pd.notna(año_fin):
                try:
                    return f"{int(año_inicio)}-{int(año_fin)}"
                except:
                    return f"{año_inicio}-{año_fin}"
            # Si solo hay uno, intentar calcularlo
            elif pd.notna(año_inicio):
                try:
                    año_inicio_int = int(año_inicio)
                    return f"{año_inicio_int}-{año_inicio_int + 1}"
                except:
                    return str(año_inicio)
            elif pd.notna(año_fin):
                try:
                    año_fin_int = int(año_fin)
                    return f"{año_fin_int - 1}-{año_fin_int}"
                except:
                    return str(año_fin)
            else:
                return "SIN_AÑO_FISCAL"
        
        df_excel['AÑO FISCAL'] = df_excel.apply(combinar_años, axis=1)
        
        # Eliminar columnas temporales
        df_excel = df_excel.drop(columns=['_año_inicio', '_año_fin'])
        
        print(f"   ✅ Columna 'AÑO FISCAL' creada")
    elif '_año_inicio' in df_excel.columns:
        # Si solo existe current_fiscal_year, crear año fiscal con +1
        print(f"⚠️ Solo existe 'current_fiscal_year', calculando año fiscal...")
        df_excel['AÑO FISCAL'] = df_excel['_año_inicio'].apply(
            lambda x: f"{int(x)}-{int(x) + 1}" if pd.notna(x) else "SIN_AÑO_FISCAL"
        )
        df_excel = df_excel.drop(columns=['_año_inicio'])
    
    # Mantener solo las columnas que están en el mapeo + AÑO FISCAL
    columnas_finales = [col for col in df_excel.columns if col in mapeo_inverso.values() or col == 'AÑO FISCAL']
    df_excel = df_excel[columnas_finales]
    
    return df_excel


def ajustar_df_a_schema_bigquery_venezuela(df, client, dataset_id, table_id):
    """
    Convierte DataFrame a los tipos esperados por el schema de BigQuery (en vivo).
    
    Args:
        df: DataFrame a convertir
        client: bigquery.Client conectado
        dataset_id, table_id: nombres de dataset y tabla en BigQuery
    
    Returns:
        DataFrame listo para cargar en BQ
    """
    # Obtén el schema BigQuery en vivo
    tabla = client.get_table(f"{dataset_id}.{table_id}")
    schema = {field.name: field.field_type for field in tabla.schema}
    print("🔍 Esquema BigQuery de la tabla:")
    for k, v in schema.items():
        print(f" - {k}: {v}")

    df2 = df.copy()
    
    # Verificar y eliminar columnas que no están en el schema de BigQuery
    columnas_no_schema = [col for col in df2.columns if col not in schema]
    if columnas_no_schema:
        print(f"⚠️  Eliminando columnas que no están en el schema: {columnas_no_schema}", flush=True)
        df2 = df2.drop(columns=columnas_no_schema)
    
    # Verificar especialmente si hay una columna "area" que no debería estar
    if 'area' in df2.columns and 'vzla_capex_diferencia_area' in df2.columns:
        print(f"⚠️  Advertencia: Se encontró columna 'area' además de 'vzla_capex_diferencia_area'. Eliminando 'area'.", flush=True)
        df2 = df2.drop(columns=['area'])
    elif 'area' in df2.columns:
        print(f"❌ Error: Se encontró columna 'area' pero no 'vzla_capex_diferencia_area'. Esto no debería pasar.", flush=True)
        df2 = df2.drop(columns=['area'])
    
    for col, tipo in schema.items():
        if col not in df2.columns:
            print(f"⚠️  Columna '{col}' no está en el DataFrame, se salta.", flush=True)
            continue
        
        # Mostrar información de la columna antes de convertir
        valores_unicos = df2[col].dropna().unique()[:5]  # Primeros 5 valores únicos
        print(f"🔄 Convirtiendo columna '{col}' a {tipo}...", flush = True)
        print(f"   Tipo actual: {df2[col].dtype}", flush = True)
        print(f"   Valores de ejemplo: {valores_unicos}", flush = True)
        
        try:
            # STRING
            if tipo == "STRING":
                df2[col] = df2[col].astype(str)
            # INTEGER
            elif tipo in ["INTEGER", "INT64"]:
                # Intentar convertir a numérico primero
                df2[col] = pd.to_numeric(df2[col], errors='coerce', downcast='integer')
                # Verificar si hay valores no convertidos (NaN que no eran NaN originalmente)
                valores_no_convertidos = df2[df2[col].isna() & df[col].notna()]
                if len(valores_no_convertidos) > 0:
                    print(f"   ⚠️  {len(valores_no_convertidos)} valores no pudieron convertirse a INTEGER", flush = True)
                    print(f"   Valores problemáticos: {valores_no_convertidos[col].unique()[:10]}", flush = True)
                    # Convertir a 0 o mantener como string según el caso
                    df2[col] = df2[col].fillna(0).astype('Int64')  # Int64 permite NaN
            # FLOAT
            elif tipo in ["FLOAT", "FLOAT64", "NUMERIC"]:
                df2[col] = pd.to_numeric(df2[col], errors='coerce')
            # BOOLEAN
            elif tipo == "BOOLEAN":
                df2[col] = df2[col].astype('bool')
            # DATE/TIMESTAMP/DATETIME
            elif tipo in ["DATE", "TIMESTAMP", "DATETIME"]:
                # Manejo especial para vzla_capex_diferencia_mes que viene en formato 'NOV-25'
                if col == 'vzla_capex_diferencia_mes':
                    # Convertir formato 'NOV-25' a fecha (primer día del mes)
                    def convertir_mes_año_a_fecha(mes_anio_str):
                        if pd.isna(mes_anio_str):
                            return pd.NaT
                        try:
                            # Formato: 'NOV-25' -> convertir a fecha del primer día del mes
                            mes_anio_str = str(mes_anio_str).strip().upper()
                            if '-' in mes_anio_str:
                                partes = mes_anio_str.split('-')
                                mes_abrev = partes[0]  # 'NOV'
                                anio_str = partes[1]   # '25'
                                
                                # Mapeo de meses abreviados en inglés
                                meses_map = {
                                    'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
                                    'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12,
                                    'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
                                    'JUL': 7, 'AGO': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12
                                }
                                
                                mes_num = meses_map.get(mes_abrev)
                                if mes_num:
                                    # Convertir año de 2 dígitos a 4 dígitos (asumir 2000-2099)
                                    anio = int('20' + anio_str) if len(anio_str) == 2 else int(anio_str)
                                    return pd.Timestamp(year=anio, month=mes_num, day=1)
                        except Exception as e:
                            print(f"   ⚠️  Error convirtiendo '{mes_anio_str}' a fecha: {e}", flush=True)
                        return pd.NaT
                    
                    df2[col] = df2[col].apply(convertir_mes_año_a_fecha)
                    print(f"   ✅ Convertido formato 'MES-AA' a DATE (primer día del mes)", flush=True)
                else:
                    # Para otras columnas de fecha, usar conversión estándar
                    df2[col] = pd.to_datetime(df2[col], errors='coerce', format='mixed')
            # Repeated or RECORD types require special custom handling
            else:
                print(f"⚠️  Tipo no manejado automáticamente: {tipo} (col: {col})", flush = True)
        except Exception as e:
            print(f"❌ Error convirtiendo columna '{col}' a {tipo}: {e}", flush = True)
            print(f"   Valores problemáticos: {df2[col].dropna().unique()[:10]}", flush = True)
            raise
    
    print("✅ DataFrame transformado según schema BigQuery", flush = True)
    return df2


def ajustar_df_a_schema_bigquery_colombia(df, client, dataset_id, table_id):
    """
    Convierte DataFrame a los tipos esperados por el schema de BigQuery (en vivo).
    
    Args:
        df: DataFrame a convertir
        client: bigquery.Client conectado
        dataset_id, table_id: nombres de dataset y tabla en BigQuery
    
    Returns:
        DataFrame listo para cargar en BQ
    """
    # Obtén el schema BigQuery en vivo
    tabla = client.get_table(f"{dataset_id}.{table_id}")
    schema = {field.name: field.field_type for field in tabla.schema}
    print("🔍 Esquema BigQuery de la tabla:")
    for k, v in schema.items():
        print(f" - {k}: {v}")

    df2 = df.copy()
    
    # Verificar y eliminar columnas que no están en el schema de BigQuery
    columnas_no_schema = [col for col in df2.columns if col not in schema]
    if columnas_no_schema:
        print(f"⚠️  Eliminando columnas que no están en el schema: {columnas_no_schema}", flush=True)
        df2 = df2.drop(columns=columnas_no_schema)
    
    # Verificar especialmente si hay una columna "area" que no debería estar
    if 'area' in df2.columns and 'col_capex_diferencia_area' in df2.columns:
        print(f"⚠️  Advertencia: Se encontró columna 'area' además de 'col_capex_diferencia_area'. Eliminando 'area'.", flush=True)
        df2 = df2.drop(columns=['area'])
    elif 'area' in df2.columns:
        print(f"❌ Error: Se encontró columna 'area' pero no 'col_capex_diferencia_area'. Esto no debería pasar.", flush=True)
        df2 = df2.drop(columns=['area'])
    
    for col, tipo in schema.items():
        if col not in df2.columns:
            print(f"⚠️  Columna '{col}' no está en el DataFrame, se salta.", flush=True)
            continue
        
        # Mostrar información de la columna antes de convertir
        valores_unicos = df2[col].dropna().unique()[:5]  # Primeros 5 valores únicos
        print(f"🔄 Convirtiendo columna '{col}' a {tipo}...", flush = True)
        print(f"   Tipo actual: {df2[col].dtype}", flush = True)
        print(f"   Valores de ejemplo: {valores_unicos}", flush = True)
        
        try:
            # STRING
            if tipo == "STRING":
                df2[col] = df2[col].astype(str)
            # INTEGER
            elif tipo in ["INTEGER", "INT64"]:
                # Intentar convertir a numérico primero
                df2[col] = pd.to_numeric(df2[col], errors='coerce', downcast='integer')
                # Verificar si hay valores no convertidos (NaN que no eran NaN originalmente)
                valores_no_convertidos = df2[df2[col].isna() & df[col].notna()]
                if len(valores_no_convertidos) > 0:
                    print(f"   ⚠️  {len(valores_no_convertidos)} valores no pudieron convertirse a INTEGER", flush = True)
                    print(f"   Valores problemáticos: {valores_no_convertidos[col].unique()[:10]}", flush = True)
                    # Convertir a 0 o mantener como string según el caso
                    df2[col] = df2[col].fillna(0).astype('Int64')  # Int64 permite NaN
            # FLOAT
            elif tipo in ["FLOAT", "FLOAT64", "NUMERIC"]:
                df2[col] = pd.to_numeric(df2[col], errors='coerce')
            # BOOLEAN
            elif tipo == "BOOLEAN":
                df2[col] = df2[col].astype('bool')
            # DATE/TIMESTAMP/DATETIME
            elif tipo in ["DATE", "TIMESTAMP", "DATETIME"]:
                # Manejo especial para col_capex_diferencia_mes que viene en formato 'NOV-25'
                if col == 'col_capex_diferencia_mes':
                    # Convertir formato 'NOV-25' a fecha (primer día del mes)
                    def convertir_mes_año_a_fecha(mes_anio_str):
                        if pd.isna(mes_anio_str):
                            return pd.NaT
                        try:
                            # Formato: 'NOV-25' -> convertir a fecha del primer día del mes
                            mes_anio_str = str(mes_anio_str).strip().upper()
                            if '-' in mes_anio_str:
                                partes = mes_anio_str.split('-')
                                mes_abrev = partes[0]  # 'NOV'
                                anio_str = partes[1]   # '25'
                                
                                # Mapeo de meses abreviados en inglés
                                meses_map = {
                                    'JAN': 1, 'FEB': 2, 'MAR': 3, 'APR': 4, 'MAY': 5, 'JUN': 6,
                                    'JUL': 7, 'AUG': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12,
                                    'ENE': 1, 'FEB': 2, 'MAR': 3, 'ABR': 4, 'MAY': 5, 'JUN': 6,
                                    'JUL': 7, 'AGO': 8, 'SEP': 9, 'OCT': 10, 'NOV': 11, 'DIC': 12
                                }
                                
                                mes_num = meses_map.get(mes_abrev)
                                if mes_num:
                                    # Convertir año de 2 dígitos a 4 dígitos (asumir 2000-2099)
                                    anio = int('20' + anio_str) if len(anio_str) == 2 else int(anio_str)
                                    return pd.Timestamp(year=anio, month=mes_num, day=1)
                        except Exception as e:
                            print(f"   ⚠️  Error convirtiendo '{mes_anio_str}' a fecha: {e}", flush=True)
                        return pd.NaT
                    
                    df2[col] = df2[col].apply(convertir_mes_año_a_fecha)
                    print(f"   ✅ Convertido formato 'MES-AA' a DATE (primer día del mes)", flush=True)
                else:
                    # Para otras columnas de fecha, usar conversión estándar
                    df2[col] = pd.to_datetime(df2[col], errors='coerce', format='mixed')
            # Repeated or RECORD types require special custom handling
            else:
                print(f"⚠️  Tipo no manejado automáticamente: {tipo} (col: {col})", flush = True)
        except Exception as e:
            print(f"❌ Error convirtiendo columna '{col}' a {tipo}: {e}", flush = True)
            print(f"   Valores problemáticos: {df2[col].dropna().unique()[:10]}", flush = True)
            raise
    
    print("✅ DataFrame transformado según schema BigQuery", flush = True)
    return df2

# =================== CLIENTE GOOGLE CLOUD STORAGE ===================

def crear_cliente_storage():
    """Crear cliente de Google Cloud Storage"""
    try:
        # Si hay un archivo de credenciales especificado y existe, usarlo
        if CREDENTIALS_FILE and os.path.exists(CREDENTIALS_FILE):
            credentials = service_account.Credentials.from_service_account_file(
                CREDENTIALS_FILE,
                scopes=["https://www.googleapis.com/auth/cloud-platform"]
            )
            client = storage.Client(credentials=credentials, project=GCP_PROJECT_ID)
            print(f"✅ Cliente GCS creado (usando archivo de credenciales)")
            return client
        else:
            # Usar Application Default Credentials (ADC) - funciona en Cloud Run, GCE, etc.
            client = storage.Client(project=GCP_PROJECT_ID)
            print(f"✅ Cliente GCS creado con ADC para proyecto: {GCP_PROJECT_ID}")
            return client
    except Exception as e:
        print(f"❌ Error creando cliente GCS: {e}")
        raise

# =================== FUNCIONES AUXILIARES GCS ===================

def listar_archivos_bucket(storage_client: storage.Client, bucket_name: str, limit: int = 10) -> List[Dict]:
    """
    Listar archivos en el bucket
    
    Args:
        storage_client: Cliente de GCS
        bucket_name: Nombre del bucket
        limit: Límite de archivos a listar
    
    Returns:
        list: Lista de diccionarios con info de los archivos
    """
    try:
        bucket = storage_client.bucket(bucket_name)
        blobs = list(bucket.list_blobs(max_results=limit))
        
        archivos = []
        for blob in blobs:
            archivos.append({
                'name': blob.name,
                'size_mb': round(blob.size / (1024 * 1024), 2),
                'created': blob.time_created.isoformat() if blob.time_created else None,
                'updated': blob.updated.isoformat() if blob.updated else None,
                'public_url': blob.public_url,
                'content_type': blob.content_type
            })
        
        return archivos
        
    except Exception as e:
        print(f"❌ Error listando archivos: {e}")
        return []

# def eliminar_archivo_gcs(storage_client: storage.Client, bucket_name: str, blob_name: str) -> bool:
#     """
#     Eliminar un archivo específico del bucket
    
#     Args:
#         storage_client: Cliente de GCS
#         bucket_name: Nombre del bucket
#         blob_name: Nombre del blob a eliminar
    
#     Returns:
#         bool: True si se eliminó exitosamente
#     """
#     try:
#         bucket = storage_client.bucket(bucket_name)
#         blob = bucket.blob(blob_name)
#         blob.delete()
        
#         print(f"✅ Archivo eliminado: {blob_name}")
#         return True
        
#     except Exception as e:
#         print(f"❌ Error eliminando archivo: {e}")
#         return False


# =================== ENDPOINTS API ===================

@app.route('/health', methods=['GET'])
def health_check():
    """Endpoint de health check"""
    return jsonify({
        'status': 'healthy',
        'service': 'CAPEX BigQuery API',
        'version': '1.0',
        'timestamp': datetime.now().isoformat()
    })


def limpiar_carpeta_tmp_gcs(storage_client: storage.Client) -> int:
    """
    Eliminar todos los archivos de la carpeta tmp/ en GCS.
    Returns: Cantidad de archivos eliminados
    """
    try:
        print(f"🗑️ Limpiando carpeta tmp/ en GCS...")
        
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blobs = bucket.list_blobs(prefix='tmp/')
        
        count = 0
        for blob in blobs:
            blob.delete()
            print(f"   ✅ Eliminado: {blob.name}")
            count += 1
        
        if count == 0:
            print(f"   ℹ️ Carpeta tmp/ ya estaba vacía")
        else:
            print(f"✅ {count} archivo(s) eliminado(s) de tmp/")
        
        return count
        
    except Exception as e:
        print(f"⚠️ Error limpiando carpeta tmp/: {e}")
        return 0


def subir_archivo_a_gcs_tmp(storage_client: storage.Client, archivo_local: str, pais: str) -> tuple:
    """
    Subir archivo a Google Cloud Storage en carpeta tmp/
    Returns: (url_publica, nombre_blob)
    """
    try:
        # Generar nombre con timestamp
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        nombre_blob = f"tmp/Bosqueto_{pais.upper()}_{timestamp}.xlsx"
        
        print(f"📤 Subiendo archivo a GCS (tmp): {nombre_blob}")
        
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(nombre_blob)
        
        # Subir archivo
        blob.upload_from_filename(archivo_local, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Hacer el blob público para que sea accesible sin autenticación
        try:
            blob.make_public()
            print(f"   ✓ Archivo configurado como público")
        except Exception as public_error:
            print(f"   ⚠️ No se pudo hacer el archivo público: {str(public_error)}")
        
        # Usar URL pública del blob
        url_publica = blob.public_url
        if not url_publica or 'storage.googleapis.com' not in url_publica:
            url_publica = f"https://storage.googleapis.com/{GCS_BUCKET_NAME}/{nombre_blob}"
        
        print(f"✅ Archivo subido exitosamente a tmp/")
        print(f"   URL pública: {url_publica}")
        
        return url_publica, nombre_blob
        
    except Exception as e:
        print(f"❌ Error subiendo archivo a GCS tmp: {e}")
        raise


def subir_archivo_a_gcs_logs(storage_client: storage.Client, archivo_local: str, pais: str) -> tuple:
    """
    Subir archivo a Google Cloud Storage en carpeta logs/{fecha_caracas}/
    Returns: (url_publica, nombre_blob)
    """
    try:
        # Obtener fecha actual en zona horaria de Caracas
        tz_caracas = ZoneInfo('America/Caracas')
        fecha_caracas = datetime.now(tz_caracas)
        fecha_str = fecha_caracas.strftime('%Y-%m-%d')
        timestamp = fecha_caracas.strftime('%Y-%m-%d_%H-%M-%S')
        
        nombre_blob = f"logs/{fecha_str}/Consolidado_{pais.upper()}_{timestamp}.xlsx"
        
        print(f"📤 Subiendo archivo a GCS (logs): {nombre_blob}")
        print(f"   📅 Fecha Caracas: {fecha_caracas.strftime('%Y-%m-%d %H:%M:%S')}")
        
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(nombre_blob)
        
        # Subir archivo
        blob.upload_from_filename(archivo_local, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Hacer el blob público
        try:
            blob.make_public()
            print(f"   ✓ Archivo configurado como público")
        except Exception as public_error:
            print(f"   ⚠️ No se pudo hacer el archivo público: {str(public_error)}")
        
        # Usar URL pública del blob
        url_publica = blob.public_url
        if not url_publica or 'storage.googleapis.com' not in url_publica:
            url_publica = f"https://storage.googleapis.com/{GCS_BUCKET_NAME}/{nombre_blob}"
        
        print(f"✅ Archivo subido exitosamente a logs/{fecha_str}/")
        print(f"   URL pública: {url_publica}")
        
        return url_publica, nombre_blob
        
    except Exception as e:
        print(f"❌ Error subiendo archivo a GCS logs: {e}")
        raise


def descargar_plantilla_gcs(storage_client: storage.Client, pais: str) -> str:
    """
    Descargar plantilla desde GCS según el país.
    Returns: Ruta del archivo local descargado
    """
    try:
        # Mapeo de país a ruta de plantilla
        plantillas = {
            'venezuela': 'template/vzla/consolidado_capex_ve_2025_2026_template.xlsx',
            'colombia': 'template/col/consolidado_capex_col_2025_2026_template.xlsx',  # Ajustar según nombre real
        }
        
        ruta_plantilla = plantillas.get(pais.lower())
        if not ruta_plantilla:
            raise ValueError(f"No hay plantilla configurada para el país: {pais}")
        
        print(f"📥 Descargando plantilla: {ruta_plantilla}")
        
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(ruta_plantilla)
        
        # Crear archivo temporal
        temp_file = f"/tmp/plantilla_{pais}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        blob.download_to_filename(temp_file)
        
        print(f"✅ Plantilla descargada: {temp_file}")
        return temp_file
        
    except Exception as e:
        print(f"❌ Error descargando plantilla: {e}")
        raise


def pegar_datos_en_plantilla(archivo_plantilla: str, df_bosqueto: pd.DataFrame, df_detalle: pd.DataFrame) -> str:
    """
    Pegar datos de BOSQUETO y DETALLE en las hojas correspondientes de la plantilla.
    La plantilla ya tiene los headers, solo se pegan los datos desde la fila 2.
    Returns: Ruta del archivo con los datos pegados
    """
    from openpyxl import load_workbook
    
    try:
        print(f"📝 Cargando plantilla Excel...", flush=True)
        wb = load_workbook(archivo_plantilla)
        print(f"   ✅ Plantilla cargada", flush=True)
        
        # Función auxiliar para verificar si un valor está vacío
        def es_valor_vacio(val):
            if pd.isna(val):
                return True
            if isinstance(val, str) and val.lower() in ('nan', 'none', 'null', ''):
                return True
            return False
        
        # Pegar BOSQUETO (datos empiezan en fila 2, headers ya están en la plantilla)
        print(f"   📋 Pegando datos en BOSQUETO ({len(df_bosqueto)} filas)...", flush=True)
        if 'BOSQUETO' in wb.sheetnames:
            ws_bosqueto = wb['BOSQUETO']
            
            for row_idx, row in enumerate(df_bosqueto.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    # Solo escribir si tiene valor real (no NaN, no "nan", no vacío)
                    if not es_valor_vacio(value):
                        ws_bosqueto.cell(row=row_idx, column=col_idx, value=value)
            
            print(f"   ✅ BOSQUETO: {len(df_bosqueto)} filas pegadas", flush=True)
        else:
            print(f"   ⚠️ Hoja 'BOSQUETO' no encontrada", flush=True)
        
        # Pegar DETALLE CORREGIDO (datos empiezan en fila 2, headers ya están en la plantilla)
        print(f"   📋 Pegando datos en Detalle Corregido ({len(df_detalle)} filas)...", flush=True)
        if 'Detalle Corregido' in wb.sheetnames:
            ws_detalle = wb['Detalle Corregido']
            
            for row_idx, row in enumerate(df_detalle.itertuples(index=False), 2):
                for col_idx, value in enumerate(row, 1):
                    # Solo escribir si tiene valor real (no NaN, no "nan", no vacío)
                    if not es_valor_vacio(value):
                        ws_detalle.cell(row=row_idx, column=col_idx, value=value)
            
            print(f"   ✅ Detalle Corregido: {len(df_detalle)} filas pegadas", flush=True)
        else:
            print(f"   ⚠️ Hoja 'Detalle Corregido' no encontrada", flush=True)
        
        # Guardar
        print(f"   💾 Guardando archivo...", flush=True)
        wb.save(archivo_plantilla)
        print(f"✅ Plantilla guardada", flush=True)
        
        return archivo_plantilla
        
    except Exception as e:
        print(f"❌ Error pegando datos en plantilla: {e}", flush=True)
        traceback.print_exc()
        raise


@app.route('/api/v1/procesar-detalle', methods=['POST'])
def procesar_detalle():
    """
    Endpoint para procesar DETALLE.
    Recibe el BOSQUETO (modificado por el usuario), carga a BigQuery,
    extrae DETALLE de BQ, y pega ambos en la plantilla.
    """
    try:
        print(f"\n{'='*70}", flush=True)
        print(f"🚀 INICIANDO PROCESAR DETALLE", flush=True)
        print(f"{'='*70}", flush=True)
        
        # Validar archivo
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No se proporcionó archivo BOSQUETO'
            }), 400
        
        file = request.files['file']
        
        if file.filename == '' or not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'error': 'Archivo inválido'
            }), 400
        
        # Obtener país (obligatorio)
        pais = request.form.get('pais', 'venezuela')
        
        print(f"🌎 País: {pais.upper()}", flush=True)
        print(f"📁 Archivo BOSQUETO recibido: {file.filename}", flush=True)

        # PASO 0: Limpiar carpetas tmp (local y GCS)
        print(f"\n🗑️ PASO 0: Limpiando carpetas tmp...", flush=True)
        
        # Limpiar /tmp local
        print(f"   🧹 Limpiando /tmp local...", flush=True)
        try:
            import glob
            archivos_tmp = glob.glob('/tmp/*.xlsx') + glob.glob('/tmp/plantilla_*.xlsx') + glob.glob('/tmp/bosqueto_*.xlsx') + glob.glob('/tmp/reporte_*.xlsx')
            for archivo in archivos_tmp:
                try:
                    os.remove(archivo)
                    print(f"      ✅ Eliminado: {os.path.basename(archivo)}", flush=True)
                except Exception as e:
                    print(f"      ⚠️ No se pudo eliminar {archivo}: {e}", flush=True)
            if not archivos_tmp:
                print(f"      ℹ️ Carpeta /tmp local ya estaba limpia", flush=True)
        except Exception as e:
            print(f"      ⚠️ Error limpiando /tmp local: {e}", flush=True)
        
        # Limpiar tmp/ en GCS
        print(f"   🧹 Limpiando tmp/ en GCS...", flush=True)
        storage_client = crear_cliente_storage()
        limpiar_carpeta_tmp_gcs(storage_client)
        print(f"✅ PASO 0 COMPLETADO: Carpetas tmp limpiadas", flush=True)

        # Guardar archivo temporalmente
        temp_bosqueto = f"/tmp/bosqueto_upload_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file.save(temp_bosqueto)
        print(f"💾 Archivo guardado temporalmente: {temp_bosqueto}", flush=True)
        
        # PASO 1: Leer BOSQUETO
        print(f"\n📖 PASO 1: LEYENDO BOSQUETO...", flush=True)
        df_bosqueto = pd.read_excel(temp_bosqueto, sheet_name='BOSQUETO')
        print(f"✅ PASO 1 COMPLETADO: {len(df_bosqueto)} filas, {len(df_bosqueto.columns)} columnas", flush=True)
        
        # PASO 2: Mapear columnas para BigQuery
        print(f"\n🔄 PASO 2: Mapeando columnas para BigQuery...", flush=True)
        if pais.lower() == 'venezuela':
            df_mapped = mapear_columnas_bosqueto_a_bigquery_venezuela(df_bosqueto)
        elif pais.lower() == 'colombia':
            df_mapped = mapear_columnas_bosqueto_a_bigquery_colombia(df_bosqueto)
        else:
            return jsonify({
                'success': False,
                'error': f'País "{pais}" no soportado'
            }), 400
        print(f"✅ PASO 2 COMPLETADO: {len(df_mapped.columns)} columnas mapeadas", flush=True)
        
        # PASO 3: Crear clientes GCP
        # PASO 3: Crear cliente BigQuery (Storage ya creado en PASO 0)
        print(f"\n🔧 PASO 3: Creando cliente BigQuery...", flush=True)
        bq_client = crear_cliente_bigquery()
        print(f"✅ PASO 3 COMPLETADO: Cliente BigQuery creado", flush=True)
        
        # PASO 4: Cargar a BigQuery (con verificación de duplicados)
        print(f"\n📤 PASO 4: Cargando a BigQuery...", flush=True)
        if pais.lower() == 'venezuela':
            resultado_carga = cargar_datos_a_bigquery_venezuela(bq_client, df_mapped)
        elif pais.lower() == 'colombia':
            resultado_carga = cargar_datos_a_bigquery_colombia(bq_client, df_mapped)
        
        if not resultado_carga['success']:
            print(f"❌ PASO 4 FALLÓ: {resultado_carga.get('error', 'Error desconocido')}", flush=True)
            if 'df_cargados' in resultado_carga:
                resultado_carga.pop('df_cargados')
            return jsonify(resultado_carga), 500
        
        print(f"✅ PASO 4 COMPLETADO: {resultado_carga['rows_loaded']} cargados, {resultado_carga['rows_duplicated']} duplicados", flush=True)
        
        # PASO 5: Extraer DETALLE CORREGIDO de BigQuery
        print(f"\n📋 PASO 5: Extrayendo DETALLE CORREGIDO de BigQuery...", flush=True)
        if pais.lower() == 'venezuela':
            df_bigquery = extraer_tabla_completa_por_lotes_venezuela(bq_client)
        elif pais.lower() == 'colombia':
            df_bigquery = extraer_tabla_completa_por_lotes_colombia(bq_client)
        
        print(f"   📊 Datos extraídos de BQ: {len(df_bigquery)} filas", flush=True)
        
        if not df_bigquery.empty:
            print(f"   🔄 Mapeando columnas de BQ a Excel...", flush=True)
            if pais.lower() == 'venezuela':
                df_detalle_corregido = mapear_bigquery_a_excel_columns_venezuela(df_bigquery)
            elif pais.lower() == 'colombia':
                df_detalle_corregido = mapear_bigquery_a_excel_columns_colombia(df_bigquery)
            print(f"✅ PASO 5 COMPLETADO: {len(df_detalle_corregido)} filas en DETALLE", flush=True)
        else:
            df_detalle_corregido = pd.DataFrame()
            print(f"⚠️ PASO 5 COMPLETADO: DETALLE vacío (sin registros en BigQuery)", flush=True)
        
        # PASO 6: Descargar plantilla
        print(f"\n📥 PASO 6: Descargando plantilla de GCS...", flush=True)
        archivo_plantilla = descargar_plantilla_gcs(storage_client, pais)
        print(f"✅ PASO 6 COMPLETADO: Plantilla descargada", flush=True)
        
        # PASO 7: Pegar datos en plantilla
        print(f"\n📝 PASO 7: Pegando datos en plantilla...", flush=True)
        archivo_final = pegar_datos_en_plantilla(archivo_plantilla, df_bosqueto, df_detalle_corregido)
        print(f"✅ PASO 7 COMPLETADO: Datos pegados en plantilla", flush=True)
        
        # PASO 8: Subir a GCS (carpeta logs/{fecha_caracas}/)
        print(f"\n☁️ PASO 8: Subiendo a Google Cloud Storage (logs)...", flush=True)
        url_descarga, nombre_archivo_gcs = subir_archivo_a_gcs_logs(storage_client, archivo_final, pais)
        print(f"✅ PASO 8 COMPLETADO: Archivo subido", flush=True)
        
        # Limpiar archivos temporales
        print(f"\n🧹 Limpiando archivos temporales...", flush=True)
        archivos_temp = [temp_bosqueto, archivo_plantilla]
        
        for archivo in archivos_temp:
            try:
                if archivo and os.path.exists(archivo):
                    os.remove(archivo)
                    print(f"   ✅ Eliminado: {os.path.basename(archivo)}", flush=True)
            except Exception as e:
                print(f"   ⚠️ No se pudo eliminar {archivo}: {e}", flush=True)
        
        # Respuesta final
        respuesta = {
            'success': True,
            'pais': pais.upper(),
            'total_rows': resultado_carga['total_rows'],
            'rows_duplicated': resultado_carga['rows_duplicated'],
            'rows_loaded': resultado_carga['rows_loaded'],
            'detalle_rows': len(df_detalle_corregido),
            'consolidado_url': url_descarga,
            'file_name': nombre_archivo_gcs,
            'timestamp': datetime.now().isoformat(),
            'message': f"Proceso completado: {resultado_carga['rows_loaded']} registros cargados a BQ, {len(df_detalle_corregido)} filas en DETALLE"
        }
        
        print(f"\n{'='*70}", flush=True)
        print(f"✅ PROCESO DETALLE COMPLETADO EXITOSAMENTE", flush=True)
        print(f"   País: {pais.upper()}", flush=True)
        print(f"   URL: {url_descarga}", flush=True)
        print(f"{'='*70}", flush=True)
        
        return jsonify(respuesta), 200
        
    except Exception as e:
        print(f"❌ Error en proceso DETALLE: {e}", flush=True)
        traceback.print_exc()

        try:
            if 'temp_bosqueto' in locals() and os.path.exists(temp_bosqueto):
                os.remove(temp_bosqueto)
            if 'archivo_plantilla' in locals() and os.path.exists(archivo_plantilla):
                os.remove(archivo_plantilla)
        except:
            pass
        
        return jsonify({
            'success': False,
            'error': str(e),
            'message': f'Error procesando DETALLE: {str(e)}'
        }), 500


@app.route('/api/v1/procesar-bosqueto', methods=['POST'])
def procesar_bosqueto():
    """
    Endpoint para procesar SOLO el BOSQUETO.
    Recibe los archivos, genera la hoja BOSQUETO, sube a GCS/tmp y retorna el Excel.
    NO carga datos a BigQuery.
    """
    try:
        # Validar archivo
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No se proporcionó archivo'
            }), 400
        
        file = request.files['file']
        
        if file.filename == '' or not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'error': 'Archivo inválido'
            }), 400
        
        # Obtener país (opcional, default venezuela)
        pais = request.form.get('pais', 'venezuela')
        
        print(f"\n{'='*70}")
        print(f"🚀 PROCESAR BOSQUETO - {pais.upper()}")
        print(f"{'='*70}")
        print(f"📁 Archivo recibido: {file.filename}")

        # PASO 0: Limpiar carpeta tmp/ en GCS
        storage_client = crear_cliente_storage()
        limpiar_carpeta_tmp_gcs(storage_client)

        # Guardar archivo principal temporalmente
        temp_reporte_pago = f"/tmp/reporte_pago_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file.save(temp_reporte_pago)
        
        # Verificar si hay Reporte Absoluto (opcional)
        temp_reporte_absoluto = None
        if 'reporte_absoluto' in request.files:
            reporte_absoluto_file = request.files['reporte_absoluto']
            if reporte_absoluto_file.filename != '':
                temp_reporte_absoluto = f"/tmp/reporte_absoluto_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                reporte_absoluto_file.save(temp_reporte_absoluto)
                print(f"📋 Reporte Absoluto recibido: {reporte_absoluto_file.filename}")
        else:
            print(f"ℹ️ No se proporcionó Reporte Absoluto")

        # PASO 1: PROCESAR Y GENERAR BOSQUETO 
        print(f"\n{'='*70}")
        print(f"🔄 PASO 1: PROCESANDO ARCHIVO Y GENERANDO BOSQUETO")
        print(f"{'='*70}")
        
        if pais.lower() == 'venezuela':
            if not VENEZUELA_MODULE_AVAILABLE:
                return jsonify({
                    'success': False,
                    'error': 'Módulo de Venezuela no disponible'
                }), 500
            
            resultado_procesamiento, processor = procesar_venezuela(temp_reporte_pago, temp_reporte_absoluto)
            
            if not resultado_procesamiento:
                return jsonify({
                    'success': False,
                    'error': 'Error al procesar archivo de Venezuela',
                    'message': 'No se pudo generar el BOSQUETO'
                }), 500

            archivo_bosqueto = resultado_procesamiento.get('archivo_salida')
        
            if not archivo_bosqueto or not os.path.exists(archivo_bosqueto):
                return jsonify({
                    'success': False,
                    'error': 'BOSQUETO no fue generado correctamente'
                }), 500
            
            print(f"✅ BOSQUETO generado: {archivo_bosqueto}")
            print(f"   Filas procesadas: {resultado_procesamiento.get('filas_procesadas', 0)}")
            print(f"   Tasa utilizada: {resultado_procesamiento.get('tasa_utilizada', 0)} VES/USD")
            
        elif pais.lower() == 'colombia':
            if not COLOMBIA_MODULE_AVAILABLE:
                return jsonify({
                    'success': False,
                    'error': 'Módulo de Colombia no disponible'
                }), 500
            
            resultado_procesamiento, processor = procesar_colombia(temp_reporte_pago, temp_reporte_absoluto)
            
            if not resultado_procesamiento:
                return jsonify({
                    'success': False,
                    'error': 'Error al procesar archivo de Colombia',
                    'message': 'No se pudo generar el BOSQUETO'
                }), 500

            archivo_bosqueto = resultado_procesamiento.get('archivo_salida')
        
            if not archivo_bosqueto or not os.path.exists(archivo_bosqueto):
                return jsonify({
                    'success': False,
                    'error': 'BOSQUETO no fue generado correctamente'
                }), 500
            
            print(f"✅ BOSQUETO generado: {archivo_bosqueto}")
            print(f"   Filas procesadas: {resultado_procesamiento.get('filas_procesadas', 0)}")
            print(f"   Tasa utilizada: {resultado_procesamiento.get('tasa_utilizada', 0)} COP/USD")
            
        else:
            return jsonify({
                'success': False,
                'error': f'País "{pais}" no soportado actualmente',
                'message': 'Solo Venezuela y Colombia están disponibles'
            }), 400
        
        # PASO 2: LEER BOSQUETO Y APLICAR CÁLCULOS 
        print(f"\n{'='*70}")
        print(f"📖 PASO 2: LEYENDO BOSQUETO Y APLICANDO CÁLCULOS")
        print(f"{'='*70}")
        
        df_bosqueto_original = pd.read_excel(archivo_bosqueto, sheet_name='BOSQUETO')
        print(f"✅ BOSQUETO leído: {len(df_bosqueto_original)} filas, {len(df_bosqueto_original.columns)} columnas")

        # Limpiar NaN → 0 en columnas numéricas
        columnas_numericas = ['Monto CAPEX EXT', 'Monto CAPEX ORD', 'Monto CADM', 'Monto', 'Pago Independiente']
        for col in columnas_numericas:
            if col in df_bosqueto_original.columns:
                df_bosqueto_original[col] = df_bosqueto_original[col].fillna(0)

        # Aplicar cálculos
        df_bosqueto_original = processor.calcular_monto_usd(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_monto_capex(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_monto_opex(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_categoria(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_validacion(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_metodo_pago(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_tipo_capex(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_monto_ord(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_monto_ext(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_dia_pago(df_bosqueto_original)
        df_bosqueto_original['SEMANA'] = processor.obtener_semana_actual()
        df_bosqueto_original['MES DE PAGO'] = processor.obtener_mes_actual()
        
        # PASO 3: Guardar BOSQUETO procesado (sobrescribe el archivo)
        print(f"\n💾 PASO 3: Guardando BOSQUETO procesado...")
        
        with pd.ExcelWriter(archivo_bosqueto, engine='openpyxl', mode='w') as writer:
            df_bosqueto_original.to_excel(writer, sheet_name='BOSQUETO', index=False)
        
        print(f"✅ BOSQUETO procesado guardado: {len(df_bosqueto_original)} filas")
        
        # PASO 4: Subir a GCS (carpeta tmp)
        print(f"\n☁️ PASO 4: Subiendo a Google Cloud Storage (tmp/)...")
        url_descarga, nombre_archivo_gcs = subir_archivo_a_gcs_tmp(storage_client, archivo_bosqueto, pais)
        
        # Limpiar archivos temporales
        print(f"\n🧹 Limpiando archivos temporales...")
        archivos_temp = [temp_reporte_pago, temp_reporte_absoluto, archivo_bosqueto]
    
        for archivo in archivos_temp:
            try:
                if archivo and os.path.exists(archivo):
                    os.remove(archivo)
                    print(f"   ✅ Eliminado: {os.path.basename(archivo)}")
            except Exception as e:
                print(f"   ⚠️ No se pudo eliminar {archivo}: {e}")
       
        # Respuesta final
        respuesta = {
            'success': True,
            'pais': pais.upper(),
            'filas_procesadas': resultado_procesamiento.get('filas_procesadas', 0),
            'tasa_utilizada': resultado_procesamiento.get('tasa_utilizada', 0),
            'bosqueto_url': url_descarga,
            'file_name': nombre_archivo_gcs,
            'timestamp': datetime.now().isoformat(),
            'message': f"BOSQUETO procesado exitosamente: {resultado_procesamiento.get('filas_procesadas', 0)} filas"
        }
        
        print(f"\n✅ PROCESO BOSQUETO COMPLETADO")
        print(f"   País: {pais.upper()}")
        print(f"   URL de descarga: {url_descarga}")
        
        return jsonify(respuesta), 200
        
    except Exception as e:
        print(f"❌ Error en proceso BOSQUETO: {e}")
        traceback.print_exc()

        try:
            if 'temp_reporte_pago' in locals() and os.path.exists(temp_reporte_pago):
                os.remove(temp_reporte_pago)
            if 'temp_reporte_absoluto' in locals() and temp_reporte_absoluto and os.path.exists(temp_reporte_absoluto):
                os.remove(temp_reporte_absoluto)
            if 'archivo_bosqueto' in locals() and os.path.exists(archivo_bosqueto):
                os.remove(archivo_bosqueto)
        except:
            pass
        
        return jsonify({
            'success': False,
            'error': str(e),
            'message': f'Error procesando BOSQUETO: {str(e)}'
        }), 500


@app.route('/api/v1/upload-bosqueto', methods=['POST'])
def upload_bosqueto():
    """
    Endpoint principal: Upload BOSQUETO, verificar duplicados, cargar a BQ,
    generar DETALLE CORREGIDO, crear Excel y subir a GCS
    """
    try:
        # Validar archivo
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'error': 'No se proporcionó archivo'
            }), 400
        
        file = request.files['file']
        
        if file.filename == '' or not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'error': 'Archivo inválido'
            }), 400
        
        # NUEVO: Obtener país (opcional, default venezuela)
        pais = request.form.get('pais')
        
        print(f"📁 Archivo recibido: {file.filename}")
        print(f"🌎 País: {pais.upper()}")

        # Guardar archivo principal temporalmente
        temp_reporte_pago = f"/tmp/reporte_pago_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file.save(temp_reporte_pago)
        
        # Verificar si hay Reporte Absoluto (opcional)
        temp_reporte_absoluto = None
        if 'reporte_absoluto' in request.files:
            reporte_absoluto_file = request.files['reporte_absoluto']
            if reporte_absoluto_file.filename != '':
                temp_reporte_absoluto = f"/tmp/reporte_absoluto_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                reporte_absoluto_file.save(temp_reporte_absoluto)
                print(f"📋 Reporte Absoluto recibido: {reporte_absoluto_file.filename}")
    
            print(f"✅ Reporte Absoluto validado: {temp_reporte_absoluto}")
        else:
            print(f"ℹ️ No se proporcionó Reporte Absoluto")

        # PASO 1: PROCESAR Y GENERAR BOSQUETO 
        print(f"\n{'='*70}")
        print(f"🔄 PASO 1: PROCESANDO ARCHIVO Y GENERANDO BOSQUETO")
        print(f"{'='*70}")
        
        if pais.lower() == 'venezuela':
            if not VENEZUELA_MODULE_AVAILABLE:
                return jsonify({
                    'success': False,
                    'error': 'Módulo de Venezuela no disponible'
                }), 500
            
            # Llamar a procesar_venezuela para generar el BOSQUETO
            resultado_procesamiento, processor = procesar_venezuela(temp_reporte_pago, temp_reporte_absoluto)
            
            if not resultado_procesamiento:
                return jsonify({
                    'success': False,
                    'error': 'Error al procesar archivo de Venezuela',
                    'message': 'No se pudo generar el BOSQUETO'
                }), 500

            # Obtener ruta del BOSQUETO generado
            archivo_bosqueto = resultado_procesamiento.get('archivo_salida')
        
            if not archivo_bosqueto or not os.path.exists(archivo_bosqueto):
                return jsonify({
                    'success': False,
                    'error': 'BOSQUETO no fue generado correctamente'
                }), 500
            
            print(f"✅ BOSQUETO generado: {archivo_bosqueto}")
            print(f"   Filas procesadas: {resultado_procesamiento.get('filas_procesadas', 0)}")
            print(f"   Tasa utilizada: {resultado_procesamiento.get('tasa_utilizada', 0)} VES/USD")
            
        elif pais.lower() == 'colombia':
            if not COLOMBIA_MODULE_AVAILABLE:
                return jsonify({
                    'success': False,
                    'error': 'Módulo de Colombia no disponible'
                }), 500
            
            # Llamar a procesar_colombia para generar el BOSQUETO
            resultado_procesamiento, processor = procesar_colombia(temp_reporte_pago, temp_reporte_absoluto)
            
            if not resultado_procesamiento:
                return jsonify({
                    'success': False,
                    'error': 'Error al procesar archivo de Colombia',
                    'message': 'No se pudo generar el BOSQUETO'
                }), 500

            # Obtener ruta del BOSQUETO generado
            archivo_bosqueto = resultado_procesamiento.get('archivo_salida')
        
            if not archivo_bosqueto or not os.path.exists(archivo_bosqueto):
                return jsonify({
                    'success': False,
                    'error': 'BOSQUETO no fue generado correctamente'
                }), 500
            
            print(f"✅ BOSQUETO generado: {archivo_bosqueto}")
            print(f"   Filas procesadas: {resultado_procesamiento.get('filas_procesadas', 0)}")
            print(f"   Tasa utilizada: {resultado_procesamiento.get('tasa_utilizada', 0)} COP/USD")
            
        else:
            return jsonify({
                'success': False,
                'error': f'País "{pais}" no soportado actualmente',
                'message': 'Solo Venezuela y Colombia están disponibles'
            }), 400
        
        # PASO 2: LEER BOSQUETO GENERADO 
        print(f"\n{'='*70}")
        print(f"📖 PASO 2: LEYENDO BOSQUETO GENERADO")
        print(f"{'='*70}")
        
        df_bosqueto_original = pd.read_excel(archivo_bosqueto, sheet_name='BOSQUETO')
        # Imprimir DataFrame completo sin truncar
        # with pd.option_context('display.max_rows', None, 
        #                     'display.max_columns', None,
        #                     'display.width', None,
        #                     'display.max_colwidth', None):
        #         print(df_bosqueto_original)
        df_bosqueto_copia = pd.read_excel(archivo_bosqueto, sheet_name='BOSQUETO')

        print(f"✅ BOSQUETO leído: {len(df_bosqueto_original)} filas, {len(df_bosqueto_original.columns)} columnas")
        # print("🔍 Diagnóstico del DataFrame:")
        # print(f"   Total filas: {len(df_bosqueto_original)}")
        # print(f"   Total columnas: {len(df_bosqueto_original.columns)}")
        # print(f"\n📋 Nombres de columnas:")
        # for i, col in enumerate(df_bosqueto_original.columns):
        #     print(f"   {i}: {col}")

        # print(f"\n📊 Primeras 3 filas de columnas clave:")
        # print(df_bosqueto_original[['Monto', 'Moneda', 'Monto CAPEX EXT', 'Monto CAPEX ORD', 'Pago Independiente']].head(3))

        # print(f"\n🔧 Valores del processor:")
        # print(f"   tasa_dolar: {processor.tasa_dolar}")
        # print(f"   moneda: {processor.moneda}")

        # 2. Limpiar NaN → 0 en df_bosqueto_original
        columnas_numericas = ['Monto CAPEX EXT', 'Monto CAPEX ORD', 'Monto CADM', 'Monto', 'Pago Independiente']
        for col in columnas_numericas:
            if col in df_bosqueto_original.columns:
                df_bosqueto_original[col] = df_bosqueto_original[col].fillna(0)

        df_bosqueto_original = processor.calcular_monto_usd(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_monto_capex(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_monto_opex(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_categoria(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_validacion(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_metodo_pago(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_tipo_capex(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_monto_ord(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_monto_ext(df_bosqueto_original)
        df_bosqueto_original = processor.calcular_dia_pago(df_bosqueto_original)
        df_bosqueto_original['SEMANA'] = processor.obtener_semana_actual()
        df_bosqueto_original['MES DE PAGO'] = processor.obtener_mes_actual()

        # PASO 2: Mapear a BigQuery
        print(f"\n🔄 PASO 2: Mapeando columnas...")
        if pais.lower() == 'venezuela':
            df_mapped = mapear_columnas_bosqueto_a_bigquery_venezuela(df_bosqueto_original)
        elif pais.lower() == 'colombia':
            df_mapped = mapear_columnas_bosqueto_a_bigquery_colombia(df_bosqueto_original)
        
        # PASO 3: Crear clientes
        print(f"\n🔧 PASO 3: Creando clientes GCP...")
        bq_client = crear_cliente_bigquery()
        storage_client = crear_cliente_storage()
        
        # PASO 4: Cargar a BigQuery (con verificación de duplicados)
        print(f"\n📤 PASO 4: Cargando a BigQuery...")
        if pais.lower() == 'venezuela':
            resultado_carga = cargar_datos_a_bigquery_venezuela(bq_client, df_mapped)
        elif pais.lower() == 'colombia':
            resultado_carga = cargar_datos_a_bigquery_colombia(bq_client, df_mapped)
        
        if not resultado_carga['success']:
            if 'df_cargados' in resultado_carga:
                resultado_carga.pop('df_cargados')
            return jsonify(resultado_carga), 500
        
        df_cargados = resultado_carga['df_cargados']

        # PASO 5: Mapear registros cargados a formato Excel (DETALLE CORREGIDO)
        print(f"\n📋 PASO 5: Generando DETALLE CORREGIDO...")
        if pais.lower() == 'venezuela':
            df_bigquery = extraer_tabla_completa_por_lotes_venezuela(bq_client)
        elif pais.lower() == 'colombia':
            df_bigquery = extraer_tabla_completa_por_lotes_colombia(bq_client)
        
        if not df_bigquery.empty:
            if pais.lower() == 'venezuela':
                df_detalle_corregido = mapear_bigquery_a_excel_columns_venezuela(df_bigquery)
            elif pais.lower() == 'colombia':
                df_detalle_corregido = mapear_bigquery_a_excel_columns_colombia(df_bigquery)
            print(f"✅ DETALLE CORREGIDO: {len(df_detalle_corregido)} filas extraídas de BigQuery.")
        else:
            df_detalle_corregido = pd.DataFrame()
            print(f"⚠️ DETALLE CORREGIDO vacío (sin registros en BigQuery)")
        

        # PASO 6: Agregar hoja DETALLE CORREGIDO al BOSQUETO existente
        print(f"\n📝 PASO 6: Agregando hoja...")
        if pais.lower() == 'venezuela':
            agregar_hoja_detalle_venezuela(archivo_bosqueto, df_detalle_corregido)
        elif pais.lower() == 'colombia':
            agregar_hoja_detalle_colombia(archivo_bosqueto, df_detalle_corregido)

        print(f"\n📊 PASO 6.5: Creando hoja CAPEX PAGADO POR RECIBO...")
        if pais.lower() == 'venezuela':
            crear_hoja_capex_venezuela(archivo_bosqueto, df_detalle_corregido)
        elif pais.lower() == 'colombia':
            crear_hoja_capex_colombia(archivo_bosqueto, df_detalle_corregido)

        print(f"\n📊 PASO 6.6: Extrayendo datos de Responsables...")
        if pais.lower() == 'venezuela':
            df_responsables = extraer_responsables_capex_venezuela(bq_client)
        elif pais.lower() == 'colombia':
            df_responsables = extraer_responsables_capex_colombia(bq_client)

        if not df_responsables.empty:
            # PASO 6.7: Crear hoja Presupuesto Mensual
            print(f"\n💰 PASO 6.7: Creando hoja Presupuesto Mensual...")
            if pais.lower() == 'venezuela':
                crear_hoja_presupuesto_venezuela(archivo_bosqueto, df_responsables)
            elif pais.lower() == 'colombia':
                crear_hoja_presupuesto_colombia(archivo_bosqueto, df_responsables)
        else:
            print(f"⚠️ No se pudo crear Presupuesto Mensual (sin datos de responsables)")
        
        
        # PASO 6.8: Extraer diferencias de BigQuery
        print(f"\n📊 PASO 6.8: Extrayendo datos de diferencias...")
        if pais.lower() == 'venezuela':
            df_diferencia = extraer_diferencia_capex_venezuela(bq_client)
        elif pais.lower() == 'colombia':
            df_diferencia = extraer_diferencia_capex_colombia(bq_client)

        # PASO 6.9: Extraer tabla 2 de CAPEX PAGADO POR RECIBO
        print(f"\n📊 PASO 6.9: Extrayendo tabla 2 de CAPEX PAGADO POR RECIBO...")
        if pais.lower() == 'venezuela':
            df_ejecutado = extraer_tabla2_venezuela(archivo_bosqueto)
        elif pais.lower() == 'colombia':
            df_ejecutado = extraer_tabla2_colombia(archivo_bosqueto)

        # PASO 6.10: Crear tabla 2 en Presupuesto Mensual
        print(f"\n📊 PASO 6.10: Creando tabla 2 (Presupuesto vs Ejecutado)...")
        if pais.lower() == 'venezuela':
            df_tabla2 = crear_tabla2_venezuela(archivo_bosqueto, df_diferencia, df_ejecutado)
        elif pais.lower() == 'colombia':
            df_tabla2 = crear_tabla2_colombia(archivo_bosqueto, df_diferencia, df_ejecutado)

        # PASO 6.11: Cargar a BigQuery
        print(f"\n📤 PASO 6.11: Cargando diferencias a BigQuery...")
        if pais.lower() == 'venezuela':
            cargar_diferencia_a_bigquery_venezuela(bq_client, df_tabla2)
        elif pais.lower() == 'colombia':
            cargar_diferencia_a_bigquery_colombia(bq_client, df_tabla2)


        # PASO 6: Generar Excel con ambas hojas (USANDO ROUTER)

        # excel_path = generar_excel_consolidado(
        #     df_bosqueto=df_bosqueto_copia,
        #     df_detalle=df_detalle_corregido,
        #     pais=pais  
        # )
        
        # PASO 7: Subir a GCS
        print(f"\n☁️ PASO 7: Subiendo a Google Cloud Storage...")
        url_descarga, nombre_archivo_gcs = subir_archivo_a_gcs(storage_client, archivo_bosqueto)
        
        # Limpiar archivos temporales
        print(f"\n🧹 Limpiando archivos temporales...")
        archivos_temp = [temp_reporte_pago, temp_reporte_absoluto, archivo_bosqueto, ]
    
        for archivo in archivos_temp:
            try:
                if archivo and os.path.exists(archivo):
                    os.remove(archivo)
                    print(f"   ✅ Eliminado: {os.path.basename(archivo)}")
            except Exception as e:
                print(f"   ⚠️ No se pudo eliminar {archivo}: {e}")
       
        # Respuesta final
        respuesta = {
            'success': True,
            'pais': pais.upper(),  # NUEVO
            'total_rows': resultado_carga['total_rows'],
            'rows_duplicated': resultado_carga['rows_duplicated'],
            'rows_loaded': resultado_carga['rows_loaded'],
            'detalle_corregido_url': url_descarga,
            'file_name': nombre_archivo_gcs,
            'timestamp': datetime.now().isoformat(),
            'message': f"Proceso completado: {resultado_carga['rows_loaded']} registros cargados, {resultado_carga['rows_duplicated']} duplicados omitidos"
        }
        
        print(f"\n✅ PROCESO COMPLETADO EXITOSAMENTE")
        print(f"   País: {pais.upper()}")
        print(f"   URL de descarga: {url_descarga}")
        
        return jsonify(respuesta), 200
        
    except Exception as e:
        print(f"❌ Error en proceso: {e}")
        traceback.print_exc()

        try:
            if 'temp_reporte_pago' in locals() and os.path.exists(temp_reporte_pago):
                os.remove(temp_reporte_pago)
            if 'temp_reporte_absoluto' in locals() and temp_reporte_absoluto and os.path.exists(temp_reporte_absoluto):
                os.remove(temp_reporte_absoluto)
            if 'archivo_bosqueto' in locals() and os.path.exists(archivo_bosqueto):
                os.remove(archivo_bosqueto)
            #if 'excel_path' in locals() and os.path.exists(excel_path):
               # os.remove(excel_path)
        except:
            pass
        
        return jsonify({
            'success': False,
            'error': str(e),
            'message': f'Error procesando solicitud: {str(e)}'
        }), 500

@app.route('/api/v1/table-info', methods=['GET'])
def table_info():
    """Endpoint para obtener información de la tabla BigQuery"""
    try:
        client = crear_cliente_bigquery()
        table_id = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET_COP}.{BIGQUERY_TABLE_COP}"
        
        table = client.get_table(table_id)
        
        return jsonify({
            'success': True,
            'project': GCP_PROJECT_ID,
            'dataset': BIGQUERY_DATASET_COP,
            'table': BIGQUERY_TABLE_COP,
            'num_rows': table.num_rows,
            'num_columns': len(table.schema),
            'created': table.created.isoformat() if table.created else None,
            'modified': table.modified.isoformat() if table.modified else None,
            'size_bytes': table.num_bytes,
            'size_mb': round(table.num_bytes / (1024 * 1024), 2)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# =================== ENDPOINT DE PRUEBA DE CONEXIÓN ===================

@app.route('/api/v1/test-connection', methods=['GET'])
def test_connection():
    """
    Endpoint para probar la conexión a BigQuery y listar recursos disponibles
    Retorna información sobre datasets y tablas accesibles
    """
    try:
        print("🔍 Probando conexión a BigQuery...")
        
        # Crear cliente
        client = crear_cliente_bigquery()
        
        # Información del proyecto
        project_info = {
            'project_id': client.project,
            'location': 'US'  # o tu región
        }
        
        # Listar datasets
        print(f"📊 Listando datasets en {client.project}...")
        datasets = list(client.list_datasets())
        
        datasets_info = []
        total_tablas = 0
        
        for dataset in datasets:
            dataset_id = dataset.dataset_id
            dataset_ref = f"{client.project}.{dataset_id}"
            
            print(f"   Dataset encontrado: {dataset_id}")
            
            # Obtener info detallada del dataset
            try:
                dataset_obj = client.get_dataset(dataset_ref)
                
                # Listar tablas en este dataset
                tablas = list(client.list_tables(dataset_ref))
                tablas_info = []
                
                for tabla in tablas:
                    tabla_id = tabla.table_id
                    tabla_ref = f"{client.project}.{dataset_id}.{tabla_id}"
                    
                    try:
                        # Obtener info de la tabla
                        tabla_obj = client.get_table(tabla_ref)
                        
                        tablas_info.append({
                            'table_id': tabla_id,
                            'table_type': tabla.table_type,
                            'full_table_id': tabla_ref,
                            'num_rows': tabla_obj.num_rows,
                            'num_columns': len(tabla_obj.schema),
                            'size_mb': round(tabla_obj.num_bytes / (1024 * 1024), 2) if tabla_obj.num_bytes else 0,
                            'created': tabla_obj.created.isoformat() if tabla_obj.created else None,
                            'modified': tabla_obj.modified.isoformat() if tabla_obj.modified else None
                        })
                        
                        total_tablas += 1
                        
                    except Exception as tabla_error:
                        print(f"      ⚠️ Error obteniendo info de tabla {tabla_id}: {tabla_error}")
                        tablas_info.append({
                            'table_id': tabla_id,
                            'error': str(tabla_error)
                        })
                
                datasets_info.append({
                    'dataset_id': dataset_id,
                    'full_dataset_id': dataset_ref,
                    'location': dataset_obj.location,
                    'created': dataset_obj.created.isoformat() if dataset_obj.created else None,
                    'modified': dataset_obj.modified.isoformat() if dataset_obj.modified else None,
                    'num_tables': len(tablas_info),
                    'tables': tablas_info
                })
                
            except Exception as dataset_error:
                print(f"   ⚠️ Error obteniendo dataset {dataset_id}: {dataset_error}")
                datasets_info.append({
                    'dataset_id': dataset_id,
                    'error': str(dataset_error)
                })
        
        # Verificar si existe nuestra tabla específica
        target_table_id = f"{GCP_PROJECT_ID}.{BIGQUERY_DATASET}.{BIGQUERY_TABLE}"
        target_exists = False
        target_info = None
        
        try:
            target_table = client.get_table(target_table_id)
            target_exists = True
            target_info = {
                'exists': True,
                'full_table_id': target_table_id,
                'num_rows': target_table.num_rows,
                'num_columns': len(target_table.schema),
                'size_mb': round(target_table.num_bytes / (1024 * 1024), 2) if target_table.num_bytes else 0,
                'created': target_table.created.isoformat() if target_table.created else None,
                'modified': target_table.modified.isoformat() if target_table.modified else None,
                'schema_fields': [field.name for field in target_table.schema]
            }
            print(f"✅ Tabla objetivo encontrada: {target_table_id}")
        except Exception as e:
            target_info = {
                'exists': False,
                'full_table_id': target_table_id,
                'error': str(e),
                'message': 'La tabla no existe. Se creará automáticamente al subir el primer BOSQUETO.'
            }
            print(f"⚠️ Tabla objetivo no encontrada: {target_table_id}")
        
        # Respuesta completa
        response = {
            'success': True,
            'connection_status': 'connected',
            'timestamp': datetime.now().isoformat(),
            'project': project_info,
            'config': {
                'gcp_project_id': GCP_PROJECT_ID,
                'dataset': BIGQUERY_DATASET,
                'table': BIGQUERY_TABLE,
                'credentials_file': CREDENTIALS_FILE,
                'credentials_exists': os.path.exists(CREDENTIALS_FILE)
            },
            'summary': {
                'total_datasets': len(datasets_info),
                'total_tables': total_tablas,
                'target_table_exists': target_exists
            },
            'datasets': datasets_info,
            'target_table': target_info
        }
        
        print("✅ Test de conexión completado exitosamente")
        
        return jsonify(response), 200
        
    except Exception as e:
        print(f"❌ Error en test de conexión: {e}")
        traceback.print_exc()
        
        return jsonify({
            'success': False,
            'connection_status': 'failed',
            'error': str(e),
            'message': 'No se pudo conectar a BigQuery. Verifica las credenciales y permisos.',
            'config': {
                'gcp_project_id': GCP_PROJECT_ID,
                'dataset': BIGQUERY_DATASET,
                'table': BIGQUERY_TABLE,
                'credentials_file': CREDENTIALS_FILE,
                'credentials_exists': os.path.exists(CREDENTIALS_FILE)
            }
        }), 500

# =================== ENDPOINT DE PRUEBA GCS ===================

@app.route('/api/v1/test-gcs', methods=['GET'])
def test_gcs():
    """
    Test de conexión a Google Cloud Storage
    """
    try:
        storage_client = crear_cliente_storage()
        
        # Verificar bucket
        try:
            bucket = storage_client.bucket(GCS_BUCKET_NAME)
            bucket.reload()
            bucket_exists = True
        except:
            bucket_exists = False
        
        # Listar archivos
        archivos = []
        if bucket_exists:
            archivos = listar_archivos_bucket(storage_client, GCS_BUCKET_NAME, limit=5)
        
        return jsonify({
            'success': True,
            'connection_status': 'connected',
            'bucket': {
                'name': GCS_BUCKET_NAME,
                'exists': bucket_exists,
                'url': f"https://console.cloud.google.com/storage/browser/{GCS_BUCKET_NAME}"
            },
            'recent_files': archivos,
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/v1/bucket-info', methods=['GET'])
def bucket_info():
    """
    Endpoint para obtener información del bucket GCS
    """
    try:
        storage_client = crear_cliente_storage()
        
        # Obtener info del bucket
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        bucket.reload()
        
        # Listar últimos archivos
        archivos = listar_archivos_bucket(storage_client, GCS_BUCKET_NAME, limit=10)
        
        return jsonify({
            'success': True,
            'bucket': {
                'name': GCS_BUCKET_NAME,
                'location': bucket.location,
                'storage_class': bucket.storage_class,
                'created': bucket.time_created.isoformat() if bucket.time_created else None,
            },
            'recent_files': archivos,
            'total_files_shown': len(archivos),
            'timestamp': datetime.now().isoformat()
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'message': f'Error obteniendo info del bucket: {str(e)}'
        }), 500


# =================== ENDPOINT LISTAR LOGS ===================

@app.route('/api/v1/logs', methods=['GET'])
def listar_logs():
    """
    Listar archivos en la carpeta logs/ del bucket, agrupados por fecha.
    Retorna links de descarga para cada archivo.
    """
    try:
        print(f"📋 Listando archivos de logs...", flush=True)
        
        storage_client = crear_cliente_storage()
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        
        # Listar todos los blobs en la carpeta logs/
        blobs = bucket.list_blobs(prefix='logs/')
        
        # Agrupar por fecha
        logs_por_fecha = {}
        total_archivos = 0
        
        for blob in blobs:
            # Ignorar la carpeta misma (si existe como objeto)
            if blob.name == 'logs/' or blob.name.endswith('/'):
                continue
            
            # Extraer fecha del path: logs/2026-01-26/archivo.xlsx
            partes = blob.name.split('/')
            if len(partes) >= 3:
                fecha = partes[1]  # 2026-01-26
                nombre_archivo = partes[2]  # Consolidado_VENEZUELA_2026-01-26_10-30-00.xlsx
            else:
                fecha = 'sin_fecha'
                nombre_archivo = blob.name
            
            # Construir URL pública
            url_publica = f"https://storage.googleapis.com/{GCS_BUCKET_NAME}/{blob.name}"
            
            # Crear entrada del archivo
            archivo_info = {
                'nombre': nombre_archivo,
                'path': blob.name,
                'url': url_publica,
                'tamaño_mb': round(blob.size / (1024 * 1024), 2) if blob.size else 0,
                'creado': blob.time_created.isoformat() if blob.time_created else None,
                'actualizado': blob.updated.isoformat() if blob.updated else None
            }
            
            # Agrupar por fecha
            if fecha not in logs_por_fecha:
                logs_por_fecha[fecha] = []
            logs_por_fecha[fecha].append(archivo_info)
            total_archivos += 1
        
        # Ordenar fechas de más reciente a más antigua
        fechas_ordenadas = sorted(logs_por_fecha.keys(), reverse=True)
        
        # Construir respuesta ordenada
        logs_ordenados = []
        for fecha in fechas_ordenadas:
            logs_ordenados.append({
                'fecha': fecha,
                'archivos': logs_por_fecha[fecha],
                'total_archivos': len(logs_por_fecha[fecha])
            })
        
        print(f"✅ Logs listados: {total_archivos} archivos en {len(logs_por_fecha)} fechas", flush=True)
        
        return jsonify({
            'success': True,
            'total_archivos': total_archivos,
            'total_fechas': len(logs_por_fecha),
            'logs': logs_ordenados,
            'bucket': GCS_BUCKET_NAME,
            'timestamp': datetime.now().isoformat()
        }), 200
        
    except Exception as e:
        print(f"❌ Error listando logs: {e}", flush=True)
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e),
            'message': f'Error listando logs: {str(e)}'
        }), 500


# =================== MAIN ===================

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    debug = os.getenv('DEBUG', 'False').lower() == 'true'
    
    print("=" * 60)
    print("🚀 CAPEX BigQuery API v1.1")
    print("=" * 60)
    print(f"📊 Proyecto GCP: {GCP_PROJECT_ID}")
    print(f"📊 Dataset: {BIGQUERY_DATASET}")
    print(f"📊 Tabla: {BIGQUERY_TABLE}")
    print(f"🔧 Puerto: {port}")
    print(f"🔧 Debug: {debug}")
    print("=" * 60)
    
    app.run(host='0.0.0.0', port=port, debug=debug)
