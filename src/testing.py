"""
Testing de cierre de mes - Simulación Semana 1 de Febrero 2026
Este script simula el proceso de actualización de plantilla cuando cambia el mes.
"""

from openpyxl import load_workbook
from datetime import datetime
from google.cloud import storage
from google.oauth2 import service_account
import os

# Configuración
GCS_BUCKET_NAME = os.getenv('GCS_BUCKET_NAME', 'your-bucket-name')
CREDENTIALS_FILE = os.getenv('GOOGLE_APPLICATION_CREDENTIALS', 'credentials.json')

# Meses en español
MESES_ES = {
    1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril',
    5: 'Mayo', 6: 'Junio', 7: 'Julio', 8: 'Agosto',
    9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
}


def es_semana_1_del_mes(fecha):
    """Verificar si la fecha está en la semana 1 del mes (días 1-7)"""
    return 1 <= fecha.day <= 7


def obtener_mes_anterior(mes_actual, año_actual):
    """Obtener el mes y año anterior"""
    if mes_actual == 1:
        return 12, año_actual - 1
    else:
        return mes_actual - 1, año_actual


def descargar_plantilla_local(ruta_local):
    """Descargar plantilla desde GCS a local (para testing)"""
    try:
        if CREDENTIALS_FILE and os.path.exists(CREDENTIALS_FILE):
            credentials = service_account.Credentials.from_service_account_file(
                CREDENTIALS_FILE,
                scopes=["https://www.googleapis.com/auth/cloud-platform"]
            )
            client = storage.Client(credentials=credentials)
        else:
            client = storage.Client()
        
        bucket = client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob('template/vzla/consolidado_capex_ve_2025_2026_template.xlsx')
        blob.download_to_filename(ruta_local)
        print(f"✅ Plantilla descargada: {ruta_local}")
        return True
    except Exception as e:
        print(f"❌ Error descargando plantilla: {e}")
        return False


def actualizar_titulos_cierre_mes(wb, mes_actual, año_actual):
    """
    Actualizar títulos cuando es semana 1 del nuevo mes.
    
    Args:
        wb: Workbook de openpyxl
        mes_actual: Número del mes actual (1-12)
        año_actual: Año actual (ej: 2026)
    """
    mes_anterior, año_anterior = obtener_mes_anterior(mes_actual, año_actual)
    
    nombre_mes_actual = MESES_ES[mes_actual]
    nombre_mes_anterior = MESES_ES[mes_anterior]
    
    print(f"\n📅 Cierre de mes detectado:")
    print(f"   Mes anterior: {nombre_mes_anterior}-{año_anterior}")
    print(f"   Mes actual: {nombre_mes_actual}-{año_actual}")
    
    # ============================================
    # 1. HOJA "Graficos" - Actualizar títulos
    # ============================================
    if 'Graficos' in wb.sheetnames:
        ws_graficos = wb['Graficos']
        
        # G6: PPTO Mes-Año
        titulo_g6 = f"PPTO {nombre_mes_actual}-{año_actual}"
        ws_graficos['G6'] = titulo_g6
        print(f"   ✅ Graficos G6: '{titulo_g6}'")
        
        # H6: Pagado Mes-Año
        titulo_h6 = f"Pagado {nombre_mes_actual}-{año_actual}"
        ws_graficos['H6'] = titulo_h6
        print(f"   ✅ Graficos H6: '{titulo_h6}'")
        
        # I6: DISPONIBLE Mes-Año
        titulo_i6 = f"DISPONIBLE {nombre_mes_actual}-{año_actual}"
        ws_graficos['I6'] = titulo_i6
        print(f"   ✅ Graficos I6: '{titulo_i6}'")
    else:
        print(f"   ⚠️ Hoja 'Graficos' no encontrada")
    
    # ============================================
    # 2. HOJA "Presupuesto Mensual" - Actualizar títulos
    # ============================================
    if 'Presupuesto Mensual' in wb.sheetnames:
        ws_presupuesto = wb['Presupuesto Mensual']
        
        # C18: Remanente Mes pasado-Año
        titulo_c18 = f"Remanente {nombre_mes_anterior}-{año_anterior}"
        ws_presupuesto['C18'] = titulo_c18
        print(f"   ✅ Presupuesto C18: '{titulo_c18}'")
        
        # D18: Presupuesto Mes actual-Año
        titulo_d18 = f"Presupuesto {nombre_mes_actual}-{año_actual}"
        ws_presupuesto['D18'] = titulo_d18
        print(f"   ✅ Presupuesto D18: '{titulo_d18}'")
        
        # E18: Ejecutado Mes actual-Año
        titulo_e18 = f"Ejecutado {nombre_mes_actual}-{año_actual}"
        ws_presupuesto['E18'] = titulo_e18
        print(f"   ✅ Presupuesto E18: '{titulo_e18}'")
    else:
        print(f"   ⚠️ Hoja 'Presupuesto Mensual' no encontrada")


def traspasar_diferencia_a_remanente(wb):
    """
    Traspasar los valores de Diferencia a Remanente.
    Filas: 20, 22-32 (saltando 21)
    
    Fórmula Diferencia: =E-D+C
    El valor calculado se copia a C (Remanente)
    """
    if 'Presupuesto Mensual' not in wb.sheetnames:
        print(f"   ⚠️ Hoja 'Presupuesto Mensual' no encontrada para cierre")
        return
    
    ws = wb['Presupuesto Mensual']
    
    # Filas a procesar (20, 22-32, saltando 21)
    filas = [20] + list(range(22, 33))
    
    print(f"\n💰 Traspasando Diferencia → Remanente:")
    
    for fila in filas:
        # Leer valores actuales
        val_c = ws[f'C{fila}'].value  # Remanente actual
        val_d = ws[f'D{fila}'].value  # Presupuesto
        val_e = ws[f'E{fila}'].value  # Ejecutado
        
        # Convertir a número (manejar None y strings)
        try:
            c = float(val_c) if val_c is not None else 0
        except (ValueError, TypeError):
            c = 0
            
        try:
            d = float(val_d) if val_d is not None else 0
        except (ValueError, TypeError):
            d = 0
            
        try:
            e = float(val_e) if val_e is not None else 0
        except (ValueError, TypeError):
            e = 0
        
        # Calcular Diferencia: E - D + C
        diferencia = e - d + c
        
        # Escribir en Remanente (C)
        ws[f'C{fila}'] = diferencia
        
        print(f"   Fila {fila}: C={c:.2f}, D={d:.2f}, E={e:.2f} → Diferencia={diferencia:.2f} → Nuevo Remanente")


def main():
    print("=" * 60)
    print("🧪 TESTING: Cierre de Mes - Simulación Semana 1 Febrero 2026")
    print("=" * 60)
    
    # Simular fecha: 3 de Febrero 2026 (Semana 1)
    fecha_simulada = datetime(2026, 2, 3)
    mes_actual = fecha_simulada.month
    año_actual = fecha_simulada.year
    
    print(f"\n📆 Fecha simulada: {fecha_simulada.strftime('%d/%m/%Y')}")
    print(f"   Mes: {MESES_ES[mes_actual]}")
    print(f"   Año: {año_actual}")
    print(f"   ¿Es semana 1?: {es_semana_1_del_mes(fecha_simulada)}")
    
    # Verificar si es semana 1
    if not es_semana_1_del_mes(fecha_simulada):
        print("\n⚠️ No es semana 1, no se ejecuta cierre de mes")
        return
    
    # Ruta de la plantilla (local para testing)
    plantilla_local = "plantilla_test.xlsx"
    
    # Intentar descargar la plantilla
    if not os.path.exists(plantilla_local):
        print(f"\n📥 Descargando plantilla desde GCS...")
        if not descargar_plantilla_local(plantilla_local):
            print("❌ No se pudo descargar la plantilla. Verifica las credenciales.")
            return
    else:
        print(f"\n📄 Usando plantilla local existente: {plantilla_local}")
    
    # Cargar plantilla
    print(f"\n📂 Cargando plantilla...")
    try:
        wb = load_workbook(plantilla_local)
        print(f"   ✅ Plantilla cargada. Hojas: {wb.sheetnames}")
    except Exception as e:
        print(f"   ❌ Error cargando plantilla: {e}")
        return
    
    # 1. Actualizar títulos
    actualizar_titulos_cierre_mes(wb, mes_actual, año_actual)
    
    # 2. Traspasar Diferencia → Remanente
    traspasar_diferencia_a_remanente(wb)
    
    # Guardar resultado
    archivo_salida = "plantilla_test_resultado.xlsx"
    print(f"\n💾 Guardando resultado: {archivo_salida}")
    wb.save(archivo_salida)
    print(f"✅ Archivo guardado exitosamente")
    
    print("\n" + "=" * 60)
    print("✅ TEST COMPLETADO")
    print(f"   Revisa el archivo: {archivo_salida}")
    print("=" * 60)


if __name__ == "__main__":
    main()
