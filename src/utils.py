# =================== ARCHIVO: utils.py (CÓDIGO COMPLETO CORREGIDO) ===================
"""
Utilidades comunes para el proyecto Consolidado CAPEX
"""

import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from pathlib import Path
import json
import datetime
import os
from typing import Optional, Dict, Any
import numpy as np

# Cargar variables de entorno desde .env
try:
    from dotenv import load_dotenv
    load_dotenv()
    print("✅ Variables de entorno cargadas desde .env")
except ImportError:
    print("⚠️ python-dotenv no instalado, usando variables de entorno del sistema")

# Importar módulo de tasas BCV desde BigQuery
try:
    from tasas import obtener_helper_tasas, obtener_tasa_bcv, precargar_tasas_bcv
    print("✅ Módulo de tasas BCV importado correctamente")
except ImportError:
    print("⚠️ No se pudo importar módulo de tasas, usando fallback")
    obtener_helper_tasas = None
    obtener_tasa_bcv = None
    precargar_tasas_bcv = None

class APIHelper:
    """Helper para consultar APIs de tasas de cambio"""
    
    def __init__(self, timeout=10):
        self.timeout = timeout
        self.tasas_ftd_cache = None  # Cache para tasas FTD
    
    def obtener_tasas_ftd(self):
        """
        Obtener todas las tasas de Farmatodo desde el endpoint TC_FTD_ENDPOINT.
        Guarda en cache para no consultar múltiples veces.
        
        Returns:
            dict: Diccionario con fecha_vigencia como clave y datos de tasa como valor
        """
        if self.tasas_ftd_cache is not None:
            return self.tasas_ftd_cache
        
        try:
            # Obtener endpoint desde variables de entorno
            endpoint = os.environ.get('TC_FTD_ENDPOINT', 'https://consulta-tasas-ftd-632121084032.europe-west1.run.app/')
            
            print(f"💱 Consultando tasas FTD desde: {endpoint}")
            response = requests.get(endpoint, timeout=self.timeout)
            
            if response.status_code == 200:
                data = response.json()
                
                if 'datos' in data:
                    # Crear diccionario indexado por fecha para búsqueda rápida
                    tasas_dict = {}
                    for item in data['datos']:
                        fecha = item.get('fecha_vigencia', '')
                        tasas_dict[fecha] = {
                            'tasa_bcv': item.get('tasa_bcv', 0),
                            'tasa_farmatodo': item.get('tasa_farmatodo', 0),
                            'tasa_referencial': item.get('tasa_referencial', 0)
                        }
                    
                    self.tasas_ftd_cache = tasas_dict
                    print(f"✅ Tasas FTD cargadas: {len(tasas_dict)} fechas disponibles")
                    
                    # Mostrar algunas muestras
                    if tasas_dict:
                        fechas_sample = list(tasas_dict.keys())[:3]
                        print(f"   💡 Fechas ejemplo: {fechas_sample}")
                    
                    return tasas_dict
                else:
                    print(f"⚠️ Respuesta sin campo 'datos': {list(data.keys())}")
            else:
                print(f"⚠️ Error HTTP {response.status_code} consultando tasas FTD")
                
        except Exception as e:
            print(f"❌ Error consultando tasas FTD: {e}")
        
        # Retornar diccionario vacío si falla
        self.tasas_ftd_cache = {}
        return {}
    
    def _normalizar_fecha_str(self, fecha):
        """
        Normalizar fecha a formato string YYYY-MM-DD
        """
        if isinstance(fecha, pd.Timestamp):
            return fecha.strftime('%Y-%m-%d')
        elif isinstance(fecha, datetime.datetime):
            return fecha.strftime('%Y-%m-%d')
        elif isinstance(fecha, datetime.date):
            return fecha.strftime('%Y-%m-%d')
        elif isinstance(fecha, str):
            # Intentar normalizar formato de fecha
            try:
                # Si viene en formato DD/MM/YYYY
                if '/' in fecha:
                    partes = fecha.split('/')
                    if len(partes) == 3:
                        if len(partes[2]) == 4:  # DD/MM/YYYY
                            return f"{partes[2]}-{partes[1].zfill(2)}-{partes[0].zfill(2)}"
                        else:  # MM/DD/YYYY o similar
                            return fecha
                else:
                    return fecha
            except:
                return str(fecha)
        else:
            return str(fecha) if fecha else ""
    
    def obtener_tasa_ftd_para_fecha(self, fecha):
        """
        Obtener la tasa Farmatodo para una fecha específica.
        
        Args:
            fecha: Fecha en formato string 'YYYY-MM-DD', datetime.date, o pd.Timestamp
        
        Returns:
            float: Tasa Farmatodo o 0 si no se encuentra
        """
        tasas = self.obtener_tasas_ftd()
        
        if not tasas:
            return 0
        
        fecha_str = self._normalizar_fecha_str(fecha)
        
        # Buscar tasa para la fecha
        if fecha_str in tasas:
            return tasas[fecha_str].get('tasa_farmatodo', 0)
        
        # Si no encuentra la fecha exacta, retornar 0
        return 0
    
    def obtener_tasa_bcv_para_fecha(self, fecha):
        """
        Obtener la tasa BCV para una fecha específica desde BigQuery.
        Usa la tabla cxp_vzla.bcv_tasas.
        
        Args:
            fecha: Fecha en formato string 'YYYY-MM-DD', datetime.date, o pd.Timestamp
        
        Returns:
            float: Tasa BCV o 0 si no se encuentra
        """
        # Usar el módulo de tasas de BigQuery si está disponible
        if obtener_tasa_bcv is not None:
            return obtener_tasa_bcv(fecha)
        
        # Fallback: retornar 0 si el módulo no está disponible
        return 0
    
    def obtener_fecha_viernes_anterior(self):
        """Obtener la fecha del viernes de la semana pasada"""
        hoy = datetime.date.today()
        dias_desde_lunes = hoy.weekday()  # 0 = Lunes, 6 = Domingo
        
        # Si es Lunes-Viernes: viernes de semana pasada
        # Si es Sábado-Domingo: viernes de esta semana
        if dias_desde_lunes <= 4:  # Lunes a Viernes
            dias_atras = dias_desde_lunes + 3
        else:  # Sábado o Domingo  
            dias_atras = dias_desde_lunes - 4
            
        fecha_viernes = hoy - datetime.timedelta(days=dias_atras)
        return fecha_viernes
        
    def obtener_tasa_venezuela_fecha_historica(self, fecha):
        """Obtener tasa histórica usando la nueva API BCV que SÍ tiene histórico"""
        try:
            fecha_str = fecha.strftime('%Y-%m-%d')
            print(f"🇻🇪 Consultando tasa BCV histórica para {fecha_str}...")
            
            # NUEVA API: https://bcv-api.rafnixg.dev/rates/YYYY-MM-DD
            url = f"https://bcv-api.rafnixg.dev/rates/{fecha_str}"
            response = requests.get(url, timeout=self.timeout)
            
            if response.status_code == 200:
                data = response.json()
                if 'dollar' in data:
                    tasa = float(data['dollar'])
                    fecha_confirmada = data.get('date', fecha_str)
                    print(f"✅ Tasa BCV histórica {fecha_confirmada}: {tasa:.4f} VES/USD")
                    return tasa, fecha
            else:
                print(f"⚠️ HTTP {response.status_code} para fecha {fecha_str}")
                    
        except Exception as e:
            print(f"⚠️ Error consultando tasa histórica para {fecha}: {e}")
        
        return None, fecha
    
    def obtener_tasa_venezuela_actual(self):
        """Obtener tasa actual usando la nueva API BCV"""
        try:
            print("🇻🇪 Consultando tasa BCV actual...")
            
            # NUEVA API: https://bcv-api.rafnixg.dev/rates/
            url = "https://bcv-api.rafnixg.dev/rates/"
            response = requests.get(url, timeout=self.timeout)
            
            if response.status_code == 200:
                data = response.json()
                if 'dollar' in data:
                    tasa = float(data['dollar'])
                    fecha_str = data.get('date', datetime.date.today().strftime('%Y-%m-%d'))
                    fecha_obj = datetime.datetime.strptime(fecha_str, '%Y-%m-%d').date()
                    print(f"✅ Tasa BCV actual: {tasa:.4f} VES/USD (fecha: {fecha_str})")
                    return tasa, fecha_obj
                    
        except Exception as e:
            print(f"⚠️ Error consultando tasa actual: {e}")
        
        return None, None
    
    def obtener_tasa_venezuela(self):
        """Obtener tasa del viernes de la semana pasada usando API con histórico REAL"""
        print("📅 OBTENIENDO TASA DEL VIERNES ANTERIOR CON API HISTÓRICA...")
        print("-" * 60)
        
        # 1. Calcular fecha del viernes anterior
        fecha_viernes = self.obtener_fecha_viernes_anterior()
        hoy = datetime.date.today()
        
        print(f"📅 Fecha actual: {hoy.strftime('%A, %Y-%m-%d')}")
        print(f"📅 Viernes objetivo: {fecha_viernes.strftime('%A, %Y-%m-%d')}")
        print(f"📅 Días atrás: {(hoy - fecha_viernes).days}")
        
        # 2. Intentar obtener tasa del viernes (AHORA SÍ DEBERÍA FUNCIONAR)
        tasa, fecha_usada = self.obtener_tasa_venezuela_fecha_historica(fecha_viernes)
        if tasa:
            return tasa, fecha_usada
        
        # 3. Respaldo: jueves anterior
        print("⚠️ No hay datos del viernes, intentando jueves...")
        fecha_jueves = fecha_viernes - datetime.timedelta(days=1)
        tasa, fecha_usada = self.obtener_tasa_venezuela_fecha_historica(fecha_jueves)
        if tasa:
            return tasa, fecha_usada
        
        # 4. Respaldo: miércoles anterior  
        print("⚠️ No hay datos del jueves, intentando miércoles...")
        fecha_miercoles = fecha_viernes - datetime.timedelta(days=2)
        tasa, fecha_usada = self.obtener_tasa_venezuela_fecha_historica(fecha_miercoles)
        if tasa:
            return tasa, fecha_usada
        
        # 5. Respaldo: martes anterior
        print("⚠️ No hay datos del miércoles, intentando martes...")
        fecha_martes = fecha_viernes - datetime.timedelta(days=3)
        tasa, fecha_usada = self.obtener_tasa_venezuela_fecha_historica(fecha_martes)
        if tasa:
            return tasa, fecha_usada
        
        # 6. Respaldo: lunes anterior
        print("⚠️ No hay datos del martes, intentando lunes...")
        fecha_lunes = fecha_viernes - datetime.timedelta(days=4)
        tasa, fecha_usada = self.obtener_tasa_venezuela_fecha_historica(fecha_lunes)
        if tasa:
            return tasa, fecha_usada
        
        # 7. Último recurso: tasa actual
        print("⚠️ Usando tasa actual como último recurso...")
        tasa, fecha_usada = self.obtener_tasa_venezuela_actual()
        if tasa:
            return tasa, fecha_usada
        
        # 8. Tasa de respaldo fija (última opción)
        tasa_respaldo = 169.98
        fecha_respaldo = hoy
        print(f"📊 Usando tasa de respaldo fija: {tasa_respaldo} VES/USD")
        return tasa_respaldo, fecha_respaldo
    
    def obtener_tasa_colombia_fecha_historica(self, fecha):
        """Obtener tasa histórica de Colombia usando la API de TRM"""
        try:
            fecha_str = fecha.strftime('%Y-%m-%d')
            print(f"🇨🇴 Consultando tasa TRM histórica para {fecha_str}...")
            
            # API: https://trm-colombia.vercel.app/?date=YYYY-MM-DD
            url = f"https://trm-colombia.vercel.app/?date={fecha_str}"
            response = requests.get(url, timeout=self.timeout)
            
            if response.status_code == 200:
                data = response.json()
                if 'data' in data and 'value' in data['data']:
                    tasa = float(data['data']['value'])
                    fecha_confirmada = data['data'].get('validityFrom', fecha_str)
                    # Extraer solo la fecha de la cadena ISO (YYYY-MM-DD)
                    if 'T' in fecha_confirmada:
                        fecha_confirmada = fecha_confirmada.split('T')[0]
                    print(f"✅ Tasa TRM histórica {fecha_confirmada}: {tasa:.2f} COP/USD")
                    fecha_obj = datetime.datetime.strptime(fecha_confirmada, '%Y-%m-%d').date()
                    return tasa, fecha_obj
            else:
                print(f"⚠️ HTTP {response.status_code} para fecha {fecha_str}")
                    
        except Exception as e:
            print(f"⚠️ Error consultando tasa histórica para {fecha}: {e}")
        
        return None, fecha
    
    def obtener_tasa_colombia_actual(self):
        """Obtener tasa actual de Colombia usando la API de TRM"""
        try:
            print("🇨🇴 Consultando tasa TRM actual...")
            hoy = datetime.date.today()
            fecha_str = hoy.strftime('%Y-%m-%d')
            
            # API: https://trm-colombia.vercel.app/?date=YYYY-MM-DD
            url = f"https://trm-colombia.vercel.app/?date={fecha_str}"
            response = requests.get(url, timeout=self.timeout)
            
            if response.status_code == 200:
                data = response.json()
                if 'data' in data and 'value' in data['data']:
                    tasa = float(data['data']['value'])
                    fecha_confirmada = data['data'].get('validityFrom', fecha_str)
                    # Extraer solo la fecha de la cadena ISO (YYYY-MM-DD)
                    if 'T' in fecha_confirmada:
                        fecha_confirmada = fecha_confirmada.split('T')[0]
                    fecha_obj = datetime.datetime.strptime(fecha_confirmada, '%Y-%m-%d').date()
                    print(f"✅ Tasa TRM actual: {tasa:.2f} COP/USD (fecha: {fecha_confirmada})")
                    return tasa, fecha_obj
                    
        except Exception as e:
            print(f"⚠️ Error consultando tasa actual: {e}")
        
        return None, None
    
    def obtener_tasa_colombia(self):
        """Obtener tasa del viernes de la semana pasada usando API con histórico"""
        print("📅 OBTENIENDO TASA DEL VIERNES ANTERIOR CON API HISTÓRICA (COLOMBIA)...")
        print("-" * 60)
        
        # 1. Calcular fecha del viernes anterior
        fecha_viernes = self.obtener_fecha_viernes_anterior()
        hoy = datetime.date.today()
        
        print(f"📅 Fecha actual: {hoy.strftime('%A, %Y-%m-%d')}")
        print(f"📅 Viernes objetivo: {fecha_viernes.strftime('%A, %Y-%m-%d')}")
        print(f"📅 Días atrás: {(hoy - fecha_viernes).days}")
        
        # 2. Intentar obtener tasa del viernes
        tasa, fecha_usada = self.obtener_tasa_colombia_fecha_historica(fecha_viernes)
        if tasa:
            return tasa, fecha_usada
        
        # 3. Respaldo: jueves anterior
        print("⚠️ No hay datos del viernes, intentando jueves...")
        fecha_jueves = fecha_viernes - datetime.timedelta(days=1)
        tasa, fecha_usada = self.obtener_tasa_colombia_fecha_historica(fecha_jueves)
        if tasa:
            return tasa, fecha_usada
        
        # 4. Respaldo: miércoles anterior  
        print("⚠️ No hay datos del jueves, intentando miércoles...")
        fecha_miercoles = fecha_viernes - datetime.timedelta(days=2)
        tasa, fecha_usada = self.obtener_tasa_colombia_fecha_historica(fecha_miercoles)
        if tasa:
            return tasa, fecha_usada
        
        # 5. Respaldo: martes anterior
        print("⚠️ No hay datos del miércoles, intentando martes...")
        fecha_martes = fecha_viernes - datetime.timedelta(days=3)
        tasa, fecha_usada = self.obtener_tasa_colombia_fecha_historica(fecha_martes)
        if tasa:
            return tasa, fecha_usada
        
        # 6. Respaldo: lunes anterior
        print("⚠️ No hay datos del martes, intentando lunes...")
        fecha_lunes = fecha_viernes - datetime.timedelta(days=4)
        tasa, fecha_usada = self.obtener_tasa_colombia_fecha_historica(fecha_lunes)
        if tasa:
            return tasa, fecha_usada
        
        # 7. Último recurso: tasa actual
        print("⚠️ Usando tasa actual como último recurso...")
        tasa, fecha_usada = self.obtener_tasa_colombia_actual()
        if tasa:
            return tasa, fecha_usada
        
        # 8. Tasa de respaldo fija (última opción)
        tasa_respaldo = 4000.0  # Tasa aproximada de respaldo para COP/USD
        fecha_respaldo = hoy
        print(f"📊 Usando tasa de respaldo fija: {tasa_respaldo} COP/USD")
        return tasa_respaldo, fecha_respaldo


class ExcelProcessor:
    """Procesador base para archivos Excel"""
        
    def __init__(self, pais, moneda, tasa_dolar, archivo_reporte_absoluto=None, lookup_solicitantes_areas=None, api_helper=None):
        self.pais = pais
        self.moneda = moneda
        self.tasa_dolar = tasa_dolar
        self.archivo_reporte_absoluto = archivo_reporte_absoluto
        self.df_absoluto = None
        self.lookup_integrado = {}
        
        # CORRECCIÓN: Nombre consistente del atributo
        self.lookup_solicitantes_areas = lookup_solicitantes_areas if lookup_solicitantes_areas is not None else {}
        
        # APIHelper para consultar tasas FTD
        self.api_helper = api_helper if api_helper is not None else APIHelper()
        
        # Precargar tasas FTD
        print(f"💱 Precargando tasas FTD...")
        self.api_helper.obtener_tasas_ftd()
        
        print(f"🔧 ExcelProcessor inicializado con {len(self.lookup_solicitantes_areas)} solicitantes mapeados")
        
        # Cargar Reporte Absoluto si existe
        if self.archivo_reporte_absoluto:
            self._cargar_reporte_absoluto_integrado()

    def obtener_area_para_solicitante(self, solicitante, proyecto=None):
        """
        Obtener área para solicitante usando lookup - Implementación fórmula Excel
        
        EXCEPCIONES:
        - Si proyecto = "A048" → retorna "AUTOPAGO"
        - Si solicitante es de TI y proyecto = "VENE" → retorna "DIR CONSTRUCCIÓN Y PROYECTOS"
        
        Args:
            solicitante: Nombre del solicitante
            proyecto: Código del proyecto (opcional)
        
        Returns:
            str: Área correspondiente
        """
        # EXCEPCIÓN 1: Si proyecto es A048, retornar AUTOPAGO directamente
        if proyecto and str(proyecto).strip().upper() == "A048":
            return "AUTOPAGO"
        
        # Si está vacío o es 0, devolver "SERVICIOS"
        if not solicitante or str(solicitante).strip() in ['', '0', 'nan', 'None', 'NAN']:
            return "SERVICIOS"
        
        # Si no hay lookup disponible
        if not self.lookup_solicitantes_areas:
            return "SIN_GOOGLE_SHEET"
        
        # Limpiar y buscar
        solicitante_clean = str(solicitante).strip().upper()
        proyecto_clean = str(proyecto).strip().upper() if proyecto else ""
        
        # Búsqueda exacta
        area_encontrada = None
        if solicitante_clean in self.lookup_solicitantes_areas:
            area_encontrada = self.lookup_solicitantes_areas[solicitante_clean]
        else:
            # Búsqueda parcial por palabras clave (apellidos)
            for sol_ref, area in self.lookup_solicitantes_areas.items():
                # Buscar por coincidencia parcial
                if solicitante_clean in sol_ref or sol_ref in solicitante_clean:
                    area_encontrada = area
                    break
                
                # Buscar por apellidos (última palabra de cada nombre)
                palabras_ref = sol_ref.split()
                palabras_buscar = solicitante_clean.split()
                
                if len(palabras_ref) > 0 and len(palabras_buscar) > 0:
                    # Comparar último apellido
                    if palabras_ref[-1] in palabras_buscar or palabras_buscar[-1] in palabras_ref:
                        area_encontrada = area
                        break
        
        # EXCEPCIÓN 2: Si solicitante es de TI y proyecto es VENE, asignar DIR CONSTRUCCIÓN Y PROYECTOS
        if area_encontrada:
            area_clean = str(area_encontrada).strip().upper()
            # Verificar si el área contiene "TI" (Tecnología de Información)
            if "TI" in area_clean or "TECNOLOGIA" in area_clean or "TECNOLOGÍA" in area_clean or "INFORMACION" in area_clean or "INFORMACIÓN" in area_clean:
                if proyecto_clean == "VENE":
                    return "DIR CONSTRUCCIÓN Y PROYECTOS"
        
        # Retornar área encontrada o área no encontrada
        if area_encontrada:
            return area_encontrada
        
        # No encontrado en Google Sheet
        return "AREA_NO_ENCONTRADA"


    def _cargar_reporte_absoluto_integrado(self):
        """Cargar y procesar el Reporte Absoluto para lookup integrado con filtros de limpieza"""
        try:
            print(f"\n📖 CARGANDO REPORTE ABSOLUTO INTEGRADO (5 CAMPOS)...")
            print("-" * 70)
            
            if not Path(self.archivo_reporte_absoluto).exists():
                print(f"❌ Reporte Absoluto no encontrado: {self.archivo_reporte_absoluto}")
                return
            
            # Leer Reporte Absoluto
            self.df_absoluto = pd.read_excel(self.archivo_reporte_absoluto)
            print(f"✅ Reporte Absoluto cargado: {len(self.df_absoluto)} filas, {len(self.df_absoluto.columns)} columnas")
            
            # ===================================================================
            # FASE 1: IDENTIFICAR COLUMNAS DE FILTRO
            # ===================================================================
            columnas_archivo = [str(col).strip() for col in self.df_absoluto.columns]
            
            col_tipo_linea = None
            col_categoria_compra = None
            col_cuenta_cargo_filtro = None
            
            for col in columnas_archivo:
                col_lower = str(col).lower()
                if 'tipo' in col_lower and 'línea' in col_lower or 'linea' in col_lower:
                    col_tipo_linea = col
                elif 'categoría' in col_lower and 'compra' in col_lower or ('categoria' in col_lower and 'compra' in col_lower):
                    col_categoria_compra = col
                elif 'cta' in col_lower and 'cargo' in col_lower and 'centro' not in col_lower and 'desc' not in col_lower:
                    col_cuenta_cargo_filtro = col
            
            print(f"\n🔍 COLUMNAS PARA FILTRADO:")
            print(f"   🔹 Tipo de Línea: '{col_tipo_linea}'" if col_tipo_linea else "   ❌ Tipo de Línea: NO ENCONTRADA")
            print(f"   🔹 Categoría de Compra: '{col_categoria_compra}'" if col_categoria_compra else "   ❌ Categoría de Compra: NO ENCONTRADA")
            print(f"   🔹 Cta. Cargo: '{col_cuenta_cargo_filtro}'" if col_cuenta_cargo_filtro else "   ❌ Cta. Cargo: NO ENCONTRADA")
            
            filas_iniciales = len(self.df_absoluto)
            
            # ===================================================================
            # FASE 2: APLICAR FILTROS DE LIMPIEZA
            # ===================================================================
            print(f"\n🧹 APLICANDO FILTROS DE LIMPIEZA...")
            
            # FILTRO 1: Tipo de Línea = "Artículo"
            if col_tipo_linea:
                self.df_absoluto = self.df_absoluto[
                    self.df_absoluto[col_tipo_linea].astype(str).str.strip().str.lower() == 'artículo'
                ]
                print(f"   ✅ Filtro 1 - Tipo de Línea='Artículo': {len(self.df_absoluto)} filas restantes")
            
            # FILTRO 2: Categoría de Compra = "CAPEX" o vacío
            if col_categoria_compra:
                def filtrar_categoria(valor):
                    if pd.isna(valor) or str(valor).strip() == '':
                        return True  # Mantener vacíos
                    valor_str = str(valor).strip().upper()
                    # Extraer la parte antes del punto
                    categoria_base = valor_str.split('.')[0] if '.' in valor_str else valor_str
                    return categoria_base == 'CAPEX'
                
                self.df_absoluto = self.df_absoluto[
                    self.df_absoluto[col_categoria_compra].apply(filtrar_categoria)
                ]
                print(f"   ✅ Filtro 2 - Categoría='CAPEX' o vacío: {len(self.df_absoluto)} filas restantes")
            
            # FILTRO 3: Cta. Cargo con segundo segmento = 110425 o 150199
            if col_cuenta_cargo_filtro:
                def filtrar_cuenta_cargo(valor):
                    if pd.isna(valor):
                        return False
                    valor_str = str(valor).strip()
                    segmentos = valor_str.split('-')
                    if len(segmentos) < 2:
                        return False
                    segundo_segmento = segmentos[1].strip()
                    return segundo_segmento in ['110425', '150199']
                
                self.df_absoluto = self.df_absoluto[
                    self.df_absoluto[col_cuenta_cargo_filtro].apply(filtrar_cuenta_cargo)
                ]
                print(f"   ✅ Filtro 3 - Cta. Cargo (segmento 2)='110425' o '150199': {len(self.df_absoluto)} filas restantes")
            
            filas_despues_filtros = len(self.df_absoluto)
            filas_eliminadas = filas_iniciales - filas_despues_filtros
            print(f"\n📊 Resumen de Filtrado:")
            print(f"   📥 Filas iniciales: {filas_iniciales}")
            print(f"   📤 Filas después de filtros: {filas_despues_filtros}")
            print(f"   🗑️  Filas eliminadas: {filas_eliminadas} ({(filas_eliminadas/filas_iniciales*100):.1f}%)")
            
            if filas_despues_filtros == 0:
                print("⚠️ ADVERTENCIA: No quedaron filas después de los filtros. Revisa los criterios.")
                return
            
            # ===================================================================
            # FASE 3: IDENTIFICAR COLUMNAS PARA LOOKUP
            # ===================================================================
            col_factura = None
            col_tienda = None
            col_ceco = None
            col_cuenta_cargo = None
            col_fecha_recibo = None
            col_descripcion = None
            
            for col in columnas_archivo:
                col_lower = str(col).lower()
                
                # Factura
                if not col_factura and any(palabra in col_lower for palabra in ['factura', 'n°']):
                    col_factura = col
                
                # Tienda (Cta. Cargo Centro Desc.)
                elif not col_tienda and all(palabra in col_lower for palabra in ['cta', 'cargo', 'centro', 'desc']):
                    col_tienda = col
                
                # CECO (Cta. Cargo Centro sin Desc.)
                elif not col_ceco and all(palabra in col_lower for palabra in ['cta', 'cargo', 'centro']) and 'desc' not in col_lower:
                    col_ceco = col
                
                # Cuenta Cargo (Cta. Cargo sin Centro)
                elif not col_cuenta_cargo and all(palabra in col_lower for palabra in ['cta', 'cargo']) and 'centro' not in col_lower:
                    col_cuenta_cargo = col
                
                # Fecha Recibo
                elif not col_fecha_recibo and 'fecha' in col_lower and ('recepción' in col_lower or 'recepcion' in col_lower):
                    col_fecha_recibo = col
                
                # Descripción
                elif not col_descripcion and ('descipción' in col_lower or 'descripcion' in col_lower or 'descripción' in col_lower):
                    col_descripcion = col
            
            # Log de columnas encontradas
            print(f"\n🔍 COLUMNAS IDENTIFICADAS PARA LOOKUP:")
            print(f"   📍 Factura: '{col_factura}'" if col_factura else "   ❌ Factura: NO ENCONTRADA")
            print(f"   🏪 Tienda: '{col_tienda}'" if col_tienda else "   ❌ Tienda: NO ENCONTRADA")
            print(f"   🏢 CECO: '{col_ceco}'" if col_ceco else "   ❌ CECO: NO ENCONTRADA")
            print(f"   📊 Cuenta Cargo: '{col_cuenta_cargo}'" if col_cuenta_cargo else "   ❌ Cuenta Cargo: NO ENCONTRADA")
            print(f"   📅 Fecha Recibo: '{col_fecha_recibo}'" if col_fecha_recibo else "   ❌ Fecha Recibo: NO ENCONTRADA")
            print(f"   📝 Descripción: '{col_descripcion}'" if col_descripcion else "   ❌ Descripción: NO ENCONTRADA")
            
            if not col_factura:
                print("❌ No se encontró columna de factura - no se puede crear lookup")
                return
            
            # ===================================================================
            # FASE 4: CREAR LOOKUP INTEGRADO
            # ===================================================================
            print(f"\n⚙️ PROCESANDO LOOKUP INTEGRADO (5 CAMPOS)...")
            
            facturas_procesadas = 0
            stats = {
                'con_tienda': 0,
                'con_ceco': 0,
                'con_proyecto': 0,
                'con_fecha_recibo': 0,
                'con_descripcion': 0
            }
            
            for idx, row in self.df_absoluto.iterrows():
                try:
                    factura = str(row[col_factura]).strip() if col_factura else ""
                    
                    if not factura or factura.lower() in ['nan', 'none', '']:
                        continue
                    
                    # Extraer todos los campos
                    datos = {}
                    
                    # TIENDA
                    datos['tienda'] = str(row[col_tienda]).strip() if col_tienda and pd.notna(row[col_tienda]) else "SIN_TIENDA"
                    
                    # CECO
                    datos['ceco'] = str(row[col_ceco]).strip() if col_ceco and pd.notna(row[col_ceco]) else "SIN_CECO"
                    
                    # PROYECTO (extraer de Cta. Cargo posición 35-38)
                    datos['proyecto'] = "SIN_PROYECTO"
                    if col_cuenta_cargo and pd.notna(row[col_cuenta_cargo]):
                        cuenta_cargo = str(row[col_cuenta_cargo]).strip()
                        if len(cuenta_cargo) >= 39:
                            datos['proyecto'] = cuenta_cargo[34:38]
                        else:
                            # Buscar patrón alternativo
                            import re
                            patron = re.search(r'-([A-Z]\d{3})-', cuenta_cargo)
                            if patron:
                                datos['proyecto'] = patron.group(1)
                    
                    # FECHA RECIBO
                    if col_fecha_recibo and pd.notna(row[col_fecha_recibo]):
                        fecha_val = row[col_fecha_recibo]
                        if isinstance(fecha_val, pd.Timestamp):
                            datos['fecha_recibo'] = fecha_val.strftime('%Y-%m-%d')
                        else:
                            datos['fecha_recibo'] = str(fecha_val).strip()
                    else:
                        # Si no hay fecha recibo, usar viernes de la semana pasada
                        from datetime import datetime, timedelta
                        
                        hoy = datetime.now()
                        # Calcular cuántos días han pasado desde el lunes (0=lunes, 6=domingo)
                        dias_desde_lunes = hoy.weekday()
                        # Retroceder al lunes de esta semana
                        lunes_esta_semana = hoy - timedelta(days=dias_desde_lunes)
                        # Retroceder 3 días más para llegar al viernes pasado
                        viernes_pasado = lunes_esta_semana - timedelta(days=3)
                        
                        datos['fecha_recibo'] = viernes_pasado.strftime('%Y-%m-%d')

                    
                    # DESCRIPCIÓN
                    datos['descripcion'] = str(row[col_descripcion]).strip() if col_descripcion and pd.notna(row[col_descripcion]) else "SIN_DESCRIPCION"
                    
                    # Guardar en lookup
                    self.lookup_integrado[factura] = datos
                    facturas_procesadas += 1
                    
                    # Actualizar estadísticas
                    if datos['tienda'] not in ['SIN_TIENDA', 'nan']:
                        stats['con_tienda'] += 1
                    if datos['ceco'] not in ['SIN_CECO', 'nan']:
                        stats['con_ceco'] += 1
                    if datos['proyecto'] not in ['SIN_PROYECTO', 'nan']:
                        stats['con_proyecto'] += 1
                    if datos['fecha_recibo'] not in ['SIN_FECHA_RECIBO', 'nan']:
                        stats['con_fecha_recibo'] += 1
                    if datos['descripcion'] not in ['SIN_DESCRIPCION', 'nan']:
                        stats['con_descripcion'] += 1
                        
                except Exception as row_error:
                    print(f"⚠️ Error procesando fila {idx}: {row_error}")
                    continue
            
            # Mostrar estadísticas
            print(f"\n📊 ESTADÍSTICAS LOOKUP INTEGRADO:")
            print(f"   🔑 Facturas procesadas: {facturas_procesadas}")
            for campo, cantidad in stats.items():
                porcentaje = (cantidad/facturas_procesadas*100) if facturas_procesadas > 0 else 0
                campo_nombre = campo.replace('con_', '').upper()
                print(f"   📋 Con {campo_nombre}: {cantidad} ({porcentaje:.1f}%)")
            
            # Mostrar muestras
            if facturas_procesadas > 0:
                print(f"\n💡 MUESTRAS DEL LOOKUP:")
                samples = list(self.lookup_integrado.items())[:2]
                for factura, datos in samples:
                    print(f"   '{factura}':")
                    print(f"      🏪 TIENDA: '{datos['tienda']}'")
                    print(f"      🏢 CECO: '{datos['ceco']}'")
                    print(f"      📊 PROYECTO: '{datos['proyecto']}'")
                    print(f"      📅 FECHA RECIBO: '{datos['fecha_recibo']}'")
                    print(f"      📝 DESCRIPCIÓN: '{datos['descripcion'][:50]}{'...' if len(datos['descripcion']) > 50 else ''}'")
            
        except Exception as e:
            print(f"❌ Error cargando Reporte Absoluto: {e}")
            import traceback
            traceback.print_exc()

    
    def obtener_datos_integrados_para_factura(self, numero_factura):
        """Obtener todos los datos para una factura específica"""
        datos_vacios = {
            'tienda': "SIN_REPORTE_ABSOLUTO",
            'ceco': "SIN_REPORTE_ABSOLUTO", 
            'proyecto': "SIN_REPORTE_ABSOLUTO",
            'fecha_recibo': "SIN_REPORTE_ABSOLUTO",
            'descripcion': "SIN_REPORTE_ABSOLUTO"
        }
        
        if not self.lookup_integrado:
            return datos_vacios
        
        factura_str = str(numero_factura).strip()
        
        # Búsqueda exacta
        if factura_str in self.lookup_integrado:
            return self.lookup_integrado[factura_str]
        
        # Búsqueda parcial
        for factura_ref, datos in self.lookup_integrado.items():
            if (factura_str.lower() in factura_ref.lower() or 
                factura_ref.lower() in factura_str.lower()):
                return datos
        
        # No encontrada
        return {
            'tienda': "FACTURA_NO_ENCONTRADA",
            'ceco': "FACTURA_NO_ENCONTRADA",
            'proyecto': "FACTURA_NO_ENCONTRADA",
            'fecha_recibo': "FACTURA_NO_ENCONTRADA",
            'descripcion': "FACTURA_NO_ENCONTRADA"
        }
    
    def crear_formula_monto_usd(self, fila, header_map):
        letra_moneda = header_map['Moneda']
        letra_monto = header_map['Monto']
        return f'=IF({letra_moneda}{fila}="{self.moneda}",{letra_monto}{fila}/{self.tasa_dolar},{letra_monto}{fila})'

    def crear_formula_categoria(self, fila, header_map):
        letra_monto_capex = header_map['MONTO A PAGAR CAPEX']  # Columna Y
        letra_monto_opex = header_map['MONTO A PAGAR OPEX']    # Columna Z
        
        return f'=IF(AND({letra_monto_capex}{fila}<>0,{letra_monto_opex}{fila}<>0),"MIXTA",IF({letra_monto_capex}{fila}<>0,"CAPEX","OPEX"))'

    def crear_formula_monto_capex(self, fila, header_map):
        letra_ext = header_map['Monto CAPEX EXT']
        letra_ord = header_map['Monto CAPEX ORD']
        letra_cadm = header_map['Monto CADM']
        letra_x = header_map['Monto USD']
        return f'=IF(AND({letra_ext}{fila}=0,{letra_ord}{fila}=0),0,(({letra_ext}{fila}+{letra_ord}{fila})/(({letra_ext}{fila}+{letra_ord}{fila})+{letra_cadm}{fila})*{letra_x}{fila}))'

    def crear_formula_monto_opex(self, fila, header_map):
        letra_ext = header_map['Monto CAPEX EXT']
        letra_ord = header_map['Monto CAPEX ORD']
        letra_cadm = header_map['Monto CADM']
        letra_x = header_map['Monto USD']
        return f'=IF(AND({letra_ext}{fila}=0,{letra_ord}{fila}=0),{letra_x}{fila},({letra_cadm}{fila}/({letra_ext}{fila}+{letra_ord}{fila}+{letra_cadm}{fila})*{letra_x}{fila}))'

    def crear_formula_validacion(self, fila, header_map):
        letra_x = header_map['Monto USD']
        letra_z = header_map['MONTO A PAGAR CAPEX']
        letra_aa = header_map['MONTO A PAGAR OPEX']
        return f'={letra_x}{fila}-{letra_z}{fila}-{letra_aa}{fila}'

    def crear_formula_metodo_pago(self, fila, header_map):
        letra_p_indep = header_map['Prioridad']
        return f'=IF(OR({letra_p_indep}{fila}=78,{letra_p_indep}{fila}=79,{letra_p_indep}{fila}=80),"VES",IF(OR({letra_p_indep}{fila}=71,{letra_p_indep}{fila}=72,{letra_p_indep}{fila}=77),"EUR","USD"))'

    def crear_formula_moneda_pago(self, fila, header_map):
        """
        Crear fórmula para MONEDA DE PAGO basada en la prioridad:
        - Si prioridad es 69,70,73,74,75,76 → USD
        - Si prioridad es 71,72,77 → EUR
        - Si prioridad es 78,79 → VES
        - De lo contrario → NA
        """
        letra_p = header_map['Prioridad']
        return f'=IF(OR({letra_p}{fila}=69,{letra_p}{fila}=70,{letra_p}{fila}=73,{letra_p}{fila}=74,{letra_p}{fila}=75,{letra_p}{fila}=76),"USD",IF(OR({letra_p}{fila}=71,{letra_p}{fila}=72,{letra_p}{fila}=77),"EUR",IF(OR({letra_p}{fila}=78,{letra_p}{fila}=79),"VES","NA")))'
    
    def crear_formula_conversion_ves(self, fila, header_map):
        """
        Crear fórmula para CONVERSION VES:
        =SI.ERROR(SI(MONEDA_PAGO="VES";MONTO_CAPEX*TC_BCV;0);0)
        """
        letra_moneda_pago = header_map['MONEDA DE PAGO']
        letra_monto_capex = header_map['MONTO A PAGAR CAPEX']
        letra_tc_bcv = header_map['TC BCV']
        return f'=IFERROR(IF({letra_moneda_pago}{fila}="VES",{letra_monto_capex}{fila}*{letra_tc_bcv}{fila},0),0)'
    
    def crear_formula_conversion_tc_ftd(self, fila, header_map):
        """
        Crear fórmula para CONVERSION TC FTD:
        =SI.ERROR(CONVERSION_VES/TC_FTD;0)
        """
        letra_conversion_ves = header_map['CONVERSION VES']
        letra_tc_ftd = header_map['TC FTD']
        return f'=IFERROR({letra_conversion_ves}{fila}/{letra_tc_ftd}{fila},0)'
    
    def crear_formula_real_convertido(self, fila, header_map):
        """
        Crear fórmula para REAL CONVERTIDO:
        =SI(MONEDA_PAGO="VES";CONVERSION_TC_FTD;MONTO_CAPEX)
        """
        letra_moneda_pago = header_map['MONEDA DE PAGO']
        letra_conversion_tc_ftd = header_map['CONVERSION TC FTD']
        letra_monto_capex = header_map['MONTO A PAGAR CAPEX']
        return f'=IF({letra_moneda_pago}{fila}="VES",{letra_conversion_tc_ftd}{fila},{letra_monto_capex}{fila})'
    
    def crear_formula_real_mes_convertido(self, fila, header_map):
        """
        Crear fórmula para REAL MES CONVERTIDO:
        Es una copia de REAL CONVERTIDO
        """
        letra_real_convertido = header_map['REAL CONVERTIDO']
        return f'={letra_real_convertido}{fila}'

    def crear_formula_tipo_capex(self, fila, header_map):
        letra_ext = header_map['Monto CAPEX EXT']
        letra_ord = header_map['Monto CAPEX ORD']
        return f'=IF(AND({letra_ext}{fila}<>0,{letra_ord}{fila}<>0),"MIXTA",IF({letra_ext}{fila}<>0,"EXT",IF({letra_ord}{fila}<>0,"ORD","N/A")))'

    def crear_formula_monto_ord(self, fila, header_map):
        letra_tipo = header_map['TIPO DE CAPEX']
        letra_z = header_map['Monto CAPEX EXT']
        letra_ord = header_map['Monto CAPEX ORD']
        return f'=IF({letra_tipo}{fila}="N/A",0,IF({letra_tipo}{fila}="EXT",0,IF({letra_tipo}{fila}="ORD",{letra_z}{fila},{letra_z}{fila}*({letra_ord}{fila}/({letra_z}{fila}+{letra_ord}{fila})))))'

    def crear_formula_monto_ext(self, fila, header_map):
        letra_tipo = header_map['TIPO DE CAPEX']
        letra_z = header_map['Monto CAPEX EXT']
        letra_ext = header_map['Monto CAPEX EXT']
        letra_ord = header_map['Monto CAPEX ORD']
        return f'=IF({letra_tipo}{fila}="N/A",0,IF({letra_tipo}{fila}="ORD",0,IF({letra_tipo}{fila}="EXT",{letra_z}{fila},{letra_z}{fila}*({letra_ext}{fila}/({letra_ext}{fila}+{letra_ord}{fila})))))'

    def crear_formula_dia_pago(self, fila, header_map):
        letra_p_indep = header_map['Prioridad']
        return f'=IF(OR({letra_p_indep}{fila}=78,{letra_p_indep}{fila}=79,{letra_p_indep}{fila}=80),"JUEVES","VIERNES")'

    def calcular_monto_usd(self, df):
        """
        Calcula Monto USD: si Moneda == moneda, divide Monto/tasa; si no, retorna Monto.
        """
        df['Monto USD'] = df.apply(
            lambda row: row['Monto'] / self.tasa_dolar if row['Moneda'] == self.moneda else row['Monto'],
            axis=1
        )
        return df

    def calcular_categoria(self, df):
        """
        CATEGORIA: MIXTA si ambos montos CAPEX > 0, CAPEX si solo EXT > 0, sino OPEX.
        """
        def categoria_fila(row):
            capex_ext = row['MONTO A PAGAR CAPEX']
            capex_ord = row['MONTO A PAGAR OPEX']
            if capex_ext != 0 and capex_ord != 0:
                return "MIXTA"
            elif capex_ext != 0:
                return "CAPEX"
            else:
                return "OPEX"
        
        df['CATEGORIA'] = df.apply(categoria_fila, axis=1)
        return df

    def calcular_monto_capex(self, df):
        """
        MONTO A PAGAR CAPEX: si ambos CAPEX=0 retorna 0, sino calcula proporción sobre Monto USD.
        """
        def monto_capex_fila(row):
            ext = row['Monto CAPEX EXT']
            ord = row['Monto CAPEX ORD']
            cadm = row['Monto CADM']
            usd = row['Monto USD']
            if ext == 0 and ord == 0:
                return 0
            return ((ext + ord) / ((ext + ord) + cadm)) * usd
        
        df['MONTO A PAGAR CAPEX'] = df.apply(monto_capex_fila, axis=1)
        return df

    def calcular_monto_opex(self, df):
        """
        MONTO A PAGAR OPEX: si ambos CAPEX=0 retorna Monto USD, sino proporción de CADM.
        """
        def monto_opex_fila(row):
            ext = row['Monto CAPEX EXT']
            ord = row['Monto CAPEX ORD']
            cadm = row['Monto CADM']
            usd = row['Monto USD']
            if ext == 0 and ord == 0:
                return usd
            return (cadm / (ext + ord + cadm)) * usd
        
        df['MONTO A PAGAR OPEX'] = df.apply(monto_opex_fila, axis=1)
        return df

    def calcular_validacion(self, df):
        """
        VALIDACION: Monto USD - MONTO A PAGAR CAPEX - MONTO A PAGAR OPEX
        """
        df['VALIDACION'] = df['Monto USD'] - df['MONTO A PAGAR CAPEX'] - df['MONTO A PAGAR OPEX']
        return df

    def calcular_metodo_pago(self, df):
        """
        METODO DE PAGO: VES si Pago Independiente = 78, 79, 80; EUR si 71, 72, 77; sino USD.
        """
        def metodo_pago_fila(row):
            p = row['Prioridad']
            if p in [78, 79, 80]:
                return "VES"
            elif p in [71, 72, 77]:
                return "EUR"
            else:
                return "USD"
        
        df['METODO DE PAGO'] = df.apply(metodo_pago_fila, axis=1)
        return df

    def calcular_moneda_pago(self, df):
        """
        MONEDA DE PAGO basada en la prioridad:
        - Si prioridad es 69,70,73,74,75,76 → USD
        - Si prioridad es 71,72,77 → EUR
        - Si prioridad es 78,79 → VES
        - De lo contrario → NA
        """
        def moneda_pago_fila(row):
            p = row['Prioridad']
            if p in [69, 70, 73, 74, 75, 76]:
                return "USD"
            elif p in [71, 72, 77]:
                return "EUR"
            elif p in [78, 79]:
                return "VES"
            else:
                return "NA"
        
        df['MONEDA DE PAGO'] = df.apply(moneda_pago_fila, axis=1)
        return df
    
    def calcular_conversion_ves(self, df):
        """
        CONVERSION VES: Si MONEDA DE PAGO es "VES", entonces MONTO A PAGAR CAPEX * TC BCV, sino 0
        """
        def conversion_ves_fila(row):
            try:
                if row['MONEDA DE PAGO'] == "VES":
                    monto = row.get('MONTO A PAGAR CAPEX', 0) or 0
                    tc_bcv = row.get('TC BCV', 0) or 0
                    return monto * tc_bcv
                return 0
            except:
                return 0
        
        df['CONVERSION VES'] = df.apply(conversion_ves_fila, axis=1)
        return df
    
    def calcular_conversion_tc_ftd(self, df):
        """
        CONVERSION TC FTD: CONVERSION VES / TC FTD (con manejo de división por cero)
        """
        def conversion_tc_ftd_fila(row):
            try:
                conversion_ves = row.get('CONVERSION VES', 0) or 0
                tc_ftd = row.get('TC FTD', 0) or 0
                if tc_ftd != 0:
                    return conversion_ves / tc_ftd
                return 0
            except:
                return 0
        
        df['CONVERSION TC FTD'] = df.apply(conversion_tc_ftd_fila, axis=1)
        return df
    
    def calcular_real_convertido(self, df):
        """
        REAL CONVERTIDO: Si MONEDA DE PAGO es "VES", usa CONVERSION TC FTD, sino usa MONTO A PAGAR CAPEX
        """
        def real_convertido_fila(row):
            try:
                if row.get('MONEDA DE PAGO', '') == "VES":
                    return row.get('CONVERSION TC FTD', 0) or 0
                return row.get('MONTO A PAGAR CAPEX', 0) or 0
            except:
                return 0
        
        df['REAL CONVERTIDO'] = df.apply(real_convertido_fila, axis=1)
        return df
    
    def calcular_real_mes_convertido(self, df):
        """
        REAL MES CONVERTIDO: Copia de REAL CONVERTIDO
        """
        df['REAL MES CONVERTIDO'] = df['REAL CONVERTIDO']
        return df

    def calcular_tipo_capex(self, df):
        """
        TIPO DE CAPEX: MIXTA si ambos > 0, EXT si solo EXT > 0, ORD si solo ORD > 0, sino N/A.
        """
        def tipo_capex_fila(row):
            ext = row['Monto CAPEX EXT']
            ord = row['Monto CAPEX ORD']
            if ext != 0 and ord != 0:
                return "MIXTA"
            elif ext != 0:
                return "EXT"
            elif ord != 0:
                return "ORD"
            else:
                return "N/A"
        
        df['TIPO DE CAPEX'] = df.apply(tipo_capex_fila, axis=1)
        return df

    def calcular_monto_ord(self, df):
        """
        MONTO ORD: distribuye según TIPO DE CAPEX y proporciones.
        """
        def monto_ord_fila(row):
            tipo = row['TIPO DE CAPEX']
            ext = row['Monto CAPEX EXT']
            ord = row['Monto CAPEX ORD']
            capex = row['MONTO A PAGAR CAPEX']
            
            if tipo == "N/A" or tipo == "EXT":
                return 0
            elif tipo == "ORD":
                return capex
            else:  # MIXTA
                return capex * (ord / (ext + ord))
        
        df['MONTO ORD'] = df.apply(monto_ord_fila, axis=1)
        return df

    def calcular_monto_ext(self, df):
        """
        MONTO EXT: distribuye según TIPO DE CAPEX y proporciones.
        """
        def monto_ext_fila(row):
            tipo = row['TIPO DE CAPEX']
            ext = row['Monto CAPEX EXT']
            ord = row['Monto CAPEX ORD']
            capex = row['MONTO A PAGAR CAPEX']
            
            if tipo == "N/A" or tipo == "ORD":
                return 0
            elif tipo == "EXT":
                return capex
            else:  # MIXTA
                return capex * (ext / (ext + ord))
        
        df['MONTO EXT'] = df.apply(monto_ext_fila, axis=1)
        return df

    def calcular_dia_pago(self, df):
        """
        DIA DE PAGO: JUEVES si Prioridad = 78, 79, 80; sino VIERNES.
        """
        df['DIA DE PAGO'] = df['Prioridad'].apply(
            lambda p: "JUEVES" if p in [78, 79, 80] else "VIERNES"
        )
        return df


    
    def _obtener_viernes_pasado(self):
        """
        Calcula la fecha del viernes de la semana pasada.
        Ejemplo: Si hoy es lunes 1 de diciembre, retorna el viernes 28 de noviembre.
        """
        hoy = datetime.date.today()
        dia_semana_actual = hoy.weekday()  # lunes=0, viernes=4, domingo=6
        
        # Calcular días hasta el viernes de esta semana
        dias_hasta_viernes_esta_semana = (4 - dia_semana_actual) % 7
        
        # Si hoy es viernes (dias_hasta_viernes_esta_semana = 0), el viernes pasado fue hace 7 días
        # Si no, el viernes pasado fue hace (dias_hasta_viernes_esta_semana + 7) días
        if dias_hasta_viernes_esta_semana == 0:
            dias_retroceso = 7
        else:
            dias_retroceso = dias_hasta_viernes_esta_semana + 7
        
        return hoy - datetime.timedelta(days=dias_retroceso)
    
    def obtener_semana_actual(self):
        """
        Obtiene el número de semana del mes basado en el viernes de la semana pasada.
        Ejemplo: Si hoy es lunes 1 de diciembre, toma el viernes pasado (28 de noviembre),
        entonces semana = 4 (cuarta semana de noviembre).
        
        Regla especial: Si el viernes pasado está en el mes anterior y hoy es del mes siguiente
        (no inclusivo del lunes), entonces el viernes pasado es semana 4.
        """
        viernes_pasado = self._obtener_viernes_pasado()
        hoy = datetime.date.today()
        
        # Calcular semana del mes basada en el viernes pasado
        primer_dia_mes = viernes_pasado.replace(day=1)
        dias_transcurridos = (viernes_pasado - primer_dia_mes).days
        
        # Calcular qué día de la semana es el día 1 del mes (lunes=0, domingo=6)
        dia_semana_primer_dia = primer_dia_mes.weekday()
        
        # Calcular en qué semana del mes está el viernes pasado
        # La semana 1 empieza el lunes de la semana que contiene el día 1
        # Si el día 1 es lunes, semana 1 = días 1-7
        # Si el día 1 es martes, semana 1 incluye el lunes anterior (último día del mes anterior)
        # Necesitamos calcular cuántas semanas completas han pasado desde el lunes de la semana del día 1
        
        # Encontrar el lunes de la semana que contiene el día 1
        dias_retroceso_lunes = dia_semana_primer_dia  # días desde el lunes hasta el día 1
        lunes_semana_1 = primer_dia_mes - datetime.timedelta(days=dias_retroceso_lunes)
        
        # Calcular días desde el lunes de la semana 1 hasta el viernes pasado
        dias_desde_lunes_semana_1 = (viernes_pasado - lunes_semana_1).days
        
        # Calcular semana (cada 7 días es una semana, empezando desde 1)
        semana = (dias_desde_lunes_semana_1 // 7) + 1
        
        # Regla especial: Si el viernes pasado está en el mes anterior y hoy es del mes siguiente
        # (especialmente si hoy es lunes), entonces el viernes pasado es semana 4
        if viernes_pasado.month < hoy.month:
            # Si el viernes pasado está en los días 22-31 del mes anterior, es semana 4
            if viernes_pasado.day >= 22:
                semana = 4
        
        # Asegurar que si el viernes pasado está en días 22-28, es semana 4
        # (independientemente del mes, si está en esos días, es la cuarta semana)
        if viernes_pasado.day >= 22 and viernes_pasado.day <= 28:
            semana = 4
        
        return semana
    
    def obtener_mes_actual(self):
        """
        Obtiene el mes basado en el viernes de la semana pasada.
        Ejemplo: Si hoy es lunes 1 de diciembre, toma el viernes pasado (28 de noviembre),
        entonces mes = "NOVIEMBRE".
        """
        viernes_pasado = self._obtener_viernes_pasado()
        
        meses = {
            1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL",
            5: "MAYO", 6: "JUNIO", 7: "JULIO", 8: "AGOSTO",
            9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE"
        }
        return meses[viernes_pasado.month]
    
    def obtener_anio_fiscal_actual(self):
        """
        Obtener año fiscal actual basado en ciclo Agosto-Julio
        Ejemplos:
        - Septiembre 2025 → "2025-2026" 
        - Julio 2026 → "2025-2026"
        - Agosto 2026 → "2026-2027"
        """
        hoy = datetime.date.today()
        
        # Si estamos en Agosto o después (Agosto = 8), el año fiscal comienza este año
        if hoy.month >= 8:
            anio_inicio = hoy.year
            anio_fin = hoy.year + 1
        else:
            # Si estamos antes de Agosto (Enero-Julio), el año fiscal comenzó el año anterior
            anio_inicio = hoy.year - 1
            anio_fin = hoy.year
        
        return f"{anio_inicio}-{anio_fin}"

    def obtener_anio_fiscal_para_fecha(self, fecha):
        """
        Obtener año fiscal para una fecha específica
        Args:
            fecha: datetime.date, datetime.datetime, string, o timestamp de pandas
        """
        try:
            # Convertir diferentes tipos de fecha a datetime.date
            if isinstance(fecha, str):
                # Intentar varios formatos de fecha
                formatos_fecha = [
                    '%Y-%m-%d',
                    '%d/%m/%Y', 
                    '%m/%d/%Y',
                    '%Y-%m-%d %H:%M:%S'
                ]
                
                fecha_obj = None
                for formato in formatos_fecha:
                    try:
                        fecha_obj = datetime.datetime.strptime(fecha, formato).date()
                        break
                    except ValueError:
                        continue
                
                if fecha_obj is None:
                    return "FECHA_INVALIDA"
                    
            elif isinstance(fecha, pd.Timestamp):
                fecha_obj = fecha.date()
            elif isinstance(fecha, datetime.datetime):
                fecha_obj = fecha.date()
            elif isinstance(fecha, datetime.date):
                fecha_obj = fecha
            else:
                return "FECHA_INVALIDA"
            
            # Calcular año fiscal para la fecha específica
            if fecha_obj.month >= 8:
                anio_inicio = fecha_obj.year
                anio_fin = fecha_obj.year + 1
            else:
                anio_inicio = fecha_obj.year - 1
                anio_fin = fecha_obj.year
            
            return f"{anio_inicio}-{anio_fin}"
            
        except Exception as e:
            return "FECHA_INVALIDA"

    def crear_archivo_consolidado(self, df, nombre_archivo):
        """Crear archivo Excel consolidado con manejo robusto de errores"""
        try:
            print(f"📝 Creando archivo: {nombre_archivo}")
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "BOSQUETO"
            
            # Headers completos - EXACTAMENTE 49 columnas (se agrega REAL CONVERTIDO, REAL MES CONVERTIDO)
            headers_venezuela = [
                # Columnas originales (1-22) - SIN Proveedor Remito
                "Numero de Factura", "Numero de OC", "Tipo Factura", "Nombre Lote",
                "Proveedor", "RIF", "Fecha Documento", "Tienda", "Sucursal",
                "Monto", "Moneda", "Fecha Vencimiento", "Cuenta", "Id Cta",
                "Método de Pago", "Pago Independiente", "Prioridad",
                "Monto CAPEX EXT", "Monto CAPEX ORD", "Monto CADM",
                "Fecha Creación", "Solicitante", 
                # Columnas calculadas (23-49) - 27 columnas
                "Monto USD", "CATEGORIA", "MONTO A PAGAR CAPEX", "MONEDA DE PAGO", "FECHA PAGO", "TC FTD",
                "TC BCV", "CONVERSION VES", "CONVERSION TC FTD", "REAL CONVERTIDO", "REAL MES CONVERTIDO",
                "MONTO A PAGAR OPEX", "VALIDACION", "METODO DE PAGO", "SEMANA", "MES DE PAGO",
                "TIPO DE CAPEX", "MONTO ORD", "MONTO EXT", "DIA DE PAGO",
                "TIENDA_LOOKUP", "CECO", "PROYECTO", "AREA", "FECHA RECIBO", "DESCRIPCIÓN",
                "AÑO FISCAL"
            ]

            def col_letra(idx):
                """Devuelve la letra Excel para un índice base 1"""
                letras = ''
                while idx > 0:
                    idx, rem = divmod(idx-1, 26)
                    letras = chr(65 + rem) + letras
                return letras

            header_map = {header: col_letra(idx + 1) for idx, header in enumerate(headers_venezuela)}
            
            # DEBUG: Verificar que Prioridad esté en el header_map
            if 'Prioridad' in header_map:
                print(f"✅ Prioridad está en header_map: columna {header_map['Prioridad']}")
            else:
                print(f"❌ ERROR: Prioridad NO está en header_map")
                print(f"   Headers disponibles: {list(header_map.keys())}")
            
            # DEBUG: Mostrar fórmula de ejemplo para MONEDA DE PAGO
            formula_ejemplo = self.crear_formula_moneda_pago(2, header_map)
            print(f"📝 Fórmula MONEDA DE PAGO ejemplo: {formula_ejemplo[:80]}...")
            
            # Verificar conteo correcto
            total_headers = len(headers_venezuela)
            print(f"📋 Headers consolidado: {total_headers} columnas")
            print(f"📊 Columna MONEDA DE PAGO en posición 26 (Z)")
            print(f"📊 Columna FECHA PAGO en posición 27 (AA)")
            print(f"📊 Columna TC FTD en posición 28 (AB)")
            print(f"📊 Columna TC BCV en posición 29 (AC)")
            print(f"📊 Columna CONVERSION VES en posición 30 (AD)")
            print(f"📊 Columna CONVERSION TC FTD en posición 31 (AE)")
            print(f"📊 Columna REAL CONVERTIDO en posición 32 (AF)")
            print(f"📊 Columna REAL MES CONVERTIDO en posición 33 (AG)")
            

            # Obtener valores actuales
            semana_actual = self.obtener_semana_actual()
            mes_actual = self.obtener_mes_actual()
            anio_fiscal_actual = self.obtener_anio_fiscal_actual()
            
            print(f"📅 Semana actual: {semana_actual}")
            print(f"📅 Mes actual: {mes_actual}")
            print(f"📅 Año fiscal actual: {anio_fiscal_actual}")
            
            # Escribir headers
            for col_idx, header in enumerate(headers_venezuela, 1):
                try:
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                except Exception as header_error:
                    print(f"❌ Error escribiendo header {col_idx}: {header_error}")
                    return False
            
            # Estadísticas
            stats = {
                'tiendas_encontradas': 0,
                'cecos_encontrados': 0,
                'proyectos_encontrados': 0,
                'fechas_recibo_encontradas': 0,
                'descripciones_encontradas': 0,
                'facturas_no_encontradas': 0
            }
            
            # OPTIMIZACIÓN: Buscar columna "Fecha de Pago" UNA VEZ antes del loop
            col_fecha_pago = None
            for col_name in df.columns:
                if 'fecha' in str(col_name).lower() and 'pago' in str(col_name).lower():
                    col_fecha_pago = col_name
                    print(f"📅 Columna Fecha de Pago encontrada: '{col_fecha_pago}'")
                    break
            
            if col_fecha_pago is None:
                print(f"⚠️ Columna 'Fecha de Pago' no encontrada en el DataFrame")
            
            # ================================================================
            # OPTIMIZACIÓN DE RENDIMIENTO: Pre-calcular tasas para fechas únicas
            # ================================================================
            print(f"\n⚡ OPTIMIZACIÓN: Pre-calculando tasas para fechas únicas...")
            
            # Extraer todas las fechas de pago únicas
            fechas_unicas = set()
            if col_fecha_pago is not None:
                for valor_fecha in df[col_fecha_pago].dropna():
                    if pd.notna(valor_fecha):
                        if isinstance(valor_fecha, (pd.Timestamp, datetime.datetime, datetime.date)):
                            fecha_str = valor_fecha.strftime('%Y-%m-%d') if hasattr(valor_fecha, 'strftime') else str(valor_fecha)
                        else:
                            fecha_str = str(valor_fecha)
                        fechas_unicas.add(fecha_str)
            
            print(f"   📅 Fechas únicas encontradas: {len(fechas_unicas)}")
            
            # Pre-cargar cache de tasas FTD (una sola vez)
            tasas_cache_ftd = self.api_helper.obtener_tasas_ftd()
            print(f"   💱 Cache de tasas FTD cargado: {len(tasas_cache_ftd)} fechas disponibles")
            
            # Pre-cargar cache de tasas BCV desde BigQuery (una sola vez)
            tasas_cache_bcv = {}
            if precargar_tasas_bcv is not None:
                print(f"   💱 Cargando tasas BCV desde BigQuery...")
                tasas_cache_bcv = precargar_tasas_bcv()
                print(f"   💱 Cache de tasas BCV cargado: {len(tasas_cache_bcv)} fechas disponibles")
            else:
                print(f"   ⚠️ Módulo de tasas BCV no disponible, TC_BCV será 0")
            
            # Pre-calcular tasas para cada fecha única
            # TC_FTD: desde endpoint FTD (tasa_farmatodo)
            # TC_BCV: desde BigQuery (cxp_vzla.bcv_tasas)
            tasas_por_fecha = {}
            
            for fecha_str in fechas_unicas:
                # Normalizar fecha
                fecha_normalizada = self.api_helper._normalizar_fecha_str(fecha_str)
                
                # TC_FTD: desde endpoint FTD
                tc_ftd = 0
                if fecha_normalizada in tasas_cache_ftd:
                    tc_ftd = tasas_cache_ftd[fecha_normalizada].get('tasa_farmatodo', 0)
                
                # TC_BCV: desde BigQuery (cache pre-cargado)
                tc_bcv = tasas_cache_bcv.get(fecha_normalizada, 0)
                
                tasas_por_fecha[fecha_str] = {
                    'tc_ftd': tc_ftd,
                    'tc_bcv': tc_bcv
                }
            
            print(f"   ✅ Tasas pre-calculadas para {len(tasas_por_fecha)} fechas")
            
            # Mostrar algunas muestras
            if tasas_por_fecha:
                muestras = list(tasas_por_fecha.items())[:2]
                for fecha, tasas in muestras:
                    print(f"      📌 {fecha}: TC_FTD={tasas['tc_ftd']}, TC_BCV={tasas['tc_bcv']}")
            
            # ================================================================
            # FIN OPTIMIZACIÓN
            # ================================================================
            
            # Escribir datos fila por fila
            print(f"\n📝 Procesando {len(df)} filas...")
            for row_idx in range(len(df)):
                try:
                    fila_excel = row_idx + 2
                    
                    # Copiar datos originales (22 columnas)
                    for col_idx in range(min(len(df.columns), 22)):
                        valor = df.iloc[row_idx, col_idx]
                        # Manejar valores problemáticos
                        if pd.isna(valor):
                            valor = ""
                        
                        # CORRECCIÓN: Convertir Prioridad a número entero para que las fórmulas funcionen
                        # La columna Prioridad está en posición 16 (índice 16, columna Q)
                        if col_idx == 17:  # Columna Prioridad
                            try:
                                if valor != "" and pd.notna(valor):
                                    valor = int(float(str(valor)))
                            except (ValueError, TypeError):
                                pass  # Mantener valor original si no se puede convertir
                        
                        ws.cell(row=fila_excel, column=col_idx + 1, value=valor)
                    
                    # Obtener datos integrados
                    numero_factura = df.iloc[row_idx, 0]
                    datos_integrados = self.obtener_datos_integrados_para_factura(numero_factura)
                    
                    # NUEVA LÓGICA: Obtener área desde Google Sheets usando Solicitante (columna 22)
                    solicitante = df.iloc[row_idx, 21] if len(df.columns) > 21 else ""  # Columna W (22)
                    proyecto = datos_integrados['proyecto']
                    area_calculada = self.obtener_area_para_solicitante(solicitante, proyecto)
                    
                    # OPTIMIZADO: Obtener Fecha de Pago usando la columna pre-identificada
                    fecha_pago = ""
                    if col_fecha_pago is not None:
                        valor_fecha = df.iloc[row_idx][col_fecha_pago]
                        if pd.notna(valor_fecha):
                            if isinstance(valor_fecha, pd.Timestamp):
                                fecha_pago = valor_fecha.strftime('%Y-%m-%d')
                            elif isinstance(valor_fecha, datetime.datetime):
                                fecha_pago = valor_fecha.strftime('%Y-%m-%d')
                            elif isinstance(valor_fecha, datetime.date):
                                fecha_pago = valor_fecha.strftime('%Y-%m-%d')
                            else:
                                fecha_pago = str(valor_fecha)
                    
                    # OPTIMIZADO: Obtener TC FTD y TC BCV desde el cache pre-calculado
                    # En lugar de llamar a funciones para cada fila, usamos el diccionario pre-calculado
                    if fecha_pago and fecha_pago in tasas_por_fecha:
                        tc_ftd = tasas_por_fecha[fecha_pago]['tc_ftd']
                        tc_bcv = tasas_por_fecha[fecha_pago]['tc_bcv']
                    else:
                        tc_ftd = 0
                        tc_bcv = 0
                    
                    # ================================================================
                    # CALCULAR MONEDA DE PAGO DIRECTAMENTE (sin fórmulas de Excel)
                    # ================================================================
                    # Diccionario de prioridades -> moneda de pago
                    PRIORIDADES_USD = [60, 69, 70, 73, 74, 75, 76]
                    PRIORIDADES_EUR = [71, 72, 77]
                    PRIORIDADES_VES = [78, 79, 80, 91]
                    
                    # Obtener el valor de Prioridad de la fila actual (columna 17, índice 16)
                    prioridad_valor = df.iloc[row_idx, 16] if len(df.columns) > 16 else None
                    
                    # Convertir a entero para comparación
                    try:
                        prioridad_int = int(float(str(prioridad_valor))) if pd.notna(prioridad_valor) and str(prioridad_valor).strip() != '' else None
                    except (ValueError, TypeError):
                        prioridad_int = None
                    
                    # Determinar MONEDA DE PAGO basado en la prioridad
                    if prioridad_int in PRIORIDADES_USD:
                        moneda_pago = "USD"
                    elif prioridad_int in PRIORIDADES_EUR:
                        moneda_pago = "EUR"
                    elif prioridad_int in PRIORIDADES_VES:
                        moneda_pago = "VES"
                    else:
                        moneda_pago = "NA"
                    
                    # DEBUG: Mostrar cálculo de MONEDA DE PAGO para las primeras 3 filas
                    if row_idx < 3:
                        print(f"   🔍 Fila {row_idx+1}: Prioridad={prioridad_valor} → int={prioridad_int} → MONEDA_PAGO={moneda_pago}")
                    
                    # ================================================================
                    # CALCULAR CONVERSIONES DIRECTAMENTE (sin fórmulas de Excel)
                    # Necesitamos el valor de MONTO A PAGAR CAPEX primero
                    # ================================================================
                    
                    # Obtener valores necesarios para cálculos
                    monto_capex_ext = df.iloc[row_idx, 17] if len(df.columns) > 17 else 0  # Columna R (18)
                    monto_capex_ord = df.iloc[row_idx, 18] if len(df.columns) > 18 else 0  # Columna S (19)
                    monto_cadm = df.iloc[row_idx, 19] if len(df.columns) > 19 else 0       # Columna T (20)
                    monto_original = df.iloc[row_idx, 9] if len(df.columns) > 9 else 0    # Columna J (10) - Monto
                    moneda_original = df.iloc[row_idx, 10] if len(df.columns) > 10 else "" # Columna K (11) - Moneda
                    
                    # Convertir a números seguros
                    try:
                        monto_capex_ext = float(monto_capex_ext) if pd.notna(monto_capex_ext) else 0
                    except:
                        monto_capex_ext = 0
                    try:
                        monto_capex_ord = float(monto_capex_ord) if pd.notna(monto_capex_ord) else 0
                    except:
                        monto_capex_ord = 0
                    try:
                        monto_cadm = float(monto_cadm) if pd.notna(monto_cadm) else 0
                    except:
                        monto_cadm = 0
                    try:
                        monto_original = float(monto_original) if pd.notna(monto_original) else 0
                    except:
                        monto_original = 0
                    
                    # Calcular MONTO USD (igual que la fórmula)
                    if str(moneda_original).strip().upper() == self.moneda:
                        monto_usd = monto_original / self.tasa_dolar if self.tasa_dolar != 0 else 0
                    else:
                        monto_usd = monto_original
                    
                    # Calcular MONTO A PAGAR CAPEX
                    if monto_capex_ext == 0 and monto_capex_ord == 0:
                        monto_a_pagar_capex = 0
                    else:
                        suma_capex = monto_capex_ext + monto_capex_ord
                        total = suma_capex + monto_cadm
                        if total != 0:
                            monto_a_pagar_capex = (suma_capex / total) * monto_usd
                        else:
                            monto_a_pagar_capex = 0
                    
                    # Calcular CONVERSION VES: Si MONEDA_PAGO="VES", entonces MONTO_CAPEX * TC_BCV
                    if moneda_pago == "VES" and tc_bcv != 0:
                        conversion_ves = monto_a_pagar_capex * tc_bcv
                    else:
                        conversion_ves = 0
                    
                    # Calcular CONVERSION TC FTD: CONVERSION_VES / TC_FTD
                    if tc_ftd != 0 and conversion_ves != 0:
                        conversion_tc_ftd = conversion_ves / tc_ftd
                    else:
                        conversion_tc_ftd = 0
                    
                    # Calcular REAL CONVERTIDO: Si MONEDA_PAGO="VES" usar CONVERSION_TC_FTD, sino usar MONTO_CAPEX
                    if moneda_pago == "VES":
                        real_convertido = conversion_tc_ftd
                    else:
                        real_convertido = monto_a_pagar_capex
                    
                    # REAL MES CONVERTIDO es igual a REAL CONVERTIDO
                    real_mes_convertido = real_convertido
                    
                    # DEBUG: Mostrar cálculo de conversiones para las primeras 3 filas
                    if row_idx < 3:
                        print(f"   💱 Fila {row_idx+1}: TC_FTD={tc_ftd}, TC_BCV={tc_bcv}, CONV_VES={conversion_ves:.2f}, CONV_TC_FTD={conversion_tc_ftd:.2f}, REAL={real_convertido:.2f}")

                    # Actualizar estadísticas
                    valores_no_encontrados = ["SIN_REPORTE_ABSOLUTO", "FACTURA_NO_ENCONTRADA"]
                    
                    if datos_integrados['tienda'] not in valores_no_encontrados + ["SIN_TIENDA"]:
                        stats['tiendas_encontradas'] += 1
                    if datos_integrados['ceco'] not in valores_no_encontrados + ["SIN_CECO"]:
                        stats['cecos_encontrados'] += 1
                    if datos_integrados['proyecto'] not in valores_no_encontrados + ["SIN_PROYECTO"]:
                        stats['proyectos_encontrados'] += 1
                    if datos_integrados['fecha_recibo'] not in valores_no_encontrados + ["SIN_FECHA_RECIBO"]:
                        stats['fechas_recibo_encontradas'] += 1
                    if datos_integrados['descripcion'] not in valores_no_encontrados + ["SIN_DESCRIPCION"]:
                        stats['descripciones_encontradas'] += 1
                    if datos_integrados['tienda'] == "FACTURA_NO_ENCONTRADA":
                        stats['facturas_no_encontradas'] += 1
                    
                    valores_calculados = [
                        self.crear_formula_monto_usd(fila_excel, header_map),
                        self.crear_formula_categoria(fila_excel, header_map),
                        self.crear_formula_monto_capex(fila_excel, header_map),
                        moneda_pago,           # MONEDA DE PAGO (valor calculado directamente)
                        fecha_pago,            # FECHA PAGO (del archivo de entrada)
                        tc_ftd,                # TC FTD (Tasa Farmatodo según fecha de pago)
                        tc_bcv,                # TC BCV (Tasa BCV según fecha de pago)
                        conversion_ves,        # CONVERSION VES (valor calculado directamente)
                        conversion_tc_ftd,     # CONVERSION TC FTD (valor calculado directamente)
                        real_convertido,       # REAL CONVERTIDO (valor calculado directamente)
                        real_mes_convertido,   # REAL MES CONVERTIDO (valor calculado directamente)
                        self.crear_formula_monto_opex(fila_excel, header_map),
                        self.crear_formula_validacion(fila_excel, header_map),
                        self.crear_formula_metodo_pago(fila_excel, header_map),
                        semana_actual,
                        mes_actual,
                        self.crear_formula_tipo_capex(fila_excel, header_map),
                        self.crear_formula_monto_ord(fila_excel, header_map),
                        self.crear_formula_monto_ext(fila_excel, header_map),
                        self.crear_formula_dia_pago(fila_excel, header_map),
                        datos_integrados['tienda'],
                        datos_integrados['ceco'],
                        datos_integrados['proyecto'],
                        area_calculada,    
                        datos_integrados['fecha_recibo'],
                        datos_integrados['descripcion'],
                        anio_fiscal_actual
                    ]

                    
                    # Verificar 27 valores (se agregó REAL CONVERTIDO, REAL MES CONVERTIDO)
                    if len(valores_calculados) != 27:
                        print(f"⚠️ Error: Se esperaban 27 valores calculados, pero hay {len(valores_calculados)}")
                        return False
                    
                    # Escribir columnas calculadas
                    for i, valor in enumerate(valores_calculados):
                        try:
                            # Manejar valores problemáticos
                            if pd.isna(valor):
                                valor = ""
                            ws.cell(row=fila_excel, column=23 + i, value=valor)
                        except Exception as cell_error:
                            print(f"❌ Error escribiendo celda [{fila_excel}, {24 + i}]: {cell_error}")
                            ws.cell(row=fila_excel, column=23 + i, value="ERROR")
                    
                except Exception as row_error:
                    print(f"❌ Error procesando fila {row_idx + 1}: {row_error}")
                    continue
            
            # Color verde para la hoja
            ws.sheet_properties.tabColor = "00FF00"
            
            # Autoajustar columnas con manejo de errores
            try:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            except Exception as adjust_error:
                print(f"⚠️ Error autoajustando columnas: {adjust_error}")
            
            # Guardar archivo
            wb.save(nombre_archivo)
            print(f"✅ Archivo creado: {nombre_archivo}")
            
            # Estadísticas finales
            total_filas = len(df)
            print(f"\n📊 ESTADÍSTICAS FINALES:")
            print(f"   📋 Total filas procesadas: {total_filas}")
            print(f"   🏪 TIENDA encontradas: {stats['tiendas_encontradas']} ({(stats['tiendas_encontradas']/total_filas*100):.1f}%)")
            print(f"   🏢 CECO encontrados: {stats['cecos_encontrados']} ({(stats['cecos_encontrados']/total_filas*100):.1f}%)")
            print(f"   📊 PROYECTO encontrados: {stats['proyectos_encontrados']} ({(stats['proyectos_encontrados']/total_filas*100):.1f}%)")
            print(f"   📅 FECHA RECIBO encontradas: {stats['fechas_recibo_encontradas']} ({(stats['fechas_recibo_encontradas']/total_filas*100):.1f}%)")
            print(f"   📝 DESCRIPCIÓN encontradas: {stats['descripciones_encontradas']} ({(stats['descripciones_encontradas']/total_filas*100):.1f}%)")
            print(f"   ❌ Facturas no encontradas: {stats['facturas_no_encontradas']} ({(stats['facturas_no_encontradas']/total_filas*100):.1f}%)")
            
            print(f"\n📍 MAPEO FINAL DE COLUMNAS:")
            print(f"   W  (22) - Solicitante (para AREA)")
            print(f"   J  (10) - Monto")
            print(f"   K  (11) - Moneda")
            print(f"   Y  (25) - MONTO A PAGAR CAPEX")
            print(f"   Z  (26) - MONEDA DE PAGO")
            print(f"   AA (27) - FECHA PAGO (del archivo de entrada)")
            print(f"   AB (28) - TC FTD (Tasa Farmatodo)")
            print(f"   AC (29) - TC BCV (Tasa BCV)")
            print(f"   AD (30) - CONVERSION VES")
            print(f"   AE (31) - CONVERSION TC FTD")
            print(f"   AF (32) - REAL CONVERTIDO ✨")
            print(f"   AG (33) - REAL MES CONVERTIDO ✨")
            print(f"   AR (43) - TIENDA_LOOKUP")
            print(f"   AS (44) - CECO")
            print(f"   AT (45) - PROYECTO")
            print(f"   AU (46) - AREA (desde Google Sheets)")
            print(f"   AV (47) - FECHA RECIBO")
            print(f"   AW (48) - DESCRIPCIÓN")
            print(f"   AX (49) - AÑO FISCAL (Agosto-Julio)")

            
            return True
            
        except Exception as e:
            print(f"❌ ERROR CRÍTICO creando archivo: {e}")
            import traceback
            traceback.print_exc()
            return False


def obtener_filas_a_saltar(archivo, max_filas_buscar=10):
    """Detectar automáticamente cuántas filas hay que saltar para encontrar los headers reales"""
    print(f"\n🔍 DETECTANDO HEADERS REALES...")
    print("-" * 40)
    
    try:
        for skip_rows in range(max_filas_buscar):
            print(f"🔍 Probando saltar {skip_rows} filas...")
            
            try:
                df_test = pd.read_excel(archivo, skiprows=skip_rows, nrows=1)
                columnas_leidas = [str(col).strip() for col in df_test.columns]
                
                print(f"   📋 Columnas encontradas: {len(columnas_leidas)}")
                
                primeras_3 = columnas_leidas[:3] if len(columnas_leidas) >= 3 else columnas_leidas
                print(f"   📝 Primeras columnas: {primeras_3}")
                
                unnamed_count = sum(1 for col in columnas_leidas if 'unnamed' in col.lower())
                print(f"   ❓ Columnas 'Unnamed': {unnamed_count}")
                
                if unnamed_count == 0:
                    columnas_criticas = ["Monto", "Moneda", "Proveedor"]
                    criticas_encontradas = 0
                    
                    for critica in columnas_criticas:
                        for col_leida in columnas_leidas:
                            if critica.lower() in col_leida.lower():
                                criticas_encontradas += 1
                                break
                    
                    print(f"   ✅ Columnas críticas encontradas: {criticas_encontradas}/3")
                    
                    if criticas_encontradas >= 2:
                        print(f"✅ HEADERS ENCONTRADOS en fila {skip_rows + 1}")
                        print(f"📋 Saltando {skip_rows} filas")
                        return skip_rows
                    else:
                        print(f"   ⚠️ Pocas columnas críticas, continuando búsqueda...")
                else:
                    print(f"   ❌ Aún contiene columnas 'Unnamed', continuando...")
                    
            except Exception as e:
                print(f"   ❌ Error leyendo con skip_rows={skip_rows}: {e}")
                continue
        
        print(f"⚠️ No se detectaron headers claros, usando fila 1 (skip_rows=0)")
        return 0
        
    except Exception as e:
        print(f"❌ Error en detección automática: {e}")
        return 0


def leer_excel_safe(archivo):
    """Leer archivo Excel de manera segura CON detección automática de headers"""
    try:
        if not Path(archivo).exists():
            print(f"❌ Archivo no encontrado: {archivo}")
            return None
            
        print(f"📖 Leyendo: {Path(archivo).name}")
        
        # Detectar filas a saltar automáticamente
        skip_rows = obtener_filas_a_saltar(archivo)
        
        # Leer con las filas correctas
        df = pd.read_excel(archivo, skiprows=skip_rows)
        
        if df.empty:
            print("❌ El archivo está vacío")
            return None
            
        print(f"✅ Leído: {len(df)} filas, {len(df.columns)} columnas")
        
        # DEBUG: Mostrar las columnas leídas
        print(f"\n🔍 DEBUG - COLUMNAS LEÍDAS DEL ARCHIVO:")
        print("-" * 50)
        for i, col in enumerate(df.columns, 1):
            col_clean = str(col).strip() if pd.notna(col) else "COLUMNA_VACÍA"
            print(f"  {i:2d}. [{len(col_clean):2d} chars] '{col_clean}'")
        
        # Verificar si aún hay columnas Unnamed después del ajuste
        unnamed_count = sum(1 for col in df.columns if 'unnamed' in str(col).lower())
        if unnamed_count > 0:
            print(f"\n⚠️ ADVERTENCIA: Aún hay {unnamed_count} columnas 'Unnamed'")
        else:
            print(f"✅ Headers detectados correctamente, sin columnas 'Unnamed'")
        
        return df
        
    except Exception as e:
        print(f"❌ Error leyendo archivo: {e}")
        return None

def validar_columnas_venezuela(df):
    """Validar estructura específica de Venezuela con DEBUG mejorado"""
    print("\n🔍 Validando estructura de Venezuela...")
    print("-" * 40)
    
    # ELIMINAR columna "Banco" si existe (no la necesitamos y desbarata el mapeo)
    if 'Banco' in df.columns:
        print(f"⚠️  Columna 'Banco' detectada. Eliminándola antes de la validación...")
        df = df.drop(columns=['Banco'])
        print(f"✅ Columna 'Banco' eliminada")
    
    columnas_esperadas = [
        "Numero de Factura", "Numero de OC", "Tipo Factura", "Nombre Lote",
        "Proveedor", "RIF", "Fecha Documento", "Tienda", "Sucursal",
        "Monto", "Moneda", "Fecha Vencimiento", "Cuenta", "Id Cta",
        "Método de Pago", "Pago Independiente", "Prioridad",
        "Monto CAPEX EXT", "Monto CAPEX ORD", "Monto CADM",
        "Fecha Creación", "Solicitante", "Proveedor Remito"
    ]
    
    columnas_archivo = [str(col).strip() for col in df.columns]
    
    print(f"📊 Columnas esperadas: {len(columnas_esperadas)}")
    print(f"📊 Columnas en archivo: {len(columnas_archivo)}")
    
    # DEBUG: Comparación detallada
    print(f"\n🔍 DEBUG - COMPARACIÓN DETALLADA:")
    print("-" * 50)
    print("ESPERADAS vs ARCHIVO:")
    
    max_len = max(len(columnas_esperadas), len(columnas_archivo))
    for i in range(max_len):
        esperada = columnas_esperadas[i] if i < len(columnas_esperadas) else "---"
        archivo = columnas_archivo[i] if i < len(columnas_archivo) else "---"
        
        if esperada == archivo:
            estado = "✅"
        elif esperada == "---":
            estado = "➕ EXTRA"  
        elif archivo == "---":
            estado = "❌ FALTA"
        else:
            estado = "🔄 DIFF"
            
        print(f"  {i+1:2d}. {estado} '{esperada}' vs '{archivo}'")
    
    # Verificar columnas críticas
    print(f"\n🔍 VERIFICANDO COLUMNAS CRÍTICAS:")
    print("-" * 40)
    
    columnas_criticas = ["Monto", "Moneda", "Proveedor"]
    faltantes = []
    encontradas = {}
    
    for col_critica in columnas_criticas:
        encontrada = None
        
        if col_critica in columnas_archivo:
            encontrada = col_critica
        else:
            for col_archivo in columnas_archivo:
                if col_critica.lower().replace(" ", "") in col_archivo.lower().replace(" ", ""):
                    encontrada = col_archivo
                    break
        
        if encontrada:
            encontradas[col_critica] = encontrada
            pos = columnas_archivo.index(encontrada) + 1
            letra = chr(64 + pos)
            print(f"  ✅ {col_critica} → '{encontrada}' (pos {pos}, col {letra})")
        else:
            faltantes.append(col_critica)
            print(f"  ❌ {col_critica} → NO ENCONTRADA")
    
    if faltantes:
        print(f"\n❌ Columnas críticas faltantes: {faltantes}")
        return False
    
    # Verificar posiciones clave
    try:
        col_monto = encontradas.get("Monto", "Monto")
        col_moneda = encontradas.get("Moneda", "Moneda") 
        col_proveedor = encontradas.get("Proveedor", "Proveedor")
        
        pos_monto = columnas_archivo.index(col_monto) + 1
        pos_moneda = columnas_archivo.index(col_moneda) + 1
        pos_proveedor = columnas_archivo.index(col_proveedor) + 1
        
        print(f"\n📍 POSICIONES CONFIRMADAS:")
        print(f"  Monto: columna {pos_monto} ({chr(64 + pos_monto)})")
        print(f"  Moneda: columna {pos_moneda} ({chr(64 + pos_moneda)})")  
        print(f"  Proveedor: columna {pos_proveedor} ({chr(64 + pos_proveedor)})")
        
    except ValueError as e:
        print(f"❌ Error ubicando columnas: {e}")
        return False
    
    print(f"\n✅ Validación de estructura EXITOSA")
    return True

def validar_monedas_venezuela(df):
    """Validar monedas específicas de Venezuela con DEBUG"""
    print(f"\n💰 VALIDANDO MONEDAS...")
    print("-" * 30)
    
    col_moneda = None
    for col in df.columns:
        if "moneda" in str(col).lower():
            col_moneda = col
            break
    
    if not col_moneda:
        print("⚠️ Columna de moneda no encontrada")
        return True
    
    print(f"📍 Usando columna: '{col_moneda}'")
    
    monedas_validas = ['VES', 'USD', 'EUR', 'VEF']
    monedas_archivo = df[col_moneda].dropna().unique()
    monedas_invalidas = set(monedas_archivo) - set(monedas_validas)
    
    if monedas_invalidas:
        print(f"⚠️ Monedas no estándar encontradas: {monedas_invalidas}")
    
    print(f"💰 Monedas en el archivo: {list(monedas_archivo)}")
    
    conteo_monedas = df[col_moneda].value_counts()
    print("📊 Distribución de monedas:")
    for moneda, cantidad in conteo_monedas.items():
        print(f"   {moneda}: {cantidad} registros")
    
    return True


def validar_columnas_colombia(df):
    """Validar estructura específica de Venezuela con DEBUG mejorado"""
    print("\n🔍 Validando estructura de Venezuela...")
    print("-" * 40)
    
    # ELIMINAR columna "Banco" si existe (no la necesitamos y desbarata el mapeo)
    if 'Banco' in df.columns:
        print(f"⚠️  Columna 'Banco' detectada. Eliminándola antes de la validación...")
        df = df.drop(columns=['Banco'])
        print(f"✅ Columna 'Banco' eliminada")
    
    columnas_esperadas = [
        "Numero de Factura", "Numero de OC", "Tipo Factura", "Nombre Lote",
        "Proveedor", "RIF", "Fecha Documento", "Tienda", "Sucursal",
        "Monto", "Moneda", "Fecha Vencimiento", "Cuenta", "Id Cta",
        "Método de Pago", "Pago Independiente", "Prioridad",
        "Monto CAPEX EXT", "Monto CAPEX ORD", "Monto CADM",
        "Fecha Creación", "Solicitante", "Proveedor Remito"
    ]
    
    columnas_archivo = [str(col).strip() for col in df.columns]
    
    print(f"📊 Columnas esperadas: {len(columnas_esperadas)}")
    print(f"📊 Columnas en archivo: {len(columnas_archivo)}")
    
    # DEBUG: Comparación detallada
    print(f"\n🔍 DEBUG - COMPARACIÓN DETALLADA:")
    print("-" * 50)
    print("ESPERADAS vs ARCHIVO:")
    
    max_len = max(len(columnas_esperadas), len(columnas_archivo))
    for i in range(max_len):
        esperada = columnas_esperadas[i] if i < len(columnas_esperadas) else "---"
        archivo = columnas_archivo[i] if i < len(columnas_archivo) else "---"
        
        if esperada == archivo:
            estado = "✅"
        elif esperada == "---":
            estado = "➕ EXTRA"  
        elif archivo == "---":
            estado = "❌ FALTA"
        else:
            estado = "🔄 DIFF"
            
        print(f"  {i+1:2d}. {estado} '{esperada}' vs '{archivo}'")
    
    # Verificar columnas críticas
    print(f"\n🔍 VERIFICANDO COLUMNAS CRÍTICAS:")
    print("-" * 40)
    
    columnas_criticas = ["Monto", "Moneda", "Proveedor"]
    faltantes = []
    encontradas = {}
    
    for col_critica in columnas_criticas:
        encontrada = None
        
        if col_critica in columnas_archivo:
            encontrada = col_critica
        else:
            for col_archivo in columnas_archivo:
                if col_critica.lower().replace(" ", "") in col_archivo.lower().replace(" ", ""):
                    encontrada = col_archivo
                    break
        
        if encontrada:
            encontradas[col_critica] = encontrada
            pos = columnas_archivo.index(encontrada) + 1
            letra = chr(64 + pos)
            print(f"  ✅ {col_critica} → '{encontrada}' (pos {pos}, col {letra})")
        else:
            faltantes.append(col_critica)
            print(f"  ❌ {col_critica} → NO ENCONTRADA")
    
    if faltantes:
        print(f"\n❌ Columnas críticas faltantes: {faltantes}")
        return False
    
    # Verificar posiciones clave
    try:
        col_monto = encontradas.get("Monto", "Monto")
        col_moneda = encontradas.get("Moneda", "Moneda") 
        col_proveedor = encontradas.get("Proveedor", "Proveedor")
        
        pos_monto = columnas_archivo.index(col_monto) + 1
        pos_moneda = columnas_archivo.index(col_moneda) + 1
        pos_proveedor = columnas_archivo.index(col_proveedor) + 1
        
        print(f"\n📍 POSICIONES CONFIRMADAS:")
        print(f"  Monto: columna {pos_monto} ({chr(64 + pos_monto)})")
        print(f"  Moneda: columna {pos_moneda} ({chr(64 + pos_moneda)})")  
        print(f"  Proveedor: columna {pos_proveedor} ({chr(64 + pos_proveedor)})")
        
    except ValueError as e:
        print(f"❌ Error ubicando columnas: {e}")
        return False
    
    print(f"\n✅ Validación de estructura EXITOSA")
    return True

def validar_monedas_colombia(df):
    """Validar monedas específicas de Venezuela con DEBUG"""
    print(f"\n💰 VALIDANDO MONEDAS...")
    print("-" * 30)
    
    col_moneda = None
    for col in df.columns:
        if "moneda" in str(col).lower():
            col_moneda = col
            break
    
    if not col_moneda:
        print("⚠️ Columna de moneda no encontrada")
        return True
    
    print(f"📍 Usando columna: '{col_moneda}'")
    
    monedas_validas = ['COP', 'USD', 'EUR']
    monedas_archivo = df[col_moneda].dropna().unique()
    monedas_invalidas = set(monedas_archivo) - set(monedas_validas)
    
    if monedas_invalidas:
        print(f"⚠️ Monedas no estándar encontradas: {monedas_invalidas}")
    
    print(f"💰 Monedas en el archivo: {list(monedas_archivo)}")
    
    conteo_monedas = df[col_moneda].value_counts()
    print("📊 Distribución de monedas:")
    for moneda, cantidad in conteo_monedas.items():
        print(f"   {moneda}: {cantidad} registros")
    
    return True

def validar_reporte_absoluto(archivo_reporte_absoluto):
    """Validar la estructura del Reporte Absoluto"""
    if not archivo_reporte_absoluto:
        print("ℹ️ No se proporcionó Reporte Absoluto")
        return False
        
    try:
        print(f"\n🔍 VALIDANDO REPORTE ABSOLUTO...")
        print("-" * 40)
        
        if not Path(archivo_reporte_absoluto).exists():
            print(f"❌ Archivo Reporte Absoluto no encontrado: {archivo_reporte_absoluto}")
            return False
        
        df_absoluto = pd.read_excel(archivo_reporte_absoluto)
        
        print(f"✅ Reporte Absoluto leído: {len(df_absoluto)} filas, {len(df_absoluto.columns)} columnas")
        
        columnas_archivo = [str(col).strip() for col in df_absoluto.columns]
        
        print(f"🔍 Primeras 5 columnas del Reporte Absoluto:")
        for i, col in enumerate(columnas_archivo[:5], 1):
            print(f"  {i:2d}. {col}")
        
        # Buscar las columnas objetivo (ahora 5)
        columnas_objetivo = ["Cta. Cargo Centro Desc.", "Cta. Cargo Centro", "Cta. Cargo", "Fecha Recepción", "Descipción"]
        for objetivo in columnas_objetivo:
            encontrada = False
            for i, col in enumerate(columnas_archivo, 1):
                # Ajustar búsqueda para términos específicos
                if objetivo == "Fecha Recepción":
                    if all(palabra in col.lower() for palabra in ['fecha', 'recepción']) or all(palabra in col.lower() for palabra in ['fecha', 'recepcion']):
                        print(f"✅ Columna '{objetivo}' encontrada: '{col}' (posición {i})")
                        encontrada = True
                        break
                elif objetivo == "Descipción":
                    if 'descipción' in col.lower() or 'descripcion' in col.lower() or 'descripción' in col.lower():
                        print(f"✅ Columna '{objetivo}' encontrada: '{col}' (posición {i})")
                        encontrada = True
                        break
                else:
                    palabras_objetivo = objetivo.lower().split()
                    if all(palabra in col.lower() for palabra in palabras_objetivo):
                        print(f"✅ Columna '{objetivo}' encontrada: '{col}' (posición {i})")
                        encontrada = True
                        break
            
            if not encontrada:
                print(f"⚠️ Columna '{objetivo}' no encontrada")
        
        # Verificar si hay facturas para hacer match
        facturas_sample = df_absoluto.iloc[:3, 0].tolist()
        print(f"💼 Facturas ejemplo: {facturas_sample}")
        
        return True
        
    except Exception as e:
        print(f"❌ Error validando Reporte Absoluto: {e}")
        return False

def analizar_estructura_archivo(archivo, max_filas_buscar=10):
    """Función de diagnóstico completo para analizar la estructura del archivo"""
    print(f"\n🔬 ANÁLISIS COMPLETO DE ESTRUCTURA")
    print("=" * 60)
    
    try:
        print("📖 Leyendo primeras 10 filas sin procesar...")
        
        for fila in range(10):
            try:
                df_raw = pd.read_excel(archivo, skiprows=fila, nrows=1, header=None)
                
                if not df_raw.empty:
                    valores = df_raw.iloc[0].tolist()
                    valores_clean = [str(v)[:30] + "..." if len(str(v)) > 30 else str(v) 
                                   for v in valores[:8]]
                    
                    print(f"  Fila {fila+1:2d}: {valores_clean}")
                    
                    texto_count = sum(1 for v in valores if isinstance(v, str) and len(str(v)) > 5)
                    if texto_count >= 5:
                        print(f"         👆 Posible fila de headers ({texto_count} textos largos)")
                        
            except Exception as e:
                print(f"  Fila {fila+1:2d}: Error - {e}")
                
        skip_detectado = obtener_filas_a_saltar(archivo)
        print(f"\n🎯 RECOMENDACIÓN: Usar skiprows={skip_detectado}")
        
        print(f"\n📋 RESULTADO CON DETECCIÓN AUTOMÁTICA:")
        df_final = pd.read_excel(archivo, skiprows=skip_detectado)
        
        print(f"   Shape: {df_final.shape}")
        print(f"   Columnas con 'Unnamed': {sum(1 for col in df_final.columns if 'unnamed' in str(col).lower())}")
        
        if len(df_final.columns) >= 3:
            print(f"   Primeras columnas: {list(df_final.columns[:3])}")
            
        return skip_detectado
        
    except Exception as e:
        print(f"❌ Error en análisis: {e}")
        return 0
